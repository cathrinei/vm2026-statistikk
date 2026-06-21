#!/usr/bin/env python3
"""
add_heatmap_sheet.py
Lager arket "Heatmap" med ett minutt per celle (1–90 + tilleggstid),
fargekodet etter antall mål. Kilde: lagstatistikk_cache.json.
"""
import io, json, shutil, sys
from dataclasses import dataclass, field
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
BASE_DIR = Path(__file__).parent

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    sys.exit(f"Mangler pakke: {e}")

EXCEL_PATH  = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
STAT_CACHE  = BASE_DIR / "lagstatistikk_cache.json"
SHEET_NAME  = "Heatmap"

GOAL_TYPES  = {"Goal!", "Penalty Goal", "Own goal"}

# Fargegradient: 0 mål → hvit, 7 mål → navy
GRADIENT = [
    ("FFFFFF", "AAAAAA"),  # 0 — hvit, grå tekst
    ("DEEBF7", "1A1A2E"),  # 1 — svært lys blå
    ("BDD7EE", "1A1A2E"),  # 2 — lys blå
    ("5B9BD5", "FFFFFF"),  # 3 — mellomblå
    ("2171B5", "FFFFFF"),  # 4 — sterk blå
    ("1A3C6B", "FFFFFF"),  # 5 — mørk blå (designsystem)
    ("0F2044", "FFFFFF"),  # 6+ — navy (designsystem)
]

def _colors(count: int) -> tuple[str, str]:
    idx = min(count, len(GRADIENT) - 1)
    return GRADIENT[idx]


def _parse_min(s: str) -> tuple[int, int] | None:
    if not s:
        return None
    s = str(s).strip().rstrip("'")
    if "'+" in s:
        parts = s.replace("'", "").split("+")
        try:
            return int(parts[0]), int(parts[1])
        except Exception:
            return None
    try:
        return int(s), 0
    except Exception:
        return None


@dataclass
class GoalCounts:
    regular: dict[int, int] = field(default_factory=dict)  # ekte min 1–90
    ht_et:   dict[int, int] = field(default_factory=dict)  # 45+N, nøkkel=N
    ft_et:   dict[int, int] = field(default_factory=dict)  # 90+N, nøkkel=N
    et1:     dict[int, int] = field(default_factory=dict)  # ekstraomg. 1: abs. min 91–105
    et1_et:  dict[int, int] = field(default_factory=dict)  # 105+N, nøkkel=N
    et2:     dict[int, int] = field(default_factory=dict)  # ekstraomg. 2: abs. min 106–120
    et2_et:  dict[int, int] = field(default_factory=dict)  # 120+N, nøkkel=N


def tell_per_minutt() -> GoalCounts:
    if not Path(STAT_CACHE).exists():
        sys.exit(f"FEIL: {STAT_CACHE} finnes ikke.")
    with open(STAT_CACHE, encoding="utf-8") as f:
        cache = json.load(f)
    result = GoalCounts()
    for mid, events in cache.items():
        for e in events:
            if e.get("TypeLocalized") not in GOAL_TYPES:
                continue
            parsed = _parse_min(e.get("MatchMinute"))
            if parsed is None:
                continue
            base, extra = parsed
            if extra == 0 and 1 <= base <= 90:
                result.regular[base] = result.regular.get(base, 0) + 1
            elif base == 45 and extra >= 1:
                result.ht_et[extra] = result.ht_et.get(extra, 0) + 1
            elif base == 90 and extra >= 1:
                result.ft_et[extra] = result.ft_et.get(extra, 0) + 1
            elif extra == 0 and 91 <= base <= 105:
                result.et1[base] = result.et1.get(base, 0) + 1
            elif base == 105 and extra >= 1:
                result.et1_et[extra] = result.et1_et.get(extra, 0) + 1
            elif extra == 0 and 106 <= base <= 120:
                result.et2[base] = result.et2.get(base, 0) + 1
            elif base == 120 and extra >= 1:
                result.et2_et[extra] = result.et2_et.get(extra, 0) + 1
    return result


def skriv_ark(counts: GoalCounts) -> None:
    backup = Path(str(EXCEL_PATH) + ".bak")
    shutil.copy2(EXCEL_PATH, backup)
    wb = load_workbook(EXCEL_PATH)

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    insert_after = "Scoringstidspunkt"
    if insert_after in wb.sheetnames:
        idx = wb.sheetnames.index(insert_after) + 1
    elif "Lagstatistikk" in wb.sheetnames:
        idx = wb.sheetnames.index("Lagstatistikk") + 1
    else:
        idx = len(wb.sheetnames)
    ws = wb.create_sheet(SHEET_NAME, idx)

    NAVY  = PatternFill("solid", fgColor="0F2044")
    BLUE  = PatternFill("solid", fgColor="1A3C6B")
    GRAY  = PatternFill("solid", fgColor="D9D9D9")
    thin  = Side(style="thin",   color="E2E8F0")
    thck  = Side(style="medium", color="0F2044")
    bot   = Border(bottom=thin)
    ctr   = Alignment(horizontal="center", vertical="center")
    lft   = Alignment(horizontal="left",   vertical="center")

    f_title = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    f_hdr   = Font(name="Calibri", bold=True, size=9,  color="FFFFFF")
    f_sub   = Font(name="Calibri", bold=True, size=8,  color="FFFFFF")
    f_gray  = Font(name="Calibri", size=8, color="888888")

    CELL_W = 4.5   # bredde per minutt-celle
    CELL_H = 20    # høyde per rad

    # ── Rad 1: Tittel ─────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1, value="Scoringstidspunkt — Heatmap (mål per minutt)")
    c.font = f_title; c.fill = NAVY; c.alignment = lft

    # ── Rad 2: Kolonneoverskrifter (siffer: +1 ... +10) ───────────────────────
    ws.row_dimensions[2].height = 16
    ws.cell(row=2, column=1, value="")
    c = ws.cell(row=2, column=1)
    c.fill = NAVY

    for col_i in range(10):
        col = col_i + 2   # B=2 ... K=11
        c = ws.cell(row=2, column=col, value=col_i + 1)
        c.font = f_hdr; c.fill = BLUE; c.alignment = ctr

    # ── Rader 3–11: minutter 1–90 (9 rader × 10 kolonner) ────────────────────
    for row_i in range(9):
        row = row_i + 3
        ws.row_dimensions[row].height = CELL_H
        row_start = row_i * 10 + 1   # 1, 11, 21, ..., 81
        row_end   = row_start + 9    # 10, 20, ..., 90
        label = f"{row_start}–{row_end}"

        # Radlabel (kolonne A)
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = f_hdr; lc.fill = NAVY; lc.alignment = ctr

        # Halvtids-separator: tykk kant under rad for minutt 41-50 (etter min 45)
        # Markeres med tykk kant under celle 41-50-raden (row_i=4, minutt 41-50)
        # Minutt 45 er col_i=4 i row_i=4
        for col_i in range(10):
            col    = col_i + 2
            minutt = row_start + col_i
            n      = counts.regular.get(minutt, 0)
            bg, fg = _colors(n)

            cell = ws.cell(row=row, column=col, value=n if n > 0 else None)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.font      = Font(name="Calibri", bold=(n > 0), size=10, color=fg)
            cell.alignment = ctr

            # Halvtidsmarkering: tykk kant til høyre for minutt 45 og til venstre for 46
            left_side  = thck if minutt == 46 else thin
            right_side = thck if minutt == 45 else thin
            top_side   = thin
            bot_side   = thin
            cell.border = Border(left=left_side, right=right_side,
                                 top=top_side, bottom=bot_side)

    # ── Rad 12: separator (før 45+ tilleggstid) ───────────────────────────────
    ws.row_dimensions[12].height = 8
    for col in range(1, 12):
        c = ws.cell(row=12, column=col)
        c.fill = GRAY

    # ── Rad 13: 45+ tilleggstid-header ────────────────────────────────────────
    ws.row_dimensions[13].height = 16
    lc = ws.cell(row=13, column=1, value="45+")
    lc.font = f_hdr; lc.fill = NAVY; lc.alignment = ctr

    for col_i in range(10):
        col = col_i + 2
        c = ws.cell(row=13, column=col, value=f"+{col_i + 1}")
        c.font = f_sub; c.fill = BLUE; c.alignment = ctr

    # ── Rad 14: 45+ tilleggstid-data (45'+1 … 45'+10) ────────────────────────
    ws.row_dimensions[14].height = CELL_H
    lc = ws.cell(row=14, column=1, value="HT")
    lc.font = f_hdr; lc.fill = NAVY; lc.alignment = ctr

    for col_i in range(10):
        col       = col_i + 2
        extra_min = col_i + 1
        n         = counts.ht_et.get(extra_min, 0)
        bg, fg    = _colors(n)
        cell = ws.cell(row=14, column=col, value=n if n > 0 else None)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.font      = Font(name="Calibri", bold=(n > 0), size=10, color=fg)
        cell.alignment = ctr
        cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Rad 15: separator (før ET/90+ tilleggstid) ────────────────────────────
    ws.row_dimensions[15].height = 8
    for col in range(1, 12):
        c = ws.cell(row=15, column=col)
        c.fill = GRAY

    # ── Rad 16: AT (Added Time) tilleggstid-header ────────────────────────────
    ws.row_dimensions[16].height = 16
    lc = ws.cell(row=16, column=1, value="AT")
    lc.font = f_hdr; lc.fill = NAVY; lc.alignment = ctr

    for col_i in range(10):
        col = col_i + 2
        c = ws.cell(row=16, column=col, value=f"+{col_i + 1}")
        c.font = f_sub; c.fill = BLUE; c.alignment = ctr

    # ── Rad 17: 90+ tilleggstid-data (90'+1 … 90'+10) ────────────────────────
    ws.row_dimensions[17].height = CELL_H
    lc = ws.cell(row=17, column=1, value="90+")
    lc.font = f_hdr; lc.fill = NAVY; lc.alignment = ctr

    for col_i in range(10):
        col       = col_i + 2
        extra_min = col_i + 1
        n         = counts.ft_et.get(extra_min, 0)
        bg, fg    = _colors(n)
        cell = ws.cell(row=17, column=col, value=n if n > 0 else None)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.font      = Font(name="Calibri", bold=(n > 0), size=10, color=fg)
        cell.alignment = ctr
        cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Rad 18: separator (før ekstraomganger) ────────────────────────────────
    ws.row_dimensions[18].height = 8
    for col in range(1, 12):
        ws.cell(row=18, column=col).fill = GRAY

    # ── Rad 19: EKSTRAOMGANGER-tittel ─────────────────────────────────────────
    ws.row_dimensions[19].height = 20
    ws.merge_cells("A19:K19")
    c = ws.cell(row=19, column=1, value="EKSTRAOMGANGER (sluttspill)")
    c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill = NAVY; c.alignment = lft

    # ── Hjelpefunksjon: skriv ET-periode ──────────────────────────────────────
    def _et_periode(first_row: int, label: str, min_start: int,
                    et_data: dict[int, int], et_added: dict[int, int]) -> None:
        """Skriver én ekstraomgangsperiode: to datarader + tilleggstid."""
        LGRAY = PatternFill("solid", fgColor="F0F0F0")
        # Header-rad: absolutte minutter som kolonneoverskrifter
        ws.row_dimensions[first_row].height = 16
        lc = ws.cell(row=first_row, column=1, value=label)
        lc.font = f_hdr; lc.fill = NAVY; lc.alignment = ctr
        for col_i in range(10):
            c = ws.cell(row=first_row, column=col_i + 2, value=min_start + col_i)
            c.font = f_sub; c.fill = BLUE; c.alignment = ctr

        # Rad A: første 10 minutter av perioden
        ws.row_dimensions[first_row + 1].height = CELL_H
        lc = ws.cell(row=first_row + 1, column=1, value=f"{min_start}–{min_start+9}")
        lc.font = f_hdr; lc.fill = NAVY; lc.alignment = ctr
        for col_i in range(10):
            minutt = min_start + col_i
            n      = et_data.get(minutt, 0)
            bg, fg = _colors(n)
            cell = ws.cell(row=first_row + 1, column=col_i + 2, value=n if n > 0 else None)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name="Calibri", bold=(n > 0), size=10, color=fg)
            cell.alignment = ctr
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Rad B: siste 5 minutter (101-105 / 116-120), resten grå
        ws.row_dimensions[first_row + 2].height = CELL_H
        lc = ws.cell(row=first_row + 2, column=1, value=f"{min_start+10}–{min_start+14}")
        lc.font = f_hdr; lc.fill = NAVY; lc.alignment = ctr
        for col_i in range(10):
            minutt = min_start + 10 + col_i
            if col_i < 5:
                n      = et_data.get(minutt, 0)
                bg, fg = _colors(n)
                cell = ws.cell(row=first_row + 2, column=col_i + 2, value=n if n > 0 else None)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Calibri", bold=(n > 0), size=10, color=fg)
            else:
                cell = ws.cell(row=first_row + 2, column=col_i + 2)
                cell.fill = LGRAY
                cell.font = Font(name="Calibri", size=10, color="CCCCCC")
            cell.alignment = ctr
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Tilleggstid-header
        ws.row_dimensions[first_row + 3].height = 16
        lc = ws.cell(row=first_row + 3, column=1, value=f"{min_start+14}+")
        lc.font = f_hdr; lc.fill = NAVY; lc.alignment = ctr
        for col_i in range(10):
            c = ws.cell(row=first_row + 3, column=col_i + 2, value=f"+{col_i + 1}")
            c.font = f_sub; c.fill = BLUE; c.alignment = ctr

        # Tilleggstid-data
        ws.row_dimensions[first_row + 4].height = CELL_H
        lc = ws.cell(row=first_row + 4, column=1, value=f"{min_start+14}+")
        lc.font = f_hdr; lc.fill = NAVY; lc.alignment = ctr
        for col_i in range(10):
            n      = et_added.get(col_i + 1, 0)
            bg, fg = _colors(n)
            cell = ws.cell(row=first_row + 4, column=col_i + 2, value=n if n > 0 else None)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name="Calibri", bold=(n > 0), size=10, color=fg)
            cell.alignment = ctr
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Rader 20–24: Ekstraomgang 1 (min 91–105 + 105+) ─────────────────────
    _et_periode(20, "ET 1", 91, counts.et1, counts.et1_et)

    # ── Rad 25: separator mellom ET-periodene ─────────────────────────────────
    ws.row_dimensions[25].height = 8
    for col in range(1, 12):
        ws.cell(row=25, column=col).fill = GRAY

    # ── Rader 26–30: Ekstraomgang 2 (min 106–120 + 120+) ────────────────────
    _et_periode(26, "ET 2", 106, counts.et2, counts.et2_et)

    # ── Rad 32–33: Forklaring ─────────────────────────────────────────────────
    ws.row_dimensions[32].height = 14
    ws.cell(row=32, column=1, value="Fargeskala:").font = Font(name="Calibri", bold=True, size=9, color="1A1A2E")

    legend_items = [
        ("0 mål",  "FFFFFF", "AAAAAA"),
        ("1 mål",  "DEEBF7", "1A1A2E"),
        ("2 mål",  "BDD7EE", "1A1A2E"),
        ("3 mål",  "5B9BD5", "FFFFFF"),
        ("4 mål",  "2171B5", "FFFFFF"),
        ("5 mål",  "1A3C6B", "FFFFFF"),
        ("6+ mål", "0F2044", "FFFFFF"),
    ]
    ws.row_dimensions[33].height = 18
    for i, (label, bg, fg) in enumerate(legend_items):
        col = i + 2
        c = ws.cell(row=33, column=col, value=label)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(name="Calibri", size=8, color=fg)
        c.alignment = ctr
        c.border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Halvtidsmarkeringsforklaring ───────────────────────────────────────────
    ws.row_dimensions[35].height = 14
    note = ws.cell(row=35, column=1,
                   value="| Tykk strek markerer halvtidspausen (mellom minutt 45 og 46)")
    note.font = Font(name="Calibri", italic=True, size=8, color="6B7A99")

    # ── Kolonnebredder ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["A"].customWidth = True
    for col_i in range(10):
        letter = chr(ord("B") + col_i)
        ws.column_dimensions[letter].width = CELL_W
        ws.column_dimensions[letter].customWidth = True
    ws.column_dimensions["L"].width = 3  # liten buffer-kolonne
    ws.column_dimensions["L"].customWidth = True

    try:
        wb.save(EXCEL_PATH)
    except Exception as e:
        shutil.copy2(backup, EXCEL_PATH)
        sys.exit(f"FEIL — rullet tilbake: {e}")


def main():
    print("=" * 55)
    print("  VM 2026 — Scoringstidspunkt Heatmap")
    print("=" * 55)

    print("\n[1/2] Teller mål per minutt fra cache...")
    counts = tell_per_minutt()
    totalt_regular = sum(counts.regular.values())
    totalt_ht_et   = sum(counts.ht_et.values())
    totalt_ft_et   = sum(counts.ft_et.values())
    totalt_et1     = sum(counts.et1.values()) + sum(counts.et1_et.values())
    totalt_et2     = sum(counts.et2.values()) + sum(counts.et2_et.values())
    totalt         = totalt_regular + totalt_ht_et + totalt_ft_et + totalt_et1 + totalt_et2
    print(f"  {totalt} mål totalt")
    print(f"    - {totalt_regular} i ordinær tid (minutt 1–90)")
    print(f"    - {totalt_ht_et} i HT tilleggstid (45'+N): {dict(sorted(counts.ht_et.items()))}")
    print(f"    - {totalt_ft_et} i FT tilleggstid (90'+N): {dict(sorted(counts.ft_et.items()))}")
    if totalt_et1:
        print(f"    - {totalt_et1} i ekstraomgang 1 (91–105 + 105'+N): {dict(sorted(counts.et1.items()))}")
    if totalt_et2:
        print(f"    - {totalt_et2} i ekstraomgang 2 (106–120 + 120'+N): {dict(sorted(counts.et2.items()))}")
    if counts.regular:
        maks_min = max(counts.regular, key=counts.regular.get)
        print(f"  Maks i ett minutt (ordinær tid): {counts.regular[maks_min]} mål (minutt {maks_min})")

    print(f"\n[2/2] Skriver ark '{SHEET_NAME}' til Excel...")
    skriv_ark(counts)
    print("\nFerdig.")


if __name__ == "__main__":
    main()
