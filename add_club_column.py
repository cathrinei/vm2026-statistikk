#!/usr/bin/env python3
"""
add_club_column.py
Setter inn "Klubb"-kolonne (N=14) mellom M (Navn) og O (Fødselsdato) i alle 12 gruppeark.

Kjøremodus:
  1. Første gang  → insert_cols(14) + fyll inn klubber
  2. Påfølgende   → kolonnen finnes allerede → bare fyll inn tomme celler

Kilde: clubs_new.json (spillernavn → klubb).
"""
import io, json, shutil, sys, unicodedata, re
from pathlib import Path
BASE_DIR = Path(__file__).parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as gcl
except ImportError:
    sys.exit("Mangler openpyxl: pip install openpyxl")

EXCEL_PATH  = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
CLUBS_JSON  = BASE_DIR / "clubs_new.json"

GROUPS      = [f"Gruppe {x}" for x in "ABCDEFGHIJKL"]
BLOCKS      = [(17, 19, 44), (46, 48, 73), (75, 77, 102), (104, 106, 131)]
COL_KLUBB   = 14   # N
KLUBB_WIDTH = 22


# ── Stil ──────────────────────────────────────────────────────────────────────

BLUE = PatternFill("solid", fgColor="1A3C6B")
EVEN = PatternFill("solid", fgColor="F0F5FB")
ODD  = PatternFill("solid", fgColor="FFFFFF")

bot = Border(bottom=Side(style="thin", color="E2E8F0"))
LFT = Alignment(horizontal="left",   vertical="center")
CTR = Alignment(horizontal="center", vertical="center")

F_HDR  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
F_CLUB = Font(name="Calibri", size=10, color="1A1A2E")


# ── Normalisering for navneoppslag ────────────────────────────────────────────

def _norm(s):
    s = (s or "").lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[-'' ]", "", s)
    return s


# ── Les klubbdata fra clubs_new.json ─────────────────────────────────────────

with open(CLUBS_JSON, encoding="utf-8") as f:
    clubs_raw = json.load(f)

name_to_club: dict[str, str] = {}
for lag_dict in clubs_raw.values():
    for player_name, club in lag_dict.items():
        if player_name and club:
            name_to_club[_norm(player_name)] = club

print(f"clubs_new.json: {len(name_to_club)} spillere med klubbverdi")

# Spillere med identisk normalisert navn i ulike nasjonallag
NAVNEKOLLISJONER: dict[str, dict[str, str]] = {
    _norm("Emiliano Martínez"): {
        _norm("Argentina"): "Aston Villa",
        _norm("Uruguay"):   "Palmeiras",
    },
}


# ── Åpne Excel ────────────────────────────────────────────────────────────────

FORVENTET = 1248

backup = Path(str(EXCEL_PATH) + ".bak")
shutil.copy2(EXCEL_PATH, backup)
print(f"Backup: {backup}")

wb = load_workbook(EXCEL_PATH)
tot_inserted = tot_updated = tot_skipped = 0

for gruppe in GROUPS:
    if gruppe not in wb.sheetnames:
        print(f"  {gruppe}: ARK MANGLER — hopper over")
        continue
    ws = wb[gruppe]

    # Sjekk om kolonnen allerede er satt inn (header-label på h_row+1)
    existing_label = ws.cell(row=BLOCKS[0][0] + 1, column=COL_KLUBB).value
    already_inserted = (existing_label == "Klubb")

    if already_inserted:
        mode = "oppdater"
    else:
        ws.insert_cols(COL_KLUBB)
        mode = "insert"

    inserted = updated = skipped = 0

    for header_row, start_row, end_row in BLOCKS:
        hdr_col_row = header_row + 1

        if not already_inserted:
            c = ws.cell(row=hdr_col_row, column=COL_KLUBB, value="Klubb")
            c.font      = F_HDR
            c.fill      = BLUE
            c.alignment = LFT

        lagnavn_norm = _norm(ws.cell(row=header_row, column=13).value or "")

        for row in range(start_row, end_row + 1):
            name_val  = ws.cell(row=row, column=13).value   # M = Navn
            club_cell = ws.cell(row=row, column=COL_KLUBB)

            if already_inserted and club_cell.value:
                skipped += 1
                continue

            navn_norm = _norm(name_val or "")
            if navn_norm in NAVNEKOLLISJONER:
                club = NAVNEKOLLISJONER[navn_norm].get(lagnavn_norm, "")
            else:
                club = name_to_club.get(navn_norm, "")
            if not club:
                continue

            # Bakgrunnsfarge fra Navn-nabocellen (M=13)
            try:
                rgb = ws.cell(row=row, column=13).fill.fgColor.rgb
                bg  = PatternFill("solid", fgColor=rgb) if rgb and rgb != "00000000" \
                      else (EVEN if (row - start_row) % 2 == 0 else ODD)
            except Exception:
                bg = EVEN if (row - start_row) % 2 == 0 else ODD

            club_cell.value     = club
            club_cell.fill      = bg
            club_cell.font      = F_CLUB
            club_cell.alignment = LFT
            club_cell.border    = bot

            if already_inserted:
                updated += 1
            else:
                inserted += 1

    if not already_inserted:
        ws.column_dimensions["N"].width = KLUBB_WIDTH

    tot_inserted += inserted
    tot_updated  += updated
    tot_skipped  += skipped
    if mode == "insert":
        print(f"  {gruppe} [ny kolonne]: {inserted} klubber lagt inn")
    else:
        print(f"  {gruppe} [oppdater]:   {updated} nye klubber, {skipped} allerede fylt")

try:
    wb.save(EXCEL_PATH)
    total_fylt = tot_inserted + tot_updated + tot_skipped
    sjekk = "✓" if total_fylt == FORVENTET else f"ADVARSEL: forventet {FORVENTET}"
    print(f"\nLagret: {EXCEL_PATH}")
    print(f"Totalt: {total_fylt} spillere ({sjekk})  —  "
          f"lagt inn: {tot_inserted + tot_updated}, allerede fylt: {tot_skipped}")
except Exception as e:
    shutil.copy2(backup, EXCEL_PATH)
    sys.exit(f"FEIL ved lagring — rullet tilbake fra backup: {e}")
