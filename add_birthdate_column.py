#!/usr/bin/env python3
"""
add_birthdate_column.py
Setter inn "Fødselsdato"-kolonne (O=15) mellom N (Klubb) og Pos i alle 12 gruppeark.

Kjøremodus:
  1. Første gang  → insert_cols(15) + fyll inn datoer
  2. Påfølgende   → kolonnen finnes allerede → bare fyll inn tomme celler

Kilde: players.json (birthdate-felt populert av add_alder_sheet.py).
"""
import io, json, shutil, sys, unicodedata, re
from pathlib import Path
from datetime import date
BASE_DIR = Path(__file__).parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as gcl
except ImportError:
    sys.exit("Mangler openpyxl: pip install openpyxl")

EXCEL_PATH   = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
PLAYERS_JSON = BASE_DIR / "players.json"

GROUPS     = [f"Gruppe {x}" for x in "ABCDEFGHIJKL"]
BLOCKS     = [(17, 19, 44), (46, 48, 73), (75, 77, 102), (104, 106, 131)]
COL_BDAY   = 15   # O
BDAY_WIDTH = 13


# ── Stil ──────────────────────────────────────────────────────────────────────

BLUE = PatternFill("solid", fgColor="1A3C6B")
EVEN = PatternFill("solid", fgColor="F0F5FB")
ODD  = PatternFill("solid", fgColor="FFFFFF")

bot  = Border(bottom=Side(style="thin", color="E2E8F0"))
CTR  = Alignment(horizontal="center", vertical="center")

F_HDR   = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
F_DATE  = Font(name="Calibri", size=10, color="6B7A99")


# ── Normalisering for navneoppslag ────────────────────────────────────────────

def _norm(s):
    s = (s or "").lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[-'' ]", "", s)
    return s


# ── Les fødselsdatoer fra players.json ───────────────────────────────────────

with open(PLAYERS_JSON, encoding="utf-8") as f:
    players_data = json.load(f)

name_to_bd: dict[str, str] = {}
missing_bd: list[str] = []

for lags in players_data.values():
    for spillere in lags.values():
        for s in spillere:
            name = s.get("name", "")
            bd   = s.get("birthdate", "")
            if name:
                if bd:
                    try:
                        d = date.fromisoformat(bd[:10])
                        name_to_bd[_norm(name)] = d.strftime("%d.%m.%Y")
                    except ValueError:
                        pass
                else:
                    missing_bd.append(name)

print(f"players.json: {len(name_to_bd)} med fødselsdato, {len(missing_bd)} uten")
if missing_bd:
    print(f"  Fremdeles uten dato: {', '.join(missing_bd[:10])}"
          + (f" … og {len(missing_bd)-10} til" if len(missing_bd) > 10 else ""))


# ── Åpne Excel ────────────────────────────────────────────────────────────────

FORVENTET = 1248

backup = Path(str(EXCEL_PATH) + ".bak")
shutil.copy2(EXCEL_PATH, backup)
print(f"\nBackup: {backup}")

wb = load_workbook(EXCEL_PATH)
tot_inserted = tot_updated = tot_skipped = 0

for gruppe in GROUPS:
    if gruppe not in wb.sheetnames:
        print(f"  {gruppe}: ARK MANGLER — hopper over")
        continue
    ws = wb[gruppe]

    # ── Finn ut om kolonnen allerede er satt inn ──────────────────────────────
    # Sjekk header-label i kolonne 15 (O) på rad 2 (første blokks kolonneoverskrift)
    existing_label = ws.cell(row=2, column=COL_BDAY).value
    already_inserted = (existing_label == "Fødselsdato")

    if already_inserted:
        # Kolonnen finnes — bare fyll inn celler som er tomme
        mode = "oppdater"
    else:
        # Første kjøring — sett inn blank kolonne, skriv header
        ws.insert_cols(COL_BDAY)
        mode = "insert"

    inserted = updated = skipped = 0

    for header_row, start_row, end_row in BLOCKS:
        hdr_col_row = header_row + 1

        if not already_inserted:
            c = ws.cell(row=hdr_col_row, column=COL_BDAY, value="Fødselsdato")
            c.font      = F_HDR
            c.fill      = BLUE
            c.alignment = CTR

        for row in range(start_row, end_row + 1):
            name_val  = ws.cell(row=row, column=13).value   # M = Navn
            bday_cell = ws.cell(row=row, column=COL_BDAY)

            if already_inserted and bday_cell.value:
                skipped += 1
                continue   # dato finnes allerede — ikke overskriv

            bd_str = name_to_bd.get(_norm(name_val or ""), "")
            if not bd_str:
                continue   # ingen dato å sette inn

            # Bakgrunnsfarge fra Pos-nabocellen (P=16)
            try:
                rgb = ws.cell(row=row, column=16).fill.fgColor.rgb
                bg  = PatternFill("solid", fgColor=rgb) if rgb and rgb != "00000000" \
                      else (EVEN if (row - start_row) % 2 == 0 else ODD)
            except Exception:
                bg = EVEN if (row - start_row) % 2 == 0 else ODD

            bday_cell.value     = bd_str
            bday_cell.fill      = bg
            bday_cell.font      = F_DATE
            bday_cell.alignment = CTR
            bday_cell.border    = bot

            if already_inserted:
                updated += 1
            else:
                inserted += 1

    if not already_inserted:
        ws.column_dimensions["O"].width = BDAY_WIDTH

    tot_inserted += inserted
    tot_updated  += updated
    tot_skipped  += skipped
    if mode == "insert":
        print(f"  {gruppe} [ny kolonne]: {inserted} datoer lagt inn")
    else:
        print(f"  {gruppe} [oppdater]:   {updated} nye datoer, {skipped} allerede fylt")

try:
    wb.save(EXCEL_PATH)
    total_fylt = tot_inserted + tot_updated + tot_skipped
    sjekk = "✓" if total_fylt == FORVENTET else f"ADVARSEL: forventet {FORVENTET}"
    print(f"\nLagret: {EXCEL_PATH}")
    print(f"Totalt: {total_fylt} spillere ({sjekk})  —  "
          f"lagt inn: {tot_inserted + tot_updated}, allerede fylt: {tot_skipped}")
except Exception as e:
    shutil.copy2(backup, EXCEL_PATH)
    sys.exit(f"FEIL ved lagring — rullet tilbake fra backup: {e}")
