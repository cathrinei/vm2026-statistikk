#!/usr/bin/env python3
"""
add_sluttspill_spillere_sheet.py
Genererer "Sluttspill – spillere"-arket med individuelle spillerstats
fra knockout-kampene (mål, assist, gule og røde kort).

Datakilder:
  sluttspill_cache.json    — match-IDer per runde, lagnavn + lag-IDer
  lagstatistikk_cache.json — timeline-events per kamp (PlayerName, IdTeam, Type, Period, HomeGoals, AwayGoals)
"""

import io, json, shutil, sys
from collections import defaultdict
from pathlib import Path

BASE_DIR   = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
SLUTTSPILL_CACHE  = BASE_DIR / "sluttspill_cache.json"
LAGSTATS_CACHE    = BASE_DIR / "lagstatistikk_cache.json"

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_SHOOTOUT_PERIOD = 9  # periode-verdi for straffesparkkonkurranse (uverifisert, juster om nødvendig)


def _xl_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin  = Side(style="thin", color="E2E8F0")
    bot   = Border(bottom=thin)
    none_ = Border()
    ctr   = Alignment(horizontal="center", vertical="center")
    lft   = Alignment(horizontal="left",   vertical="center")
    return {
        "NAVY":    PatternFill("solid", fgColor="0F2044"),
        "BLUE":    PatternFill("solid", fgColor="1A3C6B"),
        "GOLD_BG": PatternFill("solid", fgColor="FFFBE6"),
        "SILV_BG": PatternFill("solid", fgColor="F8F8F8"),
        "BRNZ_BG": PatternFill("solid", fgColor="FFF4EC"),
        "EVEN":    PatternFill("solid", fgColor="F0F5FB"),
        "ODD":     PatternFill("solid", fgColor="FFFFFF"),
        "EMPTY":   PatternFill("solid", fgColor="F8F9FA"),
        "f_title": Font(name="Calibri", bold=True,  size=13, color="FFFFFF"),
        "f_hdr":   Font(name="Calibri", bold=True,  size=10, color="FFFFFF"),
        "f_data":  Font(name="Calibri",              size=10, color="1A1A2E"),
        "f_muted": Font(name="Calibri",              size=10, color="6B7A99"),
        "f_rank1": Font(name="Calibri", bold=True,  size=10, color="92680A"),
        "f_rank2": Font(name="Calibri", bold=True,  size=10, color="5C5C5C"),
        "f_rank3": Font(name="Calibri", bold=True,  size=10, color="8B4513"),
        "f_ylw":   Font(name="Calibri", bold=True,  size=10, color="856404"),
        "f_red":   Font(name="Calibri", bold=True,  size=10, color="842029"),
        "ctr": ctr, "lft": lft,
        "bot": bot, "none": none_,
    }


def les_cacher() -> tuple[set[str], dict[str, str]]:
    """Returnerer (sluttspill_match_ids, team_id_to_name)."""
    if not SLUTTSPILL_CACHE.exists():
        return set(), {}

    with open(SLUTTSPILL_CACHE, encoding="utf-8") as f:
        sluttspill = json.load(f)

    match_ids: set[str] = set()
    team_names: dict[str, str] = {}
    for runde_kamper in sluttspill.values():
        for k in runde_kamper:
            if k.get("id"):
                match_ids.add(k["id"])
            if k.get("hjemme_id") and k.get("hjemme"):
                team_names[k["hjemme_id"]] = k["hjemme"]
            if k.get("borte_id") and k.get("borte"):
                team_names[k["borte_id"]] = k["borte"]
    return match_ids, team_names


def bygg_spillerstats(match_ids: set[str], team_names: dict[str, str]) -> tuple[list[dict], list[dict]]:
    """Bygger aggregerte stats per spiller fra lagstatistikk_cache for sluttspill-kamper."""
    if not LAGSTATS_CACHE.exists():
        return [], []

    with open(LAGSTATS_CACHE, encoding="utf-8") as f:
        cache = json.load(f)

    goals:   defaultdict[str, int] = defaultdict(int)
    assists: defaultdict[str, int] = defaultdict(int)
    gule:    defaultdict[str, int] = defaultdict(int)
    rode:    defaultdict[str, int] = defaultdict(int)
    kamper:  defaultdict[str, set] = defaultdict(set)
    navn:    dict[str, str] = {}
    lag:     dict[str, str] = {}

    for match_id, events in cache.items():
        if match_id not in match_ids:
            continue

        # Bygg lag-IDer for denne kampen fra events
        home_id = away_id = ""
        for ev in events:
            # Prøv å utlede hjemme/borte fra score-endring
            if ev.get("HomeGoals") is not None or ev.get("AwayGoals") is not None:
                pass  # bruker bare IdTeam direkte

        prev_h = prev_a = 0
        for ev in events:
            pid       = ev.get("IdPlayer") or ev.get("PlayerName")
            pname     = ev.get("PlayerName", "")
            ev_type   = ev.get("Type")
            ev_team   = ev.get("IdTeam", "")
            period    = ev.get("Period")

            if not pid:
                continue

            # Lagre navn og lag
            if pname and pid not in navn:
                navn[pid] = pname.title()
            if ev_team and pid not in lag:
                lag[pid] = team_names.get(ev_team, ev_team)

            # Kort (gule Type 2, røde Type 3) — ikke i shootout
            is_shootout = period is not None and period >= _SHOOTOUT_PERIOD
            if ev_type == 2 and not is_shootout:
                gule[pid] += 1
                kamper[pid].add(match_id)
            elif ev_type == 3 and not is_shootout:
                rode[pid] += 1
                kamper[pid].add(match_id)

            # Mål (score-change-basert, hopp over shootout)
            cur_h = ev.get("HomeGoals") if ev.get("HomeGoals") is not None else prev_h
            cur_a = ev.get("AwayGoals") if ev.get("AwayGoals") is not None else prev_a
            if cur_h <= prev_h and cur_a <= prev_a:
                prev_h, prev_a = cur_h, cur_a
                continue
            prev_h, prev_a = cur_h, cur_a

            if is_shootout:
                continue

            sub_pid  = ev.get("IdSubPlayer") if ev_type == 0 else None
            sub_name = ""

            if pid:
                goals[pid] += 1
                kamper[pid].add(match_id)
            if sub_pid:
                assists[sub_pid] += 1
                kamper[sub_pid].add(match_id)
                # Prøv å hente navn for sub_pid fra samme event hvis tilgjengelig
                if sub_pid not in navn:
                    navn[sub_pid] = sub_name or sub_pid
                if ev_team and sub_pid not in lag:
                    lag[sub_pid] = team_names.get(ev_team, ev_team)

    all_pids = set(goals) | set(assists) | set(gule) | set(rode)

    scorere = sorted(
        [{"name": navn.get(p, p), "lag": lag.get(p, ""), "kamper": len(kamper[p]),
          "goals": goals[p], "assists": assists[p]}
         for p in all_pids if goals[p] > 0 or assists[p] > 0],
        key=lambda x: (-x["goals"], -x["assists"], x["name"])
    )
    kort_liste = sorted(
        [{"name": navn.get(p, p), "lag": lag.get(p, ""),
          "gule": gule[p], "rode": rode[p]}
         for p in all_pids if gule[p] > 0 or rode[p] > 0],
        key=lambda x: (-(x["rode"]), -x["gule"], x["name"])
    )
    return scorere, kort_liste


def _rang(rows, key_fn):
    rang, cur = [], 1
    for i, p in enumerate(rows):
        rang.append(rang[-1] if i > 0 and key_fn(p) == key_fn(rows[i - 1]) else cur)
        cur = i + 2
    return rang


def skriv_ark(scorere: list[dict], kort_liste: list[dict]) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    S = _xl_styles()
    backup = Path(str(EXCEL_PATH) + ".bak")
    shutil.copy2(EXCEL_PATH, backup)
    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception:
        raise

    sheet_name = "Sluttspill – spillere"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    # Plasser arket etter "Sluttspill" hvis det finnes
    idx = None
    for probe in ("Sluttspill", "Lagstatistikk"):
        if probe in wb.sheetnames:
            idx = wb.sheetnames.index(probe) + 1
            break
    ws = wb.create_sheet(sheet_name, idx)

    row = 1

    def _tittelrad(tekst, ncols):
        nonlocal row
        ws.row_dimensions[row].height = 28
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws.cell(row=row, column=1, value=tekst)
        c.font = S["f_title"]; c.fill = S["NAVY"]; c.alignment = S["lft"]
        row += 1

    def _hdrrad(cols):
        nonlocal row
        ws.row_dimensions[row].height = 20
        for ci, (label, width, al) in enumerate(cols, 1):
            c = ws.cell(row=row, column=ci, value=label)
            c.font = S["f_hdr"]; c.fill = S["BLUE"]
            c.alignment = S["lft"] if al == "l" else S["ctr"]
            ws.column_dimensions[get_column_letter(ci)].width = width
        row += 1

    def _tomrad(ncols):
        nonlocal row
        ws.row_dimensions[row].height = 8
        for ci in range(1, ncols + 1):
            ws.cell(row=row, column=ci).fill = S["EMPTY"]
        row += 1

    # ── Seksjon 1: Scorere og assists ────────────────────────────────────────────
    scor_cols = [("#", 5, "c"), ("Spiller", 28, "l"), ("Lag", 20, "l"),
                 ("Kamper", 8, "c"), ("Mål", 7, "c"), ("Assist", 8, "c")]
    n_scor = len(scor_cols)

    _tittelrad(f"Sluttspill — Scorere og assists   {sum(p['goals'] for p in scorere)} mål · {sum(p['assists'] for p in scorere)} assists", n_scor)
    _hdrrad(scor_cols)

    if scorere:
        ranger = _rang(scorere, lambda p: p["goals"])
        for i, (p, rang) in enumerate(zip(scorere, ranger)):
            ws.row_dimensions[row].height = 17
            if rang == 1:
                bg, f_r = S["GOLD_BG"], S["f_rank1"]
            elif rang == 2:
                bg, f_r = S["SILV_BG"], S["f_rank2"]
            elif rang == 3:
                bg, f_r = S["BRNZ_BG"], S["f_rank3"]
            else:
                bg, f_r = (S["EVEN"] if i % 2 == 0 else S["ODD"]), S["f_muted"]
            vals = [rang, p["name"], p["lag"], p["kamper"] or None,
                    p["goals"] or None, p["assists"] or None]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.fill = bg; c.border = S["bot"]
                c.alignment = S["lft"] if ci in (2, 3) else S["ctr"]
                c.font = f_r if ci == 1 else S["f_data"]
            row += 1
    else:
        ws.row_dimensions[row].height = 17
        c = ws.cell(row=row, column=1, value="Ingen sluttspill-scorere registrert ennå")
        c.font = S["f_muted"]; c.alignment = S["lft"]
        ws.merge_cells(f"A{row}:{get_column_letter(n_scor)}{row}")
        row += 1

    _tomrad(n_scor)

    # ── Seksjon 2: Kort ─────────────────────────────────────────────────────────
    kort_cols = [("#", 5, "c"), ("Spiller", 28, "l"), ("Lag", 20, "l"),
                 ("Gule", 8, "c"), ("Røde", 8, "c")]
    n_kort = len(kort_cols)

    _tittelrad(f"Sluttspill — Kort   {sum(p['gule'] for p in kort_liste)} gule · {sum(p['rode'] for p in kort_liste)} røde", n_kort)
    _hdrrad(kort_cols)

    if kort_liste:
        ranger = _rang(kort_liste, lambda p: (p["rode"], p["gule"]))
        for i, (p, rang) in enumerate(zip(kort_liste, ranger)):
            ws.row_dimensions[row].height = 17
            bg = (S["EVEN"] if i % 2 == 0 else S["ODD"])
            vals = [rang, p["name"], p["lag"], p["gule"] or None, p["rode"] or None]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.fill = bg; c.border = S["bot"]
                c.alignment = S["lft"] if ci in (2, 3) else S["ctr"]
                c.font = S["f_ylw"] if ci == 4 else (S["f_red"] if ci == 5 else S["f_data"])
            row += 1
    else:
        ws.row_dimensions[row].height = 17
        c = ws.cell(row=row, column=1, value="Ingen sluttspill-kort registrert ennå")
        c.font = S["f_muted"]; c.alignment = S["lft"]
        ws.merge_cells(f"A{row}:{get_column_letter(n_kort)}{row}")
        row += 1

    try:
        wb.save(EXCEL_PATH)
    except Exception:
        shutil.copy2(backup, EXCEL_PATH)
        raise

    print(f"  → 'Sluttspill – spillere'-ark skrevet "
          f"({len(scorere)} scorere/assistister, {len(kort_liste)} kortspillere)")


def main():
    print("=" * 55)
    print("  Sluttspill – spillerstatistikk")
    print("=" * 55)

    if not SLUTTSPILL_CACHE.exists():
        print("  sluttspill_cache.json mangler — ingen sluttspill-kamper ennå.")
        return
    if not LAGSTATS_CACHE.exists():
        print("  lagstatistikk_cache.json mangler — kjør add_lagstatistikk_sheet.py først.")
        return

    match_ids, team_names = les_cacher()
    print(f"  {len(match_ids)} sluttspill-kamp-IDer, {len(team_names)} lag-IDer")

    scorere, kort_liste = bygg_spillerstats(match_ids, team_names)
    print(f"  {len(scorere)} spillere med mål/assist, {len(kort_liste)} med kort")

    skriv_ark(scorere, kort_liste)


if __name__ == "__main__":
    main()
