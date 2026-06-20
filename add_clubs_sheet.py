#!/usr/bin/env python3
"""
add_clubs_sheet.py
Legger til "Klubber"-ark i Excel med klubber rangert etter antall VM-spillere.
Kilde: clubs_new.json
"""
import io, json, shutil, sys
from collections import defaultdict, Counter
from pathlib import Path

try:
    from club_country_map import CLUB_COUNTRY
    from level_map import LAND_NO
except ImportError:
    CLUB_COUNTRY = {}
    LAND_NO = {}

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    sys.exit(f"Mangler pakke: {e}")

EXCEL_PATH  = r"C:\Claude_dev\FotballVMClaude\VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
CLUBS_JSON  = r"C:\Claude_dev\FotballVMClaude\clubs_new.json"

# Norske lagnavn
NORSK = {
    "Mexico": "Mexico", "South Africa": "Sør-Afrika", "Republic of Korea": "Sør-Korea",
    "Korea Republic": "Sør-Korea", "South Korea": "Sør-Korea",
    "Czechia": "Tsjekkia", "Czech Republic": "Tsjekkia",
    "Canada": "Canada", "Bosnia and Herzegovina": "Bosnia-Hercegovina",
    "Qatar": "Qatar", "Switzerland": "Sveits",
    "Brazil": "Brasil", "Morocco": "Marokko", "Haiti": "Haiti", "Scotland": "Skottland",
    "United States": "USA", "United States of America": "USA", "USA": "USA",
    "Paraguay": "Paraguay", "Australia": "Australia",
    "Turkey": "Tyrkia", "Türkiye": "Tyrkia",
    "Germany": "Tyskland", "Curacao": "Curaçao", "Curaçao": "Curaçao",
    "Ivory Coast": "Elfenbenskysten", "Côte d'Ivoire": "Elfenbenskysten",
    "Ecuador": "Ecuador", "Sweden": "Sverige", "Japan": "Japan",
    "New Zealand": "New Zealand", "Tunisia": "Tunisia",
    "Spain": "Spania", "Cape Verde": "Kapp Verde", "Cabo Verde": "Kapp Verde",
    "Saudi Arabia": "Saudi-Arabia", "Uruguay": "Uruguay",
    "Netherlands": "Nederland", "Belgium": "Belgia", "Egypt": "Egypt", "Jordan": "Jordan",
    "Norway": "Norge", "France": "Frankrike", "Senegal": "Senegal", "Colombia": "Colombia",
    "Argentina": "Argentina", "Panama": "Panama", "England": "England", "Algeria": "Algerie",
    "Croatia": "Kroatia", "Iran": "Iran", "IR Iran": "Iran",
    "Iraq": "Irak", "Portugal": "Portugal",
    "Ghana": "Ghana", "DR Congo": "DR Kongo", "Congo DR": "DR Kongo",
    "Uzbekistan": "Usbekistan", "Austria": "Østerrike",
}


def bygg_klubb_data() -> list[dict]:
    with open(CLUBS_JSON, encoding="utf-8") as f:
        d = json.load(f)

    club_players: dict[str, list[tuple[str, str]]] = defaultdict(list)
    for team_en, players in d.items():
        team_no = NORSK.get(team_en, team_en)
        for player, club in players.items():
            club_players[club.strip()].append((player, team_no))

    result = []
    for club, players in club_players.items():
        nation_count = Counter(team for _, team in players)
        # Bygg nasjons-streng: "Brasil (3), Tyskland (2)"
        nasjoner = ", ".join(
            f"{n} ({c})" if c > 1 else n
            for n, c in nation_count.most_common()
        )
        land_en = CLUB_COUNTRY.get(club, "")
        land_no = LAND_NO.get(land_en, land_en) if land_en else "–"
        result.append({
            "club":     club,
            "land":     land_no,
            "count":    len(players),
            "nasjoner": nasjoner,
            "players":  sorted(p for p, _ in players),
        })

    return sorted(result, key=lambda x: (-x["count"], x["club"]))


def _xl_styles():
    thin = Side(style="thin", color="E2E8F0")
    bot  = Border(bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center")
    lft  = Alignment(horizontal="left",   vertical="center")
    return {
        "NAVY":    PatternFill("solid", fgColor="0F2044"),
        "BLUE":    PatternFill("solid", fgColor="1A3C6B"),
        "GOLD_BG": PatternFill("solid", fgColor="FFFBE6"),
        "SILV_BG": PatternFill("solid", fgColor="F8F8F8"),
        "BRNZ_BG": PatternFill("solid", fgColor="FFF4EC"),
        "EVEN":    PatternFill("solid", fgColor="F0F5FB"),
        "ODD":     PatternFill("solid", fgColor="FFFFFF"),
        "f_title": Font(name="Calibri", bold=True, size=13, color="FFFFFF"),
        "f_hdr":   Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
        "f_data":  Font(name="Calibri",             size=10, color="1A1A2E"),
        "f_muted": Font(name="Calibri",             size=10, color="6B7A99"),
        "f_rank1": Font(name="Calibri", bold=True,  size=10, color="92680A"),
        "f_rank2": Font(name="Calibri", bold=True,  size=10, color="5C5C5C"),
        "f_rank3": Font(name="Calibri", bold=True,  size=10, color="8B4513"),
        "f_bold":  Font(name="Calibri", bold=True,  size=10, color="1A1A2E"),
        "ctr": ctr, "lft": lft, "bot": bot,
    }


def skriv_klubb_sheet(klubber: list[dict]) -> None:
    S = _xl_styles()

    col_defs = [
        ("#",         5,  S["ctr"]),
        ("Klubb",    30,  S["lft"]),
        ("Land",     18,  S["lft"]),
        ("Spillere",  9,  S["ctr"]),
        ("Nasjoner", 55,  S["lft"]),
    ]
    ncols = len(col_defs)

    backup = EXCEL_PATH + ".bak"
    shutil.copy2(EXCEL_PATH, backup)
    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception:
        raise

    if "Klubber" in wb.sheetnames:
        del wb["Klubber"]
    ws = wb.create_sheet("Klubber")
    ws.auto_filter.ref = "A2:E2"

    # Tittelrad
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    total_spillere = sum(k["count"] for k in klubber)
    c = ws.cell(row=1, column=1,
                value=f"VM 2026 — Klubber   {len(klubber)} klubber, {total_spillere} spillere")
    c.font = S["f_title"]; c.fill = S["NAVY"]; c.alignment = S["lft"]

    # Kolonneoverskrifter
    ws.row_dimensions[2].height = 20
    for col, (label, _, al) in enumerate(col_defs, 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = S["f_hdr"]; c.fill = S["BLUE"]; c.alignment = al

    # Datarader
    rang = 1
    for i, k in enumerate(klubber):
        row = i + 3
        ws.row_dimensions[row].height = 17

        # Lik rang ved likt antall
        if i > 0 and k["count"] == klubber[i-1]["count"]:
            pass  # behold rang
        else:
            rang = i + 1

        if rang == 1:
            bg, f_rank = S["GOLD_BG"], S["f_rank1"]
        elif rang == 2:
            bg, f_rank = S["SILV_BG"], S["f_rank2"]
        elif rang == 3:
            bg, f_rank = S["BRNZ_BG"], S["f_rank3"]
        else:
            bg, f_rank = (S["EVEN"] if i % 2 == 0 else S["ODD"]), S["f_muted"]

        vals = [rang, k["club"], k["land"], k["count"], k["nasjoner"]]
        for col, (val, (_, _, al)) in enumerate(zip(vals, col_defs), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = bg
            c.border = S["bot"]
            c.alignment = al
            if col == 1:
                c.font = f_rank
            elif col == 4:
                c.font = S["f_bold"]
            else:
                c.font = S["f_data"]

    for col, (_, width, _) in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    try:
        wb.save(EXCEL_PATH)
    except Exception:
        shutil.copy2(backup, EXCEL_PATH)
        raise

    print(f"  → 'Klubber'-sheet skrevet ({len(klubber)} klubber)")


def main():
    print("=" * 55)
    print("  VM 2026 — Klubber-ark")
    print("=" * 55)

    FORVENTET = 1248
    print("\n[1/2] Bygger klubbdata fra clubs_new.json...")
    klubber = bygg_klubb_data()
    total_spillere = sum(k["count"] for k in klubber)
    sjekk = "✓" if total_spillere == FORVENTET else f"ADVARSEL: forventet {FORVENTET}"
    print(f"  {len(klubber)} unike klubber, {total_spillere} spillere ({sjekk})")
    topp5 = ", ".join(f"{k['club']} ({k['count']})" for k in klubber[:5])
    print(f"  Topp 5: {topp5}")

    print("\n[2/2] Skriver Excel-ark...")
    skriv_klubb_sheet(klubber)

    print("\nFerdig.")


if __name__ == "__main__":
    main()
