#!/usr/bin/env python3
"""
rebuild_excel.py
Gjenoppbygger VM2026_avansert_gruppetabeller_og_sluttspill.xlsx
fra players.json og avvik_rapport.md.

Rekonstruerer:
  - 12 gruppeark (Gruppe A–L) med spillere, nr, pos, mål, assist
  - Mål/assist hentet fra avvik_rapport.md (FIFA-korrigerte verdier)
  - Gule/røde kort: tomme — kjør check_avvik.py --fix etterpå

Kan IKKE rekonstrueres:
  - Kolonne 1–11 i gruppeark (gruppe-standings/tabeller)
  - Sluttspill-ark
"""
import io, json, re, sys
from pathlib import Path
BASE_DIR = Path(__file__).parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("Mangler openpyxl: pip install openpyxl")

EXCEL_PATH   = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
PLAYERS_JSON = BASE_DIR / "players.json"
CLUBS_JSON   = BASE_DIR / "clubs_new.json"
RAPPORT_MD   = BASE_DIR / "avvik_rapport.md"

GROUPS = [f"Gruppe {x}" for x in "ABCDEFGHIJKL"]
BLOCKS = [(17, 19, 44), (46, 48, 73), (75, 77, 102), (104, 106, 131)]

COL_NR     = 12
COL_NAME   = 13
COL_KLUBB  = 14
COL_BDAY   = 15
COL_POS    = 16
COL_MÅL    = 17
COL_ASSIST = 18
COL_GULT   = 19
COL_RØDT   = 20

# ── Les players.json ──────────────────────────────────────────────────────────

with open(PLAYERS_JSON, encoding="utf-8") as f:
    players_data = json.load(f)

# ── Les klubbdata fra clubs_new.json ──────────────────────────────────────────

import unicodedata, re

def _norm_club(s):
    s = (s or "").lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[-'' ]", "", s)
    return s

with open(CLUBS_JSON, encoding="utf-8") as f:
    _clubs_raw = json.load(f)

name_to_club: dict[str, str] = {}
for _lag_dict in _clubs_raw.values():
    for _pname, _club in _lag_dict.items():
        if _pname and _club:
            name_to_club[_norm_club(_pname)] = _club

# ── Les mål/assist fra avvik_rapport.md ───────────────────────────────────────

scorer_stats: dict[str, tuple[int, int]] = {}   # name_lower → (goals, assists)

md_text = Path(RAPPORT_MD).read_text(encoding="utf-8")
in_table = False
for line in md_text.splitlines():
    if "| # | Spiller" in line:
        in_table = True
        continue
    if in_table:
        if not line.startswith("|"):
            break
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 8 or parts[1] == "---":
            continue
        try:
            name   = parts[2]
            goals  = int(parts[6])
            assists = int(parts[7])
            if goals > 0 or assists > 0:
                scorer_stats[name.lower()] = (goals, assists)
        except (ValueError, IndexError):
            continue

print(f"Lest {len(scorer_stats)} scorere fra avvik_rapport.md")

# ── Design (konsistent med Toppscorere/Kort/Kamper) ──────────────────────────

NAVY     = PatternFill("solid", fgColor="0F2044")   # lagnavn-header
BLUE     = PatternFill("solid", fgColor="1A3C6B")   # kolonneoverskrifter
EVEN_BG  = PatternFill("solid", fgColor="F0F5FB")   # annenhver rad (lys blå)
ODD_BG   = PatternFill("solid", fgColor="FFFFFF")   # annenhver rad (hvit)
GOAL_BG  = PatternFill("solid", fgColor="EBF8F2")   # rad med mål (mintgrønn)

bot_side   = Side(style="thin", color="E2E8F0")
bot_border = Border(bottom=bot_side)

F_TITLE  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
F_HDR    = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
F_NR     = Font(name="Calibri", size=10, color="6B7A99")
F_NAME   = Font(name="Calibri", size=10, color="1A1A2E")
F_MUTED  = Font(name="Calibri", size=10, color="6B7A99")
F_GOAL   = Font(name="Calibri", bold=True, size=10, color="0F5132")

CTR = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left",   vertical="center")

from openpyxl.utils import get_column_letter as gcl

# ── Bygg Excel ────────────────────────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)

FORVENTET      = 1248
total_spillere = [0]
total_klubb    = [0]
total_dato     = [0]
total_mål      = [0]

for gruppe in GROUPS:
    ws = wb.create_sheet(gruppe)
    lag_i_gruppe = players_data.get(gruppe, {})
    lag_list     = list(lag_i_gruppe.items())

    for block_idx, (header_row, start_row, end_row) in enumerate(BLOCKS):
        if block_idx >= len(lag_list):
            break
        lagnavn, spillere = lag_list[block_idx]

        # ── Lagnavn-header (merget over alle 9 kolonner) ──────────────
        merge_start = gcl(COL_NR)
        merge_end   = gcl(COL_RØDT)
        ws.merge_cells(f"{merge_start}{header_row}:{merge_end}{header_row}")
        c = ws.cell(row=header_row, column=COL_NR, value=lagnavn)
        c.font      = F_TITLE
        c.fill      = NAVY
        c.alignment = LFT
        ws.row_dimensions[header_row].height = 24

        # ── Kolonneoverskrifter ───────────────────────────────────────
        col_hdr_row = header_row + 1
        ws.row_dimensions[col_hdr_row].height = 18
        for col, label in [(COL_NR, "Nr"), (COL_NAME, "Navn"), (COL_KLUBB, "Klubb"),
                           (COL_BDAY, "Fødselsdato"), (COL_POS, "Pos"),
                           (COL_MÅL, "Mål"), (COL_ASSIST, "Assist"),
                           (COL_GULT, "Gule"), (COL_RØDT, "Røde")]:
            c = ws.cell(row=col_hdr_row, column=col, value=label)
            c.font      = F_HDR
            c.fill      = BLUE
            c.alignment = LFT if col in (COL_NAME, COL_KLUBB) else CTR

        # ── Spillerader ───────────────────────────────────────────────
        for i, spiller in enumerate(spillere):
            row = start_row + i
            if row > end_row:
                break
            nr             = spiller.get("nr", "")
            name           = spiller.get("name", "")
            pos            = spiller.get("pos", "")
            club           = name_to_club.get(_norm_club(name), "")
            goals, assists = scorer_stats.get(name.lower(), (0, 0))
            has_stat       = goals > 0 or assists > 0
            bg             = GOAL_BG if has_stat else (EVEN_BG if i % 2 == 0 else ODD_BG)
            ws.row_dimensions[row].height = 16

            bd_raw = spiller.get("birthdate", "")
            try:
                from datetime import date as _date
                bd_str = _date.fromisoformat(bd_raw[:10]).strftime("%d.%m.%Y") if bd_raw else None
            except (ValueError, TypeError):
                bd_str = None

            for col, val, al, fnt in [
                (COL_NR,     nr,                    CTR, F_NR),
                (COL_NAME,   name,                  LFT, F_NAME),
                (COL_KLUBB,  club or None,           LFT, F_NAME),
                (COL_BDAY,   bd_str,                CTR, F_MUTED),
                (COL_POS,    pos,                   CTR, F_MUTED),
                (COL_MÅL,    goals if goals > 0 else None,     CTR, F_GOAL),
                (COL_ASSIST, assists if assists > 0 else None, CTR, F_GOAL),
                (COL_GULT,   None,                  CTR, F_MUTED),
                (COL_RØDT,   None,                  CTR, F_MUTED),
            ]:
                c           = ws.cell(row=row, column=col, value=val)
                c.fill      = bg
                c.font      = fnt
                c.alignment = al
                c.border    = bot_border

    # ── Kolonnebredder ────────────────────────────────────────────────
    ws.column_dimensions["L"].width = 5    # Nr
    ws.column_dimensions["M"].width = 26   # Navn
    ws.column_dimensions["N"].width = 22   # Klubb
    ws.column_dimensions["O"].width = 13   # Fødselsdato
    ws.column_dimensions["P"].width = 5    # Pos
    ws.column_dimensions["Q"].width = 6    # Mål
    ws.column_dimensions["R"].width = 7    # Assist
    ws.column_dimensions["S"].width = 6    # Gule
    ws.column_dimensions["T"].width = 6    # Røde

    n_spillere = sum(len(v) for v in lag_i_gruppe.values())
    n_med_klubb = sum(
        1 for lags in lag_i_gruppe.values()
        for s in lags if name_to_club.get(_norm_club(s.get("name", "")))
    )
    n_med_dato = sum(
        1 for lags in lag_i_gruppe.values()
        for s in lags if s.get("birthdate")
    )
    n_med_mål = sum(
        1 for lags in lag_i_gruppe.values()
        for s in lags if scorer_stats.get(s.get("name", "").lower(), (0,0))[0] > 0
    )
    print(f"  {gruppe}: {n_spillere} spillere  "
          f"(klubb: {n_med_klubb}, dato: {n_med_dato}, mål: {n_med_mål})")
    total_spillere[0] += n_spillere
    total_klubb[0]    += n_med_klubb
    total_dato[0]     += n_med_dato
    total_mål[0]      += n_med_mål

wb.save(EXCEL_PATH)
print(f"\n{'─'*55}")
sjekk = "✓" if total_spillere[0] == FORVENTET else f"ADVARSEL: forventet {FORVENTET}"
print(f"  Totalt: {total_spillere[0]} spillere i 12 gruppeark ({sjekk})")
print(f"  Klubb fylt:     {total_klubb[0]}/{total_spillere[0]}")
print(f"  Fødselsdato:    {total_dato[0]}/{total_spillere[0]}")
print(f"  Med mål/assist: {total_mål[0]}/{total_spillere[0]}")
print(f"{'─'*55}")
print(f"\nLagret: {EXCEL_PATH}")
print("\nNeste steg:")
print("  1. python add_country_sheet.py   (gjenoppbygg Spillere etter klubbland)")
print("  2. python check_avvik.py --fix   (fyll inn gule/røde kort fra FIFA)")
