#!/usr/bin/env python3
"""
add_alder_sheet.py
Henter fødselsdato for alle spillere i VM-tropper (via FIFA API lineup-data),
og skriver et "Alder"-ark med de 10 yngste og 10 eldste spillerne.

Cache: player_alder.json  — lagres lokalt for å unngå gjentatte API-kall.
"""
import io, json, re, shutil, sys, time, unicodedata
from datetime import date, datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    import requests
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    sys.exit(f"Mangler pakke: {e}")

EXCEL_PATH   = r"C:\Claude_dev\FotballVMClaude\VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
CACHE_PATH   = r"C:\Claude_dev\FotballVMClaude\kamper_resultater.json"
ALDER_CACHE  = r"C:\Claude_dev\FotballVMClaude\player_alder.json"
LINEUP_CACHE = r"C:\Claude_dev\FotballVMClaude\lineup_cache.json"
PLAYERS_JSON = r"C:\Claude_dev\FotballVMClaude\players.json"

_FIFA_BASE   = "https://api.fifa.com/api/v3"
_FIFA_COMP   = "17"
_FIFA_SEASON = "285023"

# Turneringsstart — alder beregnes per denne datoen
TURNERING_START = date(2026, 6, 12)

# Norsk lagnavn-mapping (samme som i add_kamper_sheets.py)
NORSK = {
    "Mexico": "Mexico", "South Africa": "Sør-Afrika", "South Korea": "Sør-Korea",
    "Korea Republic": "Sør-Korea", "Czechia": "Tsjekkia", "Czech Republic": "Tsjekkia",
    "Canada": "Canada", "Bosnia and Herzegovina": "Bosnia-Hercegovina",
    "Qatar": "Qatar", "Switzerland": "Sveits",
    "Brazil": "Brasil", "Morocco": "Marokko", "Haiti": "Haiti", "Scotland": "Skottland",
    "United States": "USA", "USA": "USA", "Paraguay": "Paraguay",
    "Australia": "Australia", "Turkey": "Tyrkia", "Türkiye": "Tyrkia",
    "Germany": "Tyskland", "Curacao": "Curaçao", "Curaçao": "Curaçao",
    "Ivory Coast": "Elfenbenskysten", "Côte d'Ivoire": "Elfenbenskysten",
    "Ecuador": "Ecuador", "Sweden": "Sverige", "Japan": "Japan",
    "New Zealand": "New Zealand", "Tunisia": "Tunisia",
    "Spain": "Spania", "Cape Verde": "Kapp Verde", "Cabo Verde": "Kapp Verde",
    "Saudi Arabia": "Saudi-Arabia", "Uruguay": "Uruguay",
    "Netherlands": "Nederland", "Belgium": "Belgia", "Egypt": "Egypt", "Jordan": "Jordan",
    "Norway": "Norge", "France": "Frankrike", "Senegal": "Senegal", "Colombia": "Colombia",
    "Argentina": "Argentina", "Panama": "Panama", "England": "England", "Algeria": "Algerie",
    "Croatia": "Kroatia", "Iran": "Iran", "IR Iran": "Iran",
    "Iraq": "Irak", "Portugal": "Portugal",
    "Ghana": "Ghana", "DR Congo": "DR Kongo", "Congo DR": "DR Kongo",
    "Uzbekistan": "Usbekistan", "Austria": "Østerrike",
}

# ── Normalisering og aliases for players.json-oppdatering ────────────────────

_SPECIAL = str.maketrans({
    "ø":"o","Ø":"o","æ":"ae","Æ":"ae","å":"a","Å":"a",
    "ß":"ss","ð":"d","þ":"th","ı":"i",
})
_SPECIAL_DE = str.maketrans({
    "ü":"ue","Ü":"ue","ö":"oe","Ö":"oe","ä":"ae","Ä":"ae",
    "ø":"o","Ø":"o","æ":"ae","Æ":"ae","å":"a","Å":"a",
    "ß":"ss","ð":"d","þ":"th","ı":"i",
})
_JR_RE = re.compile(r"\bjr\.?$", re.IGNORECASE)

def _norm(s, special=_SPECIAL):
    s = (s or "").translate(special)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[-'']", "", s)
    s = " ".join(s.lower().split())
    return _JR_RE.sub("junior", s)

# Excel-navn (players.json) → FIFA API-navn (alder-cache)
_ALIASES: dict[str, str] = {
    _norm(a): _norm(b) for a, b in [
        ("Alisson Becker",           "Alisson"),
        ("Ederson Militao",          "Ederson"),
        ("Danilo Luiz",              "Danilo"),
        ("Gleison Bremer",           "Bremer"),
        ("Ahmed Fathi",              "Ahmed Fathy"),
        ("Homam Al-Amin",            "Homam Ahmed"),
        ("Mohamed Al-Mannai",        "Mohamed Manai"),
        ("Al-Hashmi Al-Hussain",     "Alhashmi Alhussein"),
        ("Ehsan Hajsafi",            "Ehsan Hajisafi"),
        ("Shojae Khalilzadeh",       "Shoja Khalilzadeh"),
        ("Rouzbeh Cheshmi",         "Roozbeh Cheshmi"),
        ("Aria Yousefi",             "Arya Yousefi"),
        ("Danial Eiri",              "Danial Iri"),
        ("Amirmohammad Razzaghinia", "Amirmohammad Razaghinia"),
        ("Shahriyar Moghanlou",      "Shahriyar Moghanloo"),
        ("Mohammad Mohebi",          "Mohammad Mohebbi"),
        ("Husam Abu Dahab",          "Husam Abudahab"),
        ("Mohammad Abu Hashish",     "Mohammad Abuhashish"),
        ("Derrick Etienne Jr.",      "Derrick Etienne"),
        ("Alejandro Zendejas",       "Alex Zendejas"),
        ("Cammy Devlin",             "Cameron Devlin"),
        ("Ko Itakura",               "Kou Itakura"),
        ("Adem Arous",               "Adam Arous"),
        ("Mohamed Amine Ben Hamida", "Mohamed Amine Ben Hmida"),
        ("Mostafa Ziko",             "Mostafa Zico"),
        ("Mohanad Lasheen",          "Mohanad Lashin"),
        ("Marwan Attia",             "Marawan Attia"),
        ("Mostafa Shobeir",          "Mostafa Shoubir"),
        ("Ayman Yahya",              "Aiman Yahya"),
        ("Jehad Thakri",             "Jehad Thikri"),
        ("Abdullah Al-Hamdan",       "Abdullah Alhamddan"),
        ("Nawaf Boushal",            "Nawaf Bu Washl"),
        ("Idrissa Gueye",            "Idrissa Gana Gueye"),
        ("Diney",                    "Diney Borges"),
        ("Yacine Titraoui",          "Yassine Titraoui"),
        ("Phillipp Mwene",           "Phillip Mwene"),
        ("Nicolas Gonzalez",         "Nico Gonzalez"),
    ]
}


def oppdater_players_json(alder_cache: dict[str, dict]) -> None:
    """Skriver birthdate-feltet inn i players.json fra alder-cachen."""
    # Bygg oppslagstabeller: norm(navn) → bd-streng
    navn_bd:    dict[str, str] = {}
    navn_bd_de: dict[str, str] = {}
    forste_ord: dict[str, str] = {}   # for ennavnsspillere i cache

    for info in alder_cache.values():
        name = info.get("name", "")
        bd   = (info.get("birthdate") or "")[:10]
        if not name or not bd:
            continue
        nk  = _norm(name)
        ndk = _norm(name, _SPECIAL_DE)
        navn_bd[nk]     = bd
        navn_bd_de[ndk] = bd
        words = nk.split()
        if len(words) == 1:
            forste_ord[words[0]] = bd

    with open(PLAYERS_JSON, encoding="utf-8") as f:
        data = json.load(f)

    treff = mangler = 0
    for lags in data.values():
        for spillere in lags.values():
            for s in spillere:
                if s.get("birthdate"):
                    treff += 1
                    continue
                key = _norm(s.get("name", ""))
                bd = (
                    navn_bd.get(key)
                    or navn_bd.get(_ALIASES.get(key, ""))
                    or navn_bd_de.get(_norm(s.get("name", ""), _SPECIAL_DE))
                    or forste_ord.get(key.split()[0] if key else "")
                    or navn_bd.get(" ".join(sorted(key.split())))
                )
                if bd:
                    s["birthdate"] = bd
                    treff += 1
                else:
                    mangler += 1

    with open(PLAYERS_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"  → players.json oppdatert: {treff} med dato, {mangler} uten")


_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json",
}

# ── Steg 1: Hent alle spilte kamp-IDer fra FIFA API ──────────────────────────

def hent_spilte_kamper() -> list[tuple[str, str, str]]:
    """Returnerer liste av (mid, gruppe) for spilte kamper."""
    # Bruk cache for gruppe-mapping
    gruppe_for_kamp: dict[str, str] = {}
    if Path(CACHE_PATH).exists():
        with open(CACHE_PATH, encoding="utf-8") as f:
            cache = json.load(f)
        for gruppe, kamper in cache.items():
            for k in kamper:
                if k.get("spilt"):
                    # Bygg nøkkel hjemme|borte for oppslag
                    gruppe_for_kamp[f"{k['hjemme']}|{k['borte']}"] = gruppe

    r = requests.get(
        f"{_FIFA_BASE}/calendar/matches?idCompetition={_FIFA_COMP}&idSeason={_FIFA_SEASON}"
        f"&count=200&language=en-GB",
        headers=_HEADERS, timeout=20
    )
    r.raise_for_status()
    spilte = []
    for m in r.json().get("Results", []):
        if m.get("MatchStatus") != 0:
            continue
        home_en = next((n["Description"] for n in (m.get("Home") or {}).get("TeamName", [])
                        if n.get("Locale") == "en-GB"), "")
        away_en = next((n["Description"] for n in (m.get("Away") or {}).get("TeamName", [])
                        if n.get("Locale") == "en-GB"), "")
        home_no = NORSK.get(home_en, home_en)
        away_no = NORSK.get(away_en, away_en)
        gruppe  = gruppe_for_kamp.get(f"{home_no}|{away_no}", "?")
        spilte.append((m["IdMatch"], home_en, away_en, gruppe))
    return spilte


# ── Steg 2: Hent spilleroppstillinger fra kampene ────────────────────────────

def hent_spillere_fra_kamper(spilte: list) -> dict[str, dict]:
    """Returnerer pid → {name_en, land_no, gruppe} for alle spillere i spilte kamper."""
    # Les lineup-cache
    lineup_cache: dict[str, list] = {}
    if Path(LINEUP_CACHE).exists():
        with open(LINEUP_CACHE, encoding="utf-8") as f:
            lineup_cache = json.load(f)

    nye = [t for t in spilte if t[0] not in lineup_cache]
    print(f"  Oppstillinger: {len(lineup_cache)} i cache, {len(nye)} nye å hente...")

    for mid, home_en, away_en, gruppe in nye:
        try:
            r = requests.get(
                f"{_FIFA_BASE}/live/football/{mid}?language=en-GB",
                headers=_HEADERS, timeout=15
            )
            if r.status_code != 200:
                lineup_cache[mid] = []
                continue
            data = r.json()
            spillere_i_kamp = []
            for side_key, team_en in (("HomeTeam", home_en), ("AwayTeam", away_en)):
                land_no = NORSK.get(team_en, team_en)
                for p in data.get(side_key, {}).get("Players", []):
                    pid = p.get("IdPlayer")
                    if not pid:
                        continue
                    name = next(
                        (n["Description"] for n in p.get("PlayerName", [])
                         if n.get("Locale") == "en-GB"), pid
                    )
                    spillere_i_kamp.append({
                        "pid":    pid,
                        "name":   name.title() if name else pid,
                        "land":   land_no,
                        "gruppe": gruppe,
                    })
            lineup_cache[mid] = spillere_i_kamp
        except Exception:
            lineup_cache[mid] = []
        time.sleep(0.1)

    # Lagre oppdatert cache
    with open(LINEUP_CACHE, "w", encoding="utf-8") as f:
        json.dump(lineup_cache, f, ensure_ascii=False)

    # Bygg spillere-dict fra cache
    spillere: dict[str, dict] = {}
    for mid, home_en, away_en, gruppe in spilte:
        for p in lineup_cache.get(mid, []):
            pid = p["pid"]
            if pid not in spillere:
                spillere[pid] = {"name": p["name"], "land": p["land"], "gruppe": p["gruppe"]}
    return spillere


# ── Steg 3: Hent/cache fødselsdatoer ─────────────────────────────────────────

def hent_aldere(spillere: dict[str, dict]) -> dict[str, dict]:
    """Slår opp fødselsdato fra cache + API. Returnerer oppdatert dict."""
    cache: dict[str, dict] = {}
    if Path(ALDER_CACHE).exists():
        with open(ALDER_CACHE, encoding="utf-8") as f:
            cache = json.load(f)

    nye = [pid for pid in spillere if pid not in cache]
    print(f"  {len(cache)} spillere i cache, {len(nye)} nye å hente...")

    for i, pid in enumerate(nye):
        try:
            r = requests.get(
                f"{_FIFA_BASE}/players/{pid}?language=en-GB",
                headers=_HEADERS, timeout=10
            )
            if r.status_code == 200:
                data = r.json()
                name = next(
                    (n["Description"] for n in data.get("Name", [])
                     if n.get("Locale") == "en-GB"),
                    ""
                )
                cache[pid] = {
                    "name":      name.title() if name else spillere[pid]["name"],
                    "birthdate": data.get("BirthDate", ""),
                }
        except Exception:
            pass
        time.sleep(0.05)
        if (i + 1) % 50 == 0:
            print(f"    ... {i+1}/{len(nye)}")

    # Lagre oppdatert cache
    with open(ALDER_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

    return cache


# ── Steg 4: Beregn alder og bygg sortert liste ────────────────────────────────

def beregn_alder(spillere: dict[str, dict], alder_cache: dict[str, dict]) -> list[dict]:
    resultat = []
    for pid, info in spillere.items():
        cached = alder_cache.get(pid, {})
        bd_raw = cached.get("birthdate", "")
        if not bd_raw:
            continue
        try:
            bd = datetime.fromisoformat(bd_raw.replace("Z", "+00:00")).date()
        except Exception:
            continue
        # Alder i år (heldagsavrunding ved turneringsstart)
        alder_dager = (TURNERING_START - bd).days
        alder_ar    = alder_dager / 365.25
        name = cached.get("name") or info["name"]
        resultat.append({
            "name":    name,
            "land":    info["land"],
            "gruppe":  info["gruppe"],
            "bd":      bd,
            "alder":   alder_ar,
            "bd_str":  bd.strftime("%d.%m.%Y"),
            "alder_str": f"{int(alder_ar)} år {int((alder_ar % 1) * 365)} dg",
        })
    return sorted(resultat, key=lambda x: x["bd"])  # eldst først


# ── Steg 5: Skriv Excel-ark ───────────────────────────────────────────────────

def _xl_styles():
    thin = Side(style="thin", color="E2E8F0")
    bot  = Border(bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center")
    lft  = Alignment(horizontal="left",   vertical="center")
    return {
        "NAVY":    PatternFill("solid", fgColor="0F2044"),
        "BLUE":    PatternFill("solid", fgColor="1A3C6B"),
        "GOLD_BG": PatternFill("solid", fgColor="FFFBE6"),
        "SILV_BG": PatternFill("solid", fgColor="F8F8F8"),
        "BRNZ_BG": PatternFill("solid", fgColor="FFF4EC"),
        "EVEN":    PatternFill("solid", fgColor="F0F5FB"),
        "ODD":     PatternFill("solid", fgColor="FFFFFF"),
        "f_title": Font(name="Calibri", bold=True, size=13, color="FFFFFF"),
        "f_hdr":   Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
        "f_data":  Font(name="Calibri",             size=10, color="1A1A2E"),
        "f_muted": Font(name="Calibri",             size=10, color="6B7A99"),
        "f_rank1": Font(name="Calibri", bold=True,  size=10, color="92680A"),
        "f_rank2": Font(name="Calibri", bold=True,  size=10, color="5C5C5C"),
        "f_rank3": Font(name="Calibri", bold=True,  size=10, color="8B4513"),
        "ctr": ctr, "lft": lft, "bot": bot,
    }


def skriv_alder_sheet(alle: list[dict], n_spillere: int) -> None:
    S = _xl_styles()

    col_defs = [
        ("#",           5,  S["ctr"]),
        ("Spiller",    28,  S["lft"]),
        ("Lag",        20,  S["lft"]),
        ("Gruppe",      9,  S["ctr"]),
        ("Fødselsdato",13,  S["ctr"]),
        ("Alder*",     13,  S["ctr"]),
    ]
    ncols = len(col_defs)

    yngste = list(reversed(alle))[:10]   # yngst = høyest fødselsdato
    eldste = alle[:10]                    # eldst = lavest fødselsdato

    backup = EXCEL_PATH + ".bak"
    shutil.copy2(EXCEL_PATH, backup)
    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception:
        raise

    if "Alder" in wb.sheetnames:
        del wb["Alder"]
    ws = wb.create_sheet("Alder")

    def write_section(tittel: str, rows: list[dict], start_row: int) -> int:
        # Tittelrad
        ws.row_dimensions[start_row].height = 28
        ws.merge_cells(f"A{start_row}:{get_column_letter(ncols)}{start_row}")
        c = ws.cell(row=start_row, column=1, value=tittel)
        c.font = S["f_title"]; c.fill = S["NAVY"]; c.alignment = S["lft"]

        # Kolonneoverskrifter
        hdr_row = start_row + 1
        ws.row_dimensions[hdr_row].height = 20
        for col, (label, _, al) in enumerate(col_defs, 1):
            c = ws.cell(row=hdr_row, column=col, value=label)
            c.font = S["f_hdr"]; c.fill = S["BLUE"]; c.alignment = al

        # Datarader
        for i, p in enumerate(rows, 1):
            row = hdr_row + i
            ws.row_dimensions[row].height = 17
            if i == 1:
                bg, f_rank = S["GOLD_BG"], S["f_rank1"]
            elif i == 2:
                bg, f_rank = S["SILV_BG"], S["f_rank2"]
            elif i == 3:
                bg, f_rank = S["BRNZ_BG"], S["f_rank3"]
            else:
                bg, f_rank = (S["EVEN"] if i % 2 == 0 else S["ODD"]), S["f_muted"]

            vals = [i, p["name"], p["land"], p["gruppe"], p["bd_str"], p["alder_str"]]
            for col, (val, (_, _, al)) in enumerate(zip(vals, col_defs), 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = bg
                c.border = S["bot"]
                c.alignment = al
                c.font = f_rank if col == 1 else S["f_data"]

        return hdr_row + len(rows) + 1  # neste ledig rad

    next_row = write_section(
        f"VM 2026 — 10 yngste spillere  (per {TURNERING_START.strftime('%d.%m.%Y')})",
        yngste, 1
    )
    next_row += 1  # blank rad mellom seksjonene
    write_section(
        f"VM 2026 — 10 eldste spillere  (per {TURNERING_START.strftime('%d.%m.%Y')})",
        eldste, next_row
    )

    # Fotnote
    note_row = next_row + len(eldste) + 3
    ws.cell(row=note_row, column=1,
            value=f"* Alder beregnet per {TURNERING_START.strftime('%d.%m.%Y')} (første kampdag)."
                  f"  Kun spillere fra lag som har spilt kamp er inkludert ({n_spillere} spillere totalt).")
    ws.cell(row=note_row, column=1).font = Font(name="Calibri", size=9, color="6B7A99", italic=True)

    for col, (_, width, _) in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    try:
        wb.save(EXCEL_PATH)
    except Exception:
        shutil.copy2(backup, EXCEL_PATH)
        raise
    print(f"  → 'Alder'-sheet skrevet ({len(yngste)} yngste + {len(eldste)} eldste)")


# ── Hovedprogram ──────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  VM 2026 — Alder-ark (yngste/eldste spillere)")
    print("=" * 55)

    print("\n[1/4] Henter spilte kamper fra FIFA API...")
    spilte = hent_spilte_kamper()
    print(f"  {len(spilte)} spilte kamper funnet")

    print("\n[2/4] Henter spilleroppstillinger...")
    spillere = hent_spillere_fra_kamper(spilte)
    print(f"  {len(spillere)} unike spillere funnet")

    print("\n[3/4] Henter fødselsdatoer...")
    alder_cache = hent_aldere(spillere)

    print("\n[4/5] Skriver Excel-ark...")
    sortert = beregn_alder(spillere, alder_cache)
    print(f"  {len(sortert)} spillere med fødselsdato")
    skriv_alder_sheet(sortert, len(sortert))
    if sortert:
        eldst  = sortert[0]
        yngst  = sortert[-1]
        print(f"  Eldst:  {eldst['name']} ({eldst['land']}) — {eldst['bd_str']} ({eldst['alder_str']})")
        print(f"  Yngst:  {yngst['name']} ({yngst['land']}) — {yngst['bd_str']} ({yngst['alder_str']})")

    print("\n[5/5] Oppdaterer players.json med fødselsdatoer...")
    oppdater_players_json(alder_cache)

    print("\nFerdig.")

if __name__ == "__main__":
    main()
