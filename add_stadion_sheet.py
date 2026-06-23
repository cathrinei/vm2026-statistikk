#!/usr/bin/env python3
"""
add_stadion_sheet.py
Lager arket "Stadionoversikt" med tilskuertall og beleggsgrad per kamp.
Kilde: kamper_resultater.json + tilskuere_cache.json.
"""
import io, json, shutil, sys
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
BASE_DIR = Path(__file__).parent

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    sys.exit(f"Mangler pakke: {e}")

EXCEL_PATH    = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
KAMPER_PATH   = BASE_DIR / "kamper_resultater.json"
TILSKUERE_PATH = BASE_DIR / "tilskuere_cache.json"
SHEET_NAME    = "Stadionoversikt"

STADION_INFO = {
    "Los Angeles Stadium":            {"navn": "SoFi Stadium",           "by": "Inglewood / Los Angeles",  "land": "USA",    "kap": 70240},
    "Mexico City Stadium":            {"navn": "Estadio Azteca",          "by": "Mexico City",              "land": "Mexico", "kap": 87523},
    "New York/New Jersey Stadium":    {"navn": "MetLife Stadium",         "by": "East Rutherford, NJ",      "land": "USA",    "kap": 82500},
    "Dallas Stadium":                 {"navn": "AT&T Stadium",            "by": "Arlington / Dallas",       "land": "USA",    "kap": 80000},
    "Houston Stadium":                {"navn": "NRG Stadium",             "by": "Houston",                  "land": "USA",    "kap": 72220},
    "Atlanta Stadium":                {"navn": "Mercedes-Benz Stadium",   "by": "Atlanta",                  "land": "USA",    "kap": 71000},
    "San Francisco Bay Area Stadium": {"navn": "Levi's Stadium",          "by": "Santa Clara / SF Bay Area","land": "USA",    "kap": 68500},
    "Kansas City Stadium":            {"navn": "Arrowhead Stadium",       "by": "Kansas City",              "land": "USA",    "kap": 76416},
    "Philadelphia Stadium":           {"navn": "Lincoln Financial Field", "by": "Philadelphia",             "land": "USA",    "kap": 69796},
    "Seattle Stadium":                {"navn": "Lumen Field",             "by": "Seattle",                  "land": "USA",    "kap": 68740},
    "Boston Stadium":                 {"navn": "Gillette Stadium",        "by": "Foxborough / Boston",      "land": "USA",    "kap": 65878},
    "Miami Stadium":                  {"navn": "Hard Rock Stadium",       "by": "Miami Gardens / Miami",    "land": "USA",    "kap": 64767},
    "BC Place Vancouver":             {"navn": "BC Place",                "by": "Vancouver",                "land": "Canada", "kap": 54500},
    "Monterrey Stadium":              {"navn": "Estadio BBVA",            "by": "San Nicolás / Monterrey",  "land": "Mexico", "kap": 53464},
    "Guadalajara Stadium":            {"navn": "Estadio Akron",           "by": "Zapopan / Guadalajara",    "land": "Mexico", "kap": 49850},
    "Toronto Stadium":                {"navn": "BMO Field",               "by": "Toronto",                  "land": "Canada", "kap": 45736},
}

LAND_FARGE = {
    "USA":    "0F2044",
    "Mexico": "006847",
    "Canada": "8B0000",
}


def _belegg_fill_font(pct: float):
    if pct >= 97:
        return PatternFill("solid", fgColor="0F5132"), Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    elif pct >= 92:
        return PatternFill("solid", fgColor="1D9E75"), Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    elif pct >= 87:
        return PatternFill("solid", fgColor="C7F0DF"), Font(name="Calibri", bold=True, size=10, color="0F5132")
    else:
        return PatternFill("solid", fgColor="FFF3CD"), Font(name="Calibri", bold=True, size=10, color="856404")


def bygg_data() -> list[dict]:
    with open(KAMPER_PATH, encoding="utf-8") as f:
        kamper = json.load(f)
    with open(TILSKUERE_PATH, encoding="utf-8") as f:
        tilskuere = json.load(f)

    rader = []
    gruppe_order = {f"Gruppe {x}": i for i, x in enumerate("ABCDEFGHIJKL")}

    for gruppe, kamp_liste in kamper.items():
        for k in kamp_liste:
            if not k.get("spilt") or not k.get("id"):
                continue
            mid = k["id"]
            if mid not in tilskuere:
                continue
            t = tilskuere[mid]
            stadion_key = t["stadion"]
            info = STADION_INFO.get(stadion_key, {})
            kap = info.get("kap")
            att = t["att"]
            pct = (att / kap * 100) if (kap and att is not None) else None

            score_h = k.get("score_h")
            score_a = k.get("score_a")
            score_str = f"{score_h}–{score_a}" if score_h is not None else "–"

            rader.append({
                "gruppe":     gruppe,
                "dato":       k["dato"],
                "hjemme":     k["hjemme"],
                "borte":      k["borte"],
                "score":      score_str,
                "arena":      info.get("navn", stadion_key),
                "by":         info.get("by", ""),
                "land":       info.get("land", ""),
                "tilskuere":  att,
                "kapasitet":  kap,
                "belegg_pct": pct,
                "grp_idx":    gruppe_order.get(gruppe, 99),
            })

    rader.sort(key=lambda r: (r["dato"], r["grp_idx"]))
    return rader


def skriv_ark(rader: list[dict]) -> None:
    backup = Path(str(EXCEL_PATH) + ".bak")
    shutil.copy2(EXCEL_PATH, backup)
    wb = load_workbook(EXCEL_PATH)

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    for after in ("Ballbesittelse", "Heatmap", "Scoringstidspunkt", "Lagstatistikk"):
        if after in wb.sheetnames:
            idx = wb.sheetnames.index(after) + 1
            break
    else:
        idx = len(wb.sheetnames)
    ws = wb.create_sheet(SHEET_NAME, idx)

    NAVY  = PatternFill("solid", fgColor="0F2044")
    BLUE  = PatternFill("solid", fgColor="1A3C6B")
    EVEN  = PatternFill("solid", fgColor="F0F5FB")
    WHITE = PatternFill("solid", fgColor="FFFFFF")
    thin  = Side(style="thin",   color="E2E8F0")
    thck  = Side(style="medium", color="0F2044")
    ctr   = Alignment(horizontal="center", vertical="center")
    lft   = Alignment(horizontal="left",   vertical="center")
    rgt   = Alignment(horizontal="right",  vertical="center")

    f_title = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    f_hdr   = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    f_grp   = Font(name="Calibri", bold=True, size=9,  color="FFFFFF")
    f_data  = Font(name="Calibri", size=10,            color="1A1A2E")
    f_muted = Font(name="Calibri", size=10,            color="6B7A99")
    f_score = Font(name="Calibri", bold=True, size=10, color="0F5132")
    f_dato  = Font(name="Calibri", size=10,            color="1A1A2E")

    # A=Dato, B=Gruppe, C=Hjemme, D=Score, E=Borte, F=Arena, G=By, H=Land, I=Tilsk., J=Kap., K=Belegg
    COL_WIDTHS = {"A": 7, "B": 9, "C": 20, "D": 7, "E": 16,
                  "F": 24, "G": 24, "H": 7, "I": 12, "J": 11, "K": 9}
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # ── Rad 1: Tittel ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:K1")
    c = ws.cell(row=1, column=1, value="Tilskuere og arenaer — VM 2026")
    c.font = f_title; c.fill = NAVY; c.alignment = lft

    # ── Rad 2: Kolonneoverskrifter ─────────────────────────────────────────────
    ws.row_dimensions[2].height = 20
    hdrs = [
        (1,  "Dato",      ctr, NAVY),
        (2,  "Gruppe",    ctr, NAVY),
        (3,  "Hjemmelag", lft, BLUE),
        (4,  "Score",     ctr, NAVY),
        (5,  "Bortelag",  lft, BLUE),
        (6,  "Arena",     lft, BLUE),
        (7,  "By",        lft, BLUE),
        (8,  "Land",      ctr, BLUE),
        (9,  "Tilskuere", rgt, BLUE),
        (10, "Kapasitet", rgt, BLUE),
        (11, "Belegg %",  ctr, BLUE),
    ]
    for col, label, al, fill in hdrs:
        c = ws.cell(row=2, column=col, value=label)
        c.font = f_hdr; c.fill = fill; c.alignment = al

    # ── Datarader ──────────────────────────────────────────────────────────────
    for i, r in enumerate(rader):
        row = 3 + i
        ws.row_dimensions[row].height = 18
        bg = EVEN if i % 2 == 0 else WHITE

        dato_fmt = r["dato"][8:10] + "." + r["dato"][5:7]

        belegg_fill, belegg_font = _belegg_fill_font(r["belegg_pct"] or 0)
        land_hex = LAND_FARGE.get(r["land"], "1A3C6B")
        land_fill = PatternFill("solid", fgColor=land_hex)

        def cell(col, val, font=None, fill=None, al=ctr, num_fmt=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = font or f_data
            c.fill      = fill or bg
            c.alignment = al
            c.border    = Border(
                bottom=thin,
                left=thck  if col in (3, 6) else thin,
                right=thck if col in (2, 5) else thin,
            )
            if num_fmt:
                c.number_format = num_fmt
            return c

        cell(1,  dato_fmt,     f_dato,   bg,         ctr)
        cell(2,  r["gruppe"],  f_grp,    NAVY,        ctr)
        cell(3,  r["hjemme"],  f_data,   bg,          lft)
        cell(4,  r["score"],   f_score,  bg,          ctr)
        cell(5,  r["borte"],   f_data,   bg,          lft)
        cell(6,  r["arena"],   f_data,   bg,          lft)
        cell(7,  r["by"],      f_muted,  bg,          lft)
        cell(8,  r["land"],    Font(name="Calibri", bold=True, size=9, color="FFFFFF"),
                               land_fill, ctr)
        cell(9,  r["tilskuere"],  f_data, bg,         rgt, "#,##0")
        cell(10, r["kapasitet"],  f_muted, bg,        rgt, "#,##0")
        if r["belegg_pct"] is not None:
            cell(11, r["belegg_pct"] / 100, belegg_font, belegg_fill, ctr, "0.0%")
        else:
            cell(11, None, f_muted, bg, ctr)

    # ── Forklaring ─────────────────────────────────────────────────────────────
    leg_row = 3 + len(rader) + 1
    ws.row_dimensions[leg_row].height = 18
    ws.cell(row=leg_row, column=1, value="Belegg:").font = Font(
        name="Calibri", bold=True, size=9, color="1A1A2E")
    legend = [
        ("≥97% (praktisk utsolgt)", "0F5132", "FFFFFF"),
        ("92–96%",                  "1D9E75", "FFFFFF"),
        ("87–91%",                  "C7F0DF", "0F5132"),
        ("<87%",                    "FFF3CD", "856404"),
    ]
    for j, (lbl, bg_hex, fg_hex) in enumerate(legend):
        c = ws.cell(row=leg_row, column=j + 2, value=lbl)
        c.fill  = PatternFill("solid", fgColor=bg_hex)
        c.font  = Font(name="Calibri", size=8, color=fg_hex)
        c.alignment = ctr
        c.border = Border(left=Side(style="thin", color="E2E8F0"),
                          right=Side(style="thin", color="E2E8F0"),
                          top=Side(style="thin",   color="E2E8F0"),
                          bottom=Side(style="thin", color="E2E8F0"))

    leg_row += 1
    ws.row_dimensions[leg_row].height = 14
    note = ws.cell(row=leg_row, column=1,
                   value="Kapasitetstall: offisielle arenakapasiteter (ikke FIFA-sertifisert VM-kapasitet som kan avvike noe)")
    note.font = Font(name="Calibri", italic=True, size=8, color="6B7A99")
    ws.merge_cells(f"A{leg_row}:K{leg_row}")

    # ── Sammendrag ─────────────────────────────────────────────────────────────
    sum_row = leg_row + 2
    ws.row_dimensions[sum_row].height = 18
    spilte = [r for r in rader if r["tilskuere"] and r["kapasitet"]]
    if spilte:
        tot_att  = sum(r["tilskuere"] for r in spilte)
        snitt_pct = sum(r["belegg_pct"] for r in spilte if r["belegg_pct"]) / len(spilte)
        max_r    = max(spilte, key=lambda r: r["tilskuere"])
        min_r    = min(spilte, key=lambda r: r["tilskuere"])

        sammendrag = [
            (1, f"Totalt tilskuere: {tot_att:,}".replace(",", " ")),
            (3, f"Snitt belegg: {snitt_pct:.1f}%"),
            (5, f"Høyest: {max_r['tilskuere']:,} ({max_r['hjemme']} – {max_r['borte']})".replace(",", " ")),
            (8, f"Lavest: {min_r['tilskuere']:,} ({min_r['hjemme']} – {min_r['borte']})".replace(",", " ")),
        ]
        for col, tekst in sammendrag:
            c = ws.cell(row=sum_row, column=col, value=tekst)
            c.font = Font(name="Calibri", italic=True, size=9, color="1A3C6B")

    try:
        wb.save(EXCEL_PATH)
    except Exception as e:
        shutil.copy2(backup, EXCEL_PATH)
        sys.exit(f"FEIL — rullet tilbake: {e}")


def main():
    print("=" * 55)
    print("  VM 2026 — Tilskuere og arenaer")
    print("=" * 55)

    for p in (KAMPER_PATH, TILSKUERE_PATH):
        if not p.exists():
            sys.exit(f"FEIL: {p} finnes ikke.")

    print("\n[1/2] Bygger data...")
    rader = bygg_data()
    print(f"  {len(rader)} kamper med tilskuertall")
    for r in rader:
        att_str = f"{r['tilskuere']:6,}" if r['tilskuere'] is not None else "   N/A"
        pct_str = f"{r['belegg_pct']:.1f}%" if r['belegg_pct'] is not None else "  N/A"
        print(f"  {r['dato']}  {r['gruppe']:9s}  {r['hjemme']:22s} {r['score']:5s} {r['borte']:22s}"
              f"  {att_str}  {pct_str}")

    print(f"\n[2/2] Skriver ark '{SHEET_NAME}'...")
    skriv_ark(rader)
    print(f"\nFerdig — {len(rader)} kamper skrevet.")


if __name__ == "__main__":
    main()
