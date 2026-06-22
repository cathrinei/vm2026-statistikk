#!/usr/bin/env python3
"""
add_kamper_sheets.py
Leser kamper fra VM_fotball_menn_2026_kamper.xlsx, henter resultater fra
FIFA API og skriver kampetabeller til gruppe-arkene i hoved-Excel-filen.

Kampene plasseres i kolonne 1–7, radene 1–9 (titel, header, 6 kamper + blank).
Spillerdata i kolonne 12–18 berøres ikke.
"""
import io, json, shutil, sys, time, unicodedata, re
from datetime import datetime, timedelta
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
BASE_DIR = Path(__file__).parent

try:
    import requests
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    sys.exit(f"Mangler pakke: {e}")

# Font/PatternFill etc. brukes inne i _xl_styles() — ikke slett importen over

EXCEL_PATH        = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
KAMPER_PATH       = BASE_DIR / "VM_fotball_menn_2026_kamper.xlsx"
CACHE_PATH        = BASE_DIR / "kamper_resultater.json"
TILSKUERE_CACHE   = BASE_DIR / "tilskuere_cache.json"

_FIFA_BASE   = "https://api.fifa.com/api/v3"
_FIFA_COMP   = "17"
_FIFA_SEASON = "285023"

GROUPS = [f"Gruppe {x}" for x in "ABCDEFGHIJKL"]

# ── Navnemapping ──────────────────────────────────────────────────────────────
# Norske landnavn — kilde: players.json-nøkler + engelske varianter fra kamper.xlsx / FIFA API
NORSK: dict[str, str] = {
    # Gruppe A
    "Mexico":                   "Mexico",
    "South Africa":             "Sør-Afrika",
    "South Korea":              "Sør-Korea",
    "Sør-Korea":                "Sør-Korea",
    "Czechia":                  "Tsjekkia",
    "Tsjekkia":                 "Tsjekkia",
    # Gruppe B
    "Canada":                   "Canada",
    "Bosnia and Herzegovina":   "Bosnia-Hercegovina",
    "Qatar":                    "Qatar",
    "Switzerland":              "Sveits",
    # Gruppe C
    "Brazil":                   "Brasil",
    "Morocco":                  "Marokko",
    "Haiti":                    "Haiti",
    "Scotland":                 "Skottland",
    # Gruppe D
    "United States":            "USA",
    "USA":                      "USA",
    "Paraguay":                 "Paraguay",
    "Australia":                "Australia",
    "Turkey":                   "Tyrkia",
    "Türkiye":                  "Tyrkia",
    # Gruppe E
    "Germany":                  "Tyskland",
    "Curacao":                  "Curaçao",
    "Curaçao":                  "Curaçao",
    "Ivory Coast":              "Elfenbenskysten",
    "Côte d'Ivoire":            "Elfenbenskysten",
    "Ecuador":                  "Ecuador",
    # Gruppe F
    "Sweden":                   "Sverige",
    "Japan":                    "Japan",
    "New Zealand":              "New Zealand",
    "Tunisia":                  "Tunisia",
    # Gruppe G
    "Spain":                    "Spania",
    "Cape Verde":               "Kapp Verde",
    "Cabo Verde":               "Kapp Verde",
    "Saudi Arabia":             "Saudi-Arabia",
    "Uruguay":                  "Uruguay",
    # Gruppe H (sjekk players.json for nøyaktige navn)
    "Netherlands":              "Nederland",
    "Belgium":                  "Belgia",
    "Egypt":                    "Egypt",
    "Jordan":                   "Jordan",
    # Gruppe I
    "Norway":                   "Norge",
    "France":                   "Frankrike",
    "Senegal":                  "Senegal",
    "Colombia":                 "Colombia",
    # Gruppe J
    "Argentina":                "Argentina",
    "Panama":                   "Panama",
    "England":                  "England",
    "Algeria":                  "Algerie",
    # Gruppe K
    "Croatia":                  "Kroatia",
    "Iran":                     "Iran",
    "IR Iran":                  "Iran",
    "Iraq":                     "Irak",
    "Portugal":                 "Portugal",
    # Gruppe L
    "Ghana":                    "Ghana",
    "DR Congo":                 "DR Kongo",
    "Congo DR":                 "DR Kongo",
    "Uzbekistan":               "Usbekistan",
    "Austria":                  "Østerrike",
    # Ekstra varianter
    "Korea Republic":           "Sør-Korea",
    "Czech Republic":           "Tsjekkia",
}

def til_norsk(name: str) -> str:
    return NORSK.get(name, name)

# Internt oppslag: normer → norsk (for matching mot FIFA API)
KAMPER_TO_NO: dict[str, str] = {k: til_norsk(k) for k in NORSK}
FIFA_TO_NO:   dict[str, str] = KAMPER_TO_NO

SPECIAL = str.maketrans({"ø":"o","Ø":"o","æ":"ae","Æ":"ae","å":"a","Å":"a"})
def norm(s: str) -> str:
    s = (s or "").translate(SPECIAL)
    s = unicodedata.normalize("NFD", s)
    return re.sub(r"\s+", "", "".join(c for c in s if unicodedata.category(c) != "Mn").lower())

def to_norsk(name: str, mapping: dict = None) -> str:
    return til_norsk(name)

def _tittel(s: str) -> str:
    """Wilton SAMPAIO -> Wilton Sampaio, STÅLE SOLBAKKEN -> Ståle Solbakken."""
    return " ".join(w.capitalize() for w in (s or "").split())


# ── Les kamper (fra cache eller kamper.xlsx) ──────────────────────────────────

def les_kamper(fra_xlsx: bool = False) -> dict[str, list[dict]]:
    """Laster fra cache hvis den finnes, ellers fra kamper.xlsx."""
    if not fra_xlsx and Path(CACHE_PATH).exists():
        with open(CACHE_PATH, encoding="utf-8") as f:
            return json.load(f)

    wb = load_workbook(KAMPER_PATH, data_only=True)
    grupper: dict[str, list[dict]] = {}
    for g in GROUPS:
        if g not in wb.sheetnames:
            continue
        ws = wb[g]
        kamper = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            dato, hjemme, borte, res = (row + (None,None,None,None))[:4]
            if not hjemme or not borte:
                continue
            hjemme_no = to_norsk(hjemme, KAMPER_TO_NO)
            borte_no  = to_norsk(borte,  KAMPER_TO_NO)
            dato_str  = dato.strftime("%Y-%m-%d") if isinstance(dato, datetime) else str(dato or "")
            kamper.append({
                "dato":      dato_str,
                "hjemme":    hjemme_no,
                "borte":     borte_no,
                "hjemme_en": hjemme,
                "borte_en":  borte,
                "score_h":   None,
                "score_a":   None,
                "spilt":     False,
            })
        grupper[g] = kamper
    return grupper


def lagre_cache(grupper: dict[str, list[dict]]) -> None:
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(grupper, f, ensure_ascii=False, indent=2)

# ── Hent resultater fra FIFA API ──────────────────────────────────────────────

def hent_resultater(grupper: dict[str, list[dict]]) -> set[str]:
    """Henter ferske resultater fra FIFA API.
    Returnerer settet av gruppenavn der noe endret seg."""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json",
    }
    print("  Henter kampdata fra FIFA API...")
    try:
        r = requests.get(
            f"{_FIFA_BASE}/calendar/matches"
            f"?idCompetition={_FIFA_COMP}&idSeason={_FIFA_SEASON}&count=200&language=en-GB",
            headers=headers, timeout=20
        )
        r.raise_for_status()
        alle = r.json().get("Results", [])
    except Exception as e:
        print(f"  ⚠️  Kunne ikke hente FIFA-data: {e}")
        return set()

    # Bygg oppslag: norm(hjemme)|norm(borte) → match-objekt
    fifa_idx: dict[str, dict] = {}
    for m in alle:
        home_raw = next((n["Description"] for n in (m.get("Home") or {}).get("TeamName", [])
                         if n.get("Locale") == "en-GB"), "")
        away_raw = next((n["Description"] for n in (m.get("Away") or {}).get("TeamName", [])
                         if n.get("Locale") == "en-GB"), "")
        home_no = to_norsk(home_raw, FIFA_TO_NO)
        away_no = to_norsk(away_raw, FIFA_TO_NO)
        key = norm(home_no) + "|" + norm(away_no)
        fifa_idx[key] = m

    # Bygg oppslag: lagnavn → FIFA team-ID (både norsk og engelsk, for robust matching)
    team_id_map: dict[str, str] = {}
    for m in alle:
        for side in ("Home", "Away"):
            tid = (m.get(side) or {}).get("IdTeam", "")
            raw = next((n["Description"] for n in (m.get(side) or {}).get("TeamName", [])
                        if n.get("Locale") == "en-GB"), "")
            if tid and raw:
                team_id_map[raw]            = tid  # engelsk FIFA-navn
                team_id_map[to_norsk(raw)]  = tid  # norsk navn

    # Aliases: Excel-headere bruker andre engelske navn enn FIFA API
    _EXCEL_ALIASES = {
        "United States": "USA",
        "Turkey":        "Türkiye",
        "Curacao":       "Curaçao",
        "Ivory Coast":   "Côte d'Ivoire",
        "Cape Verde":    "Cabo Verde",
        "DR Congo":      "Congo DR",
    }
    for excel_name, fifa_name in _EXCEL_ALIASES.items():
        if fifa_name in team_id_map:
            team_id_map[excel_name] = team_id_map[fifa_name]

    kamper_n = treff = 0
    endrede_grupper: set[str] = set()

    for g, kamper in grupper.items():
        for k in kamper:
            kamper_n += 1
            key = norm(k["hjemme"]) + "|" + norm(k["borte"])
            m = fifa_idx.get(key)
            if not m:
                continue
            treff += 1

            # Hent kickoff-dato og -tid fra FIFA API (norsk tid, CEST = UTC+2)
            dato_fifa = m.get("Date", "")
            if dato_fifa:
                try:
                    dt_utc = datetime.strptime(dato_fifa[:19], "%Y-%m-%dT%H:%M:%S")
                    dt_no  = dt_utc + timedelta(hours=2)
                    ny_dato = dt_no.strftime("%Y-%m-%d")
                    ny_tid  = dt_no.strftime("%H:%M")
                    if k.get("tid") != ny_tid or k.get("dato") != ny_dato:
                        k["tid"]  = ny_tid
                        k["dato"] = ny_dato
                        endrede_grupper.add(g)
                except Exception:
                    pass

            if m.get("MatchStatus") == 0:  # spilt
                ny_h = m.get("Home", {}).get("Score")
                ny_a = m.get("Away", {}).get("Score")
                if not k["spilt"] or k["score_h"] != ny_h or k["score_a"] != ny_a:
                    endrede_grupper.add(g)
                k["score_h"] = ny_h
                k["score_a"] = ny_a
                k["spilt"]   = True
                k["id"]      = m.get("IdMatch", "")
                officials = m.get("Officials") or []
                k["dommer"] = next(
                    (next((n["Description"] for n in (o.get("NameShort") or [])
                           if n.get("Locale") == "en-GB"), "")
                     for o in officials if o.get("OfficialType") == 1),
                    ""
                )

    print(f"  {treff}/{kamper_n} kamper matchet mot FIFA API")
    spilt = sum(1 for km in grupper.values() for k in km if k["spilt"])
    print(f"  {spilt} kamper med resultat")
    return endrede_grupper, team_id_map

# ── Hent tilskuertall og trenere fra FIFA API ────────────────────────────────

def hent_tilskuere(grupper: dict, headers: dict) -> tuple[dict[str, int], dict[str, str]]:
    """Henter tilskuertall og stadiumnavn for spilte kamper. Cache: tilskuere_cache.json."""
    raw_cache: dict = {}
    if Path(TILSKUERE_CACHE).exists():
        with open(TILSKUERE_CACHE, encoding="utf-8") as f:
            raw_cache = json.load(f)

    # Migrer gammelt format (int/null) → dict-format
    cache: dict[str, dict] = {}
    for mid, val in raw_cache.items():
        if isinstance(val, dict):
            cache[mid] = val
        else:
            cache[mid] = {"att": val, "stadion": None}

    # Hent for kamper som mangler stadion-data
    needed = {k["id"] for g, kamper in grupper.items()
              for k in kamper if k.get("spilt") and k.get("id")
              and (k["id"] not in cache or cache[k["id"]].get("stadion") is None)}

    cached_with_data = sum(1 for v in cache.values() if isinstance(v, dict) and v.get("stadion"))
    print(f"  {cached_with_data} tilskuertall i cache, {len(needed)} nye å hente...")
    for mid in needed:
        try:
            r = requests.get(f"{_FIFA_BASE}/live/football/{mid}",
                             headers=headers, timeout=10)
            if r.status_code == 200:
                data = r.json() or {}
                att = data.get("Attendance")
                stadium_obj = data.get("Stadium") or {}
                name_list = stadium_obj.get("Name") or []
                stadion = next((n["Description"] for n in name_list
                                if n.get("Locale") == "en-GB"), "")
                cache[mid] = {"att": int(att) if att else None, "stadion": stadion}
            else:
                cache[mid] = {"att": None, "stadion": ""}
        except Exception:
            cache[mid] = {"att": None, "stadion": ""}
        time.sleep(0.15)

    with open(TILSKUERE_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)

    tilskuere = {mid: v.get("att") if isinstance(v, dict) else v for mid, v in cache.items()}
    stadioner = {mid: v.get("stadion", "") if isinstance(v, dict) else "" for mid, v in cache.items()}
    return tilskuere, stadioner


def hent_trenere(headers: dict) -> dict[str, str]:
    """Henter trenere for alle lag fra FIFA API. Returnerer {team_id: navn}."""
    try:
        r = requests.get(
            f"{_FIFA_BASE}/coaches/season/{_FIFA_SEASON}?count=200&language=en-GB",
            headers=headers, timeout=15)
        r.raise_for_status()
        coaches = r.json().get("Results", [])
    except Exception as e:
        print(f"  ADVARSEL: Kunne ikke hente trenere: {e}")
        return {}

    trenere: dict[str, str] = {}
    for c in coaches:
        team_id = c.get("IdTeam", "")
        if c.get("Role") != 0:  # 0 = hovedtrener (1 = assistent)
            continue
        name_list = c.get("Name") or []
        name = next((n["Description"] for n in name_list if n.get("Locale") == "en-GB"), "")
        if team_id and name:
            trenere[team_id] = name
    return trenere


# ── Beregn gruppetabell ───────────────────────────────────────────────────────

def beregn_tabell(lag_liste: list[str], kamper: list[dict]) -> list[dict]:
    """Beregn standings fra kampresultater. Returnerer sortert liste."""
    tabell = {lag: {"lag": lag, "k": 0, "v": 0, "u": 0, "t": 0,
                    "mf": 0, "mm": 0, "p": 0} for lag in lag_liste}
    for k in kamper:
        if not k["spilt"]:
            continue
        h, a = k["hjemme"], k["borte"]
        gh, ga = k["score_h"], k["score_a"]
        if gh is None or ga is None or h not in tabell or a not in tabell:
            continue
        tabell[h]["k"] += 1; tabell[a]["k"] += 1
        tabell[h]["mf"] += gh; tabell[h]["mm"] += ga
        tabell[a]["mf"] += ga; tabell[a]["mm"] += gh
        if gh > ga:
            tabell[h]["v"] += 1; tabell[h]["p"] += 3
            tabell[a]["t"] += 1
        elif gh < ga:
            tabell[a]["v"] += 1; tabell[a]["p"] += 3
            tabell[h]["t"] += 1
        else:
            tabell[h]["u"] += 1; tabell[h]["p"] += 1
            tabell[a]["u"] += 1; tabell[a]["p"] += 1
    for r in tabell.values():
        r["md"] = r["mf"] - r["mm"]
    return sorted(tabell.values(),
                  key=lambda x: (-x["p"], -x["md"], -x["mf"], x["lag"]))

# ── Skriv til Excel ───────────────────────────────────────────────────────────

def fmt_dato(d, tid=None) -> str:
    if d is None:
        return ""
    if isinstance(d, datetime):
        dato_str = d.strftime("%d.%m")
    else:
        try:
            dato_str = datetime.strptime(str(d)[:10], "%Y-%m-%d").strftime("%d.%m")
        except Exception:
            dato_str = str(d)[:10]
    if tid:
        return f"{dato_str} {tid}"
    return dato_str

def fmt_resultat(k: dict) -> str:
    if k["spilt"]:
        h = k["score_h"] if k["score_h"] is not None else "?"
        a = k["score_a"] if k["score_a"] is not None else "?"
        return f"{h} – {a}"
    d = k["dato"]
    if d is None:
        return ""
    try:
        dt = datetime.strptime(str(d)[:10], "%Y-%m-%d")
        if dt.date() <= datetime.now().date():
            return "– – –"
        return ""
    except Exception:
        return ""

def _xl_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="E2E8F0")
    bot  = Border(bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center")
    lft  = Alignment(horizontal="left",   vertical="center")
    return {
        "NAVY":    PatternFill("solid", fgColor="0F2044"),
        "BLUE":    PatternFill("solid", fgColor="1A3C6B"),
        "PLAYED":  PatternFill("solid", fgColor="EBF8F2"),
        "EVEN":    PatternFill("solid", fgColor="F0F5FB"),
        "ODD":     PatternFill("solid", fgColor="FFFFFF"),
        "f_title": Font(name="Calibri", bold=True,  size=13, color="FFFFFF"),
        "f_hdr":   Font(name="Calibri", bold=True,  size=10, color="FFFFFF"),
        "f_data":  Font(name="Calibri",              size=10, color="1A1A2E"),
        "f_muted": Font(name="Calibri",              size=10, color="6B7A99"),
        "f_score": Font(name="Calibri", bold=True,  size=10, color="0F5132"),
        "ctr": ctr, "lft": lft, "bot": bot,
    }

_COL_HEADERS = [
    (1,  "Nr",          "center"),
    (2,  "Navn",        "left"),
    (3,  "Klubb",       "left"),
    (4,  "Fødselsdato", "center"),
    (5,  "Pos",         "center"),
    (6,  "Mål",         "center"),
    (7,  "Assist",      "center"),
    (8,  "Gule",        "center"),
    (9,  "Røde",        "center"),
    (10, "Min",         "center"),
]


def _skriv_trenere(ws, block_header_rows, team_id_map, trenere):
    """Trenernavn i col U (kursiv, NAVY) ved siden av lagnavn-merge L–T; gjenoppretter kolonneoverskrifter på rad+1."""
    from openpyxl.styles import Font as F2, PatternFill as PF2, Alignment as A2
    NAVY = PF2("solid", fgColor="0F2044")
    BLUE = PF2("solid", fgColor="1A3C6B")

    for h_row in block_header_rows:
        raw_name = ws.cell(row=h_row, column=1).value
        if not raw_name:
            continue

        # Rens vekk evt. gammel trenertekst som ble lagt inn i lagnavn-cellen
        base = str(raw_name).strip()
        if "    Trener:" in base:
            base = base.split("    Trener:")[0].strip()

        team_id = team_id_map.get(base, team_id_map.get(til_norsk(base), ""))
        coach   = trenere.get(team_id, "")
        base    = til_norsk(base)   # vis norsk lagnavn i cellen

        # Fjern alle eksisterende merges i h_row som starter i col 1
        to_remove = [mc for mc in ws.merged_cells.ranges
                     if mc.min_row == h_row and mc.min_col == 1]
        for mc in to_remove:
            ws.unmerge_cells(str(mc))
        # Fjern også E–J-merge fra forrige kjøring (starter i col 5)
        to_remove2 = [mc for mc in ws.merged_cells.ranges
                      if mc.min_row == h_row and mc.min_col == 5]
        for mc in to_remove2:
            ws.unmerge_cells(str(mc))

        ws.row_dimensions[h_row].height = 24

        # Lagnavn: merge A–D (cols 1–4) — bold, stor, NAVY
        ws.merge_cells(f"A{h_row}:D{h_row}")
        c = ws.cell(row=h_row, column=1)
        c.value     = base
        c.font      = F2(name="Calibri", bold=True, size=13, color="FFFFFF")
        c.fill      = NAVY
        c.alignment = A2(horizontal="left", vertical="center")

        # Trenernavn: merge E–J (cols 5–10) — kursiv, liten, NAVY
        ws.merge_cells(f"E{h_row}:J{h_row}")
        ct = ws.cell(row=h_row, column=5,
                     value=f"Trener: {_tittel(coach)}" if coach else "")
        ct.font      = F2(name="Calibri", italic=True, size=9, color="FFFFFF")
        ct.fill      = NAVY
        ct.alignment = A2(horizontal="right", vertical="center")

        # Gjenopprett kolonneoverskrifter på h_row+1
        col_row = h_row + 1
        to_remove2 = [mc for mc in ws.merged_cells.ranges
                      if mc.min_row == col_row and mc.min_col == 1]
        for mc in to_remove2:
            ws.unmerge_cells(str(mc))
        ws.row_dimensions[col_row].height = 18
        for col, label, al in _COL_HEADERS:
            ch = ws.cell(row=col_row, column=col, value=label)
            ch.font      = F2(name="Calibri", bold=True, size=10, color="FFFFFF")
            ch.fill      = BLUE
            ch.alignment = A2(horizontal=al, vertical="center")


def skriv_kamper_til_excel(grupper: dict[str, list[dict]],
                           kun_grupper: set[str] | None = None,
                           tilskuere: dict | None = None,
                           stadioner: dict | None = None,
                           trenere: dict | None = None,
                           team_id_map: dict | None = None) -> None:
    S = _xl_styles()
    backup = Path(str(EXCEL_PATH) + ".bak")
    shutil.copy2(EXCEL_PATH, backup)

    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception:
        raise

    runde_labels = ["Runde 1", "Runde 1", "Runde 2", "Runde 2", "Runde 3", "Runde 3"]
    BLOCK_HEADER_ROWS = [17, 46, 75, 104]

    for g in GROUPS:
        if g not in wb.sheetnames or g not in grupper:
            continue
        bare_trenere = kun_grupper is not None and g not in kun_grupper
        ws   = wb[g]
        camp = grupper[g]

        # Trenere skrives alltid (uavhengig av kun_grupper)
        if bare_trenere:
            if trenere and team_id_map:
                _skriv_trenere(ws, BLOCK_HEADER_ROWS, team_id_map, trenere)
            continue

        # Rad 1: Tittel (utvides til G for stadion-kolonne)
        ws.row_dimensions[1].height = 28
        ws.merge_cells("A1:H1")
        c = ws.cell(row=1, column=1, value=f"{g} — Kamper")
        c.font = S["f_title"]; c.fill = S["NAVY"]; c.alignment = S["lft"]

        # Rad 2: Kolonneoverskrifter (inkl. Tilskuere i kol 6, Stadion i kol 7)
        ws.row_dimensions[2].height = 20
        headers_row = ["Runde", "Dato og kl.slett", "Hjemmelag", "Resultat", "Bortelag", "Tilskuere", "Stadion", "Dommer"]
        for col, label in enumerate(headers_row, 1):
            c = ws.cell(row=2, column=col, value=label)
            c.font = S["f_hdr"]; c.fill = S["BLUE"]
            c.alignment = S["lft"] if col in (3, 5, 7, 8) else S["ctr"]
        ws.column_dimensions["B"].width = 14

        # Rad 3–8: Kamper
        for i, k in enumerate(camp):
            row = i + 3
            ws.row_dimensions[row].height = 17
            res = fmt_resultat(k)
            bg  = S["PLAYED"] if k["spilt"] else (S["EVEN"] if i % 2 == 0 else S["ODD"])

            mid = k.get("id", "")
            att = (tilskuere or {}).get(mid)
            att_str = f"{att:,}".replace(",", " ") if att else ""
            stadion_str = (stadioner or {}).get(mid, "") or ""
            dommer_str = _tittel(k.get("dommer", "") or "")

            data = [
                runde_labels[i] if i < len(runde_labels) else "",
                fmt_dato(k["dato"], k.get("tid")),
                k["hjemme"],
                res,
                k["borte"],
                att_str,
                stadion_str,
                dommer_str,
            ]
            for col, val in enumerate(data, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill   = bg
                c.border = S["bot"]
                c.alignment = S["lft"] if col in (3, 5, 7, 8) else S["ctr"]
                if col == 4 and k["spilt"]:
                    c.font = S["f_score"]
                elif col in (1, 2, 6):
                    c.font = S["f_muted"]
                else:
                    c.font = S["f_data"]

        # Rad 9: blank separator
        for col in range(1, 10):
            c = ws.cell(row=9, column=col)
            c.value = None; c.fill = S["ODD"]

        # ── Gruppetabell ──────────────────────────────────────────────
        # Rad 10: Seksjonstittel
        ws.row_dimensions[10].height = 26
        ws.merge_cells("A10:I10")
        c = ws.cell(row=10, column=1, value=f"{g} — Tabell")
        c.font = S["f_title"]; c.fill = S["NAVY"]; c.alignment = S["lft"]

        # Rad 11: Kolonneoverskrifter
        ws.row_dimensions[11].height = 20
        tab_cols = [
            ("Lag",  22, S["lft"]),
            ("K",     5, S["ctr"]),
            ("V",     5, S["ctr"]),
            ("U",     5, S["ctr"]),
            ("T",     5, S["ctr"]),
            ("MF",    6, S["ctr"]),
            ("MM",    6, S["ctr"]),
            ("MD",    6, S["ctr"]),
            ("P",     6, S["ctr"]),
        ]
        for col, (label, _, al) in enumerate(tab_cols, 1):
            c = ws.cell(row=11, column=col, value=label)
            c.font = S["f_hdr"]; c.fill = S["BLUE"]; c.alignment = al

        # Hent lagene i gruppen i riktig rekkefølge (fra kamper)
        lag_i_gruppe = list(dict.fromkeys(
            lag for k in camp for lag in (k["hjemme"], k["borte"])
        ))
        standings = beregn_tabell(lag_i_gruppe, camp)

        # Rad 12–15: Lag
        for i, row_data in enumerate(standings):
            row = 12 + i
            ws.row_dimensions[row].height = 17
            bg = S["EVEN"] if i % 2 == 0 else S["ODD"]

            # Fremhev lag som er kvalifisert (topp 2) eller på vippen (plass 3)
            if i == 0:
                bg = S["PLAYED"]  # grønn: trygg topp
            elif i == 1:
                from openpyxl.styles import PatternFill as PF
                bg = PF("solid", fgColor="EBF5FB")  # lys blå: 2. plass

            played = row_data["k"] > 0
            md     = row_data["md"]
            md_str = (f"+{md}" if md > 0 else str(md)) if played else None
            vals = [
                row_data["lag"],
                row_data["k"]  if played else None,
                row_data["v"]  if played else None,
                row_data["u"]  if played else None,
                row_data["t"]  if played else None,
                row_data["mf"] if played else None,
                row_data["mm"] if played else None,
                md_str,
                row_data["p"]  if played else None,
            ]
            for col, (val, (_, _, al)) in enumerate(zip(vals, tab_cols), 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = bg
                c.border = S["bot"]
                c.alignment = al
                if col == 9 and row_data["k"] > 0:
                    c.font = S["f_score"]
                elif col == 1:
                    c.font = S["f_data"]
                else:
                    c.font = S["f_muted"]

        # Trenere — legg trener i lag-headerraden; gjenopprett kolonneoverskrifter på rad+1
        if trenere and team_id_map:
            _skriv_trenere(ws, BLOCK_HEADER_ROWS, team_id_map, trenere)

        # Kolonnebredder — F=12 for tilskuere (erstatter MF=6 som er i F for standings)
        widths = {"A": 9, "B": 8, "C": 22, "D": 10, "E": 22, "F": 12, "G": 20, "H": 20, "I": 6}
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

    try:
        wb.save(EXCEL_PATH)
    except Exception:
        shutil.copy2(backup, EXCEL_PATH)
        raise

    n = len(kun_grupper) if kun_grupper is not None else len(grupper)
    print(f"  Kampetabeller skrevet til {n} gruppeark")

# ── Hovedprogram ──────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--full", action="store_true",
                        help="Les kamper.xlsx på nytt og skriv alle ark (ignorer cache)")
    args = parser.parse_args()

    print("=" * 55)
    print("  VM 2026 — Oppdater gruppekamper i Excel")
    print("=" * 55)

    fra_xlsx = args.full or not Path(CACHE_PATH).exists()
    kilde = "kamper.xlsx" if fra_xlsx else "cache"
    print(f"\n[1/4] Leser kamper fra {kilde}...")
    grupper = les_kamper(fra_xlsx=fra_xlsx)
    totalt = sum(len(v) for v in grupper.values())
    print(f"  {totalt} kamper lest fra {len(grupper)} grupper")

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "application/json",
    }

    print("\n[2/4] Henter ferske resultater fra FIFA API...")
    endrede, team_id_map = hent_resultater(grupper)

    if not endrede and not fra_xlsx:
        print("\n  Ingen endringer siden sist — Excel er allerede oppdatert.")
        lagre_cache(grupper)
        print("\nFerdig.")
        return

    if endrede:
        print(f"  Endringer i: {', '.join(sorted(endrede))}")
    lagre_cache(grupper)

    print("\n[3/4] Henter tilskuertall fra FIFA API...")
    tilskuere, stadioner = hent_tilskuere(grupper, headers)
    print(f"  {sum(1 for v in tilskuere.values() if v)} kamper med tilskuertall")

    trenere = {}
    if fra_xlsx:
        print("  Henter trenere (kun ved --full)...")
        trenere = hent_trenere(headers)
        print(f"  {len(trenere)} trenere funnet")

    kun = None if fra_xlsx else endrede
    print(f"\n[4/4] Skriver til Excel ({('alle grupper' if kun is None else str(len(kun)) + ' grupper')})...")
    skriv_kamper_til_excel(grupper, kun_grupper=kun,
                           tilskuere=tilskuere, stadioner=stadioner, trenere=trenere, team_id_map=team_id_map)

    print("\nFerdig. Kjør check_avvik.py --fix for å holde statistikk oppdatert.")

if __name__ == "__main__":
    main()
