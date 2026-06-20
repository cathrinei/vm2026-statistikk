#!/usr/bin/env python3
"""
add_scoringstidspunkt_sheet.py
Lager arket "Scoringstidspunkt" med fordeling av mål per 15-minutters intervall
og et innebygd søylediagram. Kilde: lagstatistikk_cache.json.
"""
import io, json, shutil, sys
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
BASE_DIR = Path(__file__).parent

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import SeriesLabel
except ImportError as e:
    sys.exit(f"Mangler pakke: {e}")

EXCEL_PATH  = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
STAT_CACHE  = BASE_DIR / "lagstatistikk_cache.json"
SHEET_NAME  = "Scoringstidspunkt"

GOAL_TYPES  = {"Goal!", "Penalty Goal", "Own goal"}

INTERVALS = [
    ("1–15",    1,  15),
    ("16–30",  16,  30),
    ("31–45",  31,  45),
    ("46–60",  46,  60),
    ("61–75",  61,  75),
    ("76–90+", 76, 999),
]


def _parse_min(s: str) -> int | None:
    if not s:
        return None
    s = str(s).strip().rstrip("'")
    if "'+'" in s or "'+" in s:
        parts = s.replace("'", "").split("+")
        try:
            return int(parts[0]) + int(parts[1])
        except Exception:
            return None
    try:
        return int(s)
    except Exception:
        return None


def tell_maal() -> tuple[list[int], int, list[dict]]:
    """Returnerer (counts per intervall, totalt, rådata per mål)."""
    if not Path(STAT_CACHE).exists():
        sys.exit(f"FEIL: {STAT_CACHE} finnes ikke — kjør add_lagstatistikk_sheet.py først.")

    with open(STAT_CACHE, encoding="utf-8") as f:
        cache = json.load(f)

    counts = [0] * len(INTERVALS)
    maal_data = []

    for mid, events in cache.items():
        for e in events:
            if e.get("TypeLocalized") not in GOAL_TYPES:
                continue
            minutt = _parse_min(e.get("MatchMinute"))
            if minutt is None:
                continue
            for i, (_, lo, hi) in enumerate(INTERVALS):
                if lo <= minutt <= hi:
                    counts[i] += 1
                    break
            maal_data.append({
                "minutt":  minutt,
                "type":    e.get("TypeLocalized"),
                "spiller": e.get("PlayerName", ""),
            })

    return counts, sum(counts), maal_data


def skriv_ark(counts: list[int], totalt: int) -> None:
    backup = Path(str(EXCEL_PATH) + ".bak")
    shutil.copy2(EXCEL_PATH, backup)
    wb = load_workbook(EXCEL_PATH)

    # Fjern gammelt ark hvis det finnes
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    # Plasser arket etter Lagstatistikk (eller sist)
    insert_after = "Lagstatistikk"
    if insert_after in wb.sheetnames:
        idx = wb.sheetnames.index(insert_after) + 1
    else:
        idx = len(wb.sheetnames)
    ws = wb.create_sheet(SHEET_NAME, idx)

    NAVY   = PatternFill("solid", fgColor="0F2044")
    BLUE   = PatternFill("solid", fgColor="1A3C6B")
    EVEN   = PatternFill("solid", fgColor="F0F5FB")
    WHITE  = PatternFill("solid", fgColor="FFFFFF")
    bot    = Border(bottom=Side(style="thin", color="E2E8F0"))
    ctr    = Alignment(horizontal="center", vertical="center")
    lft    = Alignment(horizontal="left",   vertical="center")

    f_title = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    f_hdr   = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    f_data  = Font(name="Calibri", size=10, color="1A1A2E")
    f_muted = Font(name="Calibri", size=10, color="6B7A99")
    f_score = Font(name="Calibri", bold=True, size=10, color="0F5132")

    # ── Tittel ────────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=1, value="Scoringstidspunkt — Mål per 15-minutters intervall")
    c.font = f_title; c.fill = NAVY; c.alignment = lft

    # ── Kolonneoverskrifter ───────────────────────────────────────────────────
    ws.row_dimensions[2].height = 20
    for col, label, al in [(1, "Intervall", ctr), (2, "Mål", ctr), (3, "Andel", ctr)]:
        c = ws.cell(row=2, column=col, value=label)
        c.font = f_hdr; c.fill = BLUE; c.alignment = al

    # ── Datarader ─────────────────────────────────────────────────────────────
    for i, ((label, _, __), count) in enumerate(zip(INTERVALS, counts)):
        row = 3 + i
        ws.row_dimensions[row].height = 18
        bg = EVEN if i % 2 == 0 else WHITE
        andel = count / totalt if totalt else 0

        c1 = ws.cell(row=row, column=1, value=label)
        c1.font = f_data; c1.fill = bg; c1.border = bot; c1.alignment = ctr

        c2 = ws.cell(row=row, column=2, value=count)
        c2.font = f_score; c2.fill = bg; c2.border = bot; c2.alignment = ctr

        c3 = ws.cell(row=row, column=3, value=andel)
        c3.font = f_muted; c3.fill = bg; c3.border = bot; c3.alignment = ctr
        c3.number_format = "0.0%"

    # ── Totalt-rad ────────────────────────────────────────────────────────────
    ws.row_dimensions[9].height = 18
    for col, val, font, al in [
        (1, "Totalt", f_hdr, ctr),
        (2, totalt,   f_hdr, ctr),
        (3, "",       f_hdr, ctr),
    ]:
        c = ws.cell(row=9, column=col, value=val)
        c.font = font; c.fill = BLUE; c.alignment = al

    # ── Søylediagram ──────────────────────────────────────────────────────────
    chart = BarChart()
    chart.type    = "col"
    chart.grouping = "clustered"
    chart.title   = "Mål per 15-minutters intervall"
    chart.y_axis.title = "Antall mål"
    chart.x_axis.title = "Minutt"
    chart.legend  = None
    chart.width   = 16
    chart.height  = 12

    # Data: Mål-verdier (rad 3–8, kolonne B)
    data_ref = Reference(ws, min_col=2, min_row=3, max_row=8)
    chart.add_data(data_ref)

    # Kategorier: intervallnavnene (kolonne A)
    cats = Reference(ws, min_col=1, min_row=3, max_row=8)
    chart.set_categories(cats)


    # Plasser diagrammet til høyre for tabellen (fra celle E1)
    ws.add_chart(chart, "E1")

    # ── Kolonnebredder ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 10

    try:
        wb.save(EXCEL_PATH)
    except Exception as e:
        shutil.copy2(backup, EXCEL_PATH)
        sys.exit(f"FEIL — rullet tilbake: {e}")

    # Fiks openpyxl-bugs i chart-XML
    import zipfile, re as _re, os
    tmp = Path(str(EXCEL_PATH) + ".tmp")
    with zipfile.ZipFile(EXCEL_PATH, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/charts/chart1.xml":
                xml = data.decode("utf-8")
                # 1. catAx: axPos skal være "b" (bunn) for kolonnediagram
                xml = xml.replace('<axPos val="l"/>', '<axPos val="b"/>', 1)
                # 2. Fjern ufullstendig spPr med bare prstDash (ingen linjefarge)
                xml = _re.sub(r'<spPr><a:ln[^>]*><a:prstDash[^/]*/></a:ln></spPr>', '', xml)
                # 3. Legg til crosses="autoZero" på begge akser hvis mangler
                xml = xml.replace('<majorTickMark val="none"/><minorTickMark val="none"/><crossAx',
                                  '<majorTickMark val="none"/><minorTickMark val="none"/><crosses val="autoZero"/><crossAx')
                data = xml.encode("utf-8")
            zout.writestr(item, data)
    os.replace(tmp, EXCEL_PATH)


def main():
    print("=" * 55)
    print("  VM 2026 — Scoringstidspunkt")
    print("=" * 55)

    print("\n[1/2] Teller mål per intervall fra cache...")
    counts, totalt, _ = tell_maal()
    for (label, _, __), n in zip(INTERVALS, counts):
        andel = n / totalt * 100 if totalt else 0
        print(f"  {label:8s}  {n:3d} mål  ({andel:.1f}%)")
    print(f"  {'Totalt':8s}  {totalt:3d} mål")

    print(f"\n[2/2] Skriver ark '{SHEET_NAME}' til Excel...")
    skriv_ark(counts, totalt)
    print(f"\nFerdig.")


if __name__ == "__main__":
    main()
