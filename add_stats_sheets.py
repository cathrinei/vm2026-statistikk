#!/usr/bin/env python3
"""
add_stats_sheets.py
Legger til to statistikkark i Excel-filen:
  - "Aldersfordeling"  : snittalder, yngste og eldste spiller per lag
  - "Klubbdominans"    : hvilke klubber har flest spillere i VM
"""
import io, json, shutil, sys
from collections import defaultdict
from datetime import date
from pathlib import Path
BASE_DIR = Path(__file__).parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as gcl
except ImportError:
    sys.exit("Mangler openpyxl: pip install openpyxl")

try:
    from club_country_map import CLUB_COUNTRY
    from level_map import LAND_NO
except ImportError:
    CLUB_COUNTRY = {}
    LAND_NO = {}

EXCEL_PATH   = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
PLAYERS_JSON = BASE_DIR / "players.json"
CLUBS_JSON   = BASE_DIR / "clubs_new.json"

TURNERING_START = date(2026, 6, 12)

# ── Lagnavn-oversettelse (engelsk → norsk) ────────────────────────────────────

NORSK: dict[str, str] = {
    "Mexico": "Mexico", "South Africa": "Sør-Afrika", "South Korea": "Sør-Korea",
    "Korea Republic": "Sør-Korea", "Republic of Korea": "Sør-Korea",
    "Czechia": "Tsjekkia", "Czech Republic": "Tsjekkia",
    "Canada": "Canada", "Bosnia and Herzegovina": "Bosnia-Hercegovina",
    "Qatar": "Qatar", "Switzerland": "Sveits",
    "Brazil": "Brasil", "Morocco": "Marokko", "Haiti": "Haiti", "Scotland": "Skottland",
    "United States": "USA", "United States of America": "USA", "USA": "USA",
    "Paraguay": "Paraguay", "Australia": "Australia",
    "Turkey": "Tyrkia", "Türkiye": "Tyrkia",
    "Germany": "Tyskland", "Curacao": "Curaçao", "Curaçao": "Curaçao",
    "Ivory Coast": "Elfenbenskysten", "Côte d'Ivoire": "Elfenbenskysten",
    "Ecuador": "Ecuador", "Sweden": "Sverige", "Japan": "Japan",
    "New Zealand": "New Zealand", "Tunisia": "Tunisia",
    "Spain": "Spania", "Cape Verde": "Kapp Verde", "Cabo Verde": "Kapp Verde",
    "Saudi Arabia": "Saudi-Arabia", "Uruguay": "Uruguay",
    "Netherlands": "Nederland", "Belgium": "Belgia", "Egypt": "Egypt", "Jordan": "Jordan",
    "Norway": "Norge", "France": "Frankrike", "Senegal": "Senegal", "Colombia": "Colombia",
    "Argentina": "Argentina", "Panama": "Panama", "England": "England", "Algeria": "Algerie",
    "Croatia": "Kroatia", "Iran": "Iran", "IR Iran": "Iran",
    "Iraq": "Irak", "Portugal": "Portugal",
    "Ghana": "Ghana", "DR Congo": "DR Kongo", "Congo DR": "DR Kongo",
    "Democratic Republic of the Congo": "DR Kongo",
    "Uzbekistan": "Usbekistan", "Austria": "Østerrike",
    "Sør-Korea": "Sør-Korea", "Tsjekkia": "Tsjekkia",   # allerede norsk
}

def no(name: str) -> str:
    return NORSK.get(name, name)


# ── Designsystem (felles) ─────────────────────────────────────────────────────

def _styles():
    bot = Border(bottom=Side(style="thin", color="E2E8F0"))
    return {
        "NAVY":    PatternFill("solid", fgColor="0F2044"),
        "BLUE":    PatternFill("solid", fgColor="1A3C6B"),
        "GOLD":    PatternFill("solid", fgColor="FFFBE6"),
        "SILV":    PatternFill("solid", fgColor="F8F8F8"),
        "BRNZ":    PatternFill("solid", fgColor="FFF4EC"),
        "EVEN":    PatternFill("solid", fgColor="F0F5FB"),
        "ODD":     PatternFill("solid", fgColor="FFFFFF"),
        "MINT":    PatternFill("solid", fgColor="EBF8F2"),
        "f_title": Font(name="Calibri", bold=True, size=13, color="FFFFFF"),
        "f_hdr":   Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
        "f_data":  Font(name="Calibri", size=10, color="1A1A2E"),
        "f_muted": Font(name="Calibri", size=10, color="6B7A99"),
        "f_rank1": Font(name="Calibri", bold=True, size=10, color="92680A"),
        "f_rank2": Font(name="Calibri", bold=True, size=10, color="5C5C5C"),
        "f_rank3": Font(name="Calibri", bold=True, size=10, color="8B4513"),
        "ctr": Alignment(horizontal="center", vertical="center"),
        "lft": Alignment(horizontal="left",   vertical="center"),
        "bot": bot,
    }

def _rank_style(S, rang):
    if rang == 1:   return S["GOLD"], S["f_rank1"]
    if rang == 2:   return S["SILV"], S["f_rank2"]
    if rang == 3:   return S["BRNZ"], S["f_rank3"]
    return None, None

def _write_title(ws, S, row, ncols, text):
    ws.row_dimensions[row].height = 28
    ws.merge_cells(f"A{row}:{gcl(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = S["f_title"]; c.fill = S["NAVY"]; c.alignment = S["lft"]

def _write_headers(ws, S, row, col_defs):
    ws.row_dimensions[row].height = 20
    for col, (label, _, al) in enumerate(col_defs, 1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = S["f_hdr"]; c.fill = S["BLUE"]; c.alignment = al

def _set_widths(ws, col_defs):
    for col, (_, w, _) in enumerate(col_defs, 1):
        ws.column_dimensions[gcl(col)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# ARK 1: Aldersfordeling per lag
# ─────────────────────────────────────────────────────────────────────────────

def bygg_aldersfordeling() -> list[dict]:
    with open(PLAYERS_JSON, encoding="utf-8") as f:
        data = json.load(f)

    lag_data = []
    for gruppe, lags in data.items():
        for lagnavn, spillere in lags.items():
            aldre = []
            for s in spillere:
                bd = s.get("birthdate", "")
                if not bd:
                    continue
                try:
                    d = date.fromisoformat(bd[:10])
                    alder = (TURNERING_START - d).days / 365.25
                    aldre.append((alder, s["name"], d))
                except ValueError:
                    continue

            if not aldre:
                continue

            aldre.sort(key=lambda x: x[2])   # eldst først for eldste-oppslag
            eldst  = aldre[0]
            yngst  = aldre[-1]
            snitt  = sum(a for a, _, _ in aldre) / len(aldre)

            def fmt(a): return f"{int(a)} år {int((a % 1) * 365)} dg"

            lag_data.append({
                "lag":        no(lagnavn),
                "gruppe":     gruppe,
                "snitt":      snitt,
                "snitt_str":  f"{snitt:.1f}".replace(".", ","),
                "eldst_navn": eldst[1],
                "eldst_alder":fmt(eldst[0]),
                "yngst_navn": yngst[1],
                "yngst_alder":fmt(yngst[0]),
                "antall":     len(aldre),
                "totalt":     len(spillere),
            })

    return sorted(lag_data, key=lambda x: x["snitt"])   # yngste lag øverst


def skriv_aldersfordeling(wb, lag_data: list[dict]) -> None:
    S = _styles()
    sheet = "Aldersfordeling"
    if sheet in wb.sheetnames:
        del wb[sheet]
    ws = wb.create_sheet(sheet)

    col_defs = [
        ("#",              5,  S["ctr"]),
        ("Lag",           22,  S["lft"]),
        ("Gruppe",         9,  S["ctr"]),
        ("Snittalder",    12,  S["ctr"]),
        ("Yngste spiller",26,  S["lft"]),
        ("Alder",         14,  S["ctr"]),
        ("Eldste spiller",26,  S["lft"]),
        ("Alder",         14,  S["ctr"]),
        ("Spillere m/dato",13, S["ctr"]),
    ]

    _write_title(ws, S, 1, len(col_defs),
                 f"VM 2026 — Aldersfordeling per lag  (per {TURNERING_START.strftime('%d.%m.%Y')})")
    _write_headers(ws, S, 2, col_defs)

    for i, lag in enumerate(lag_data, 1):
        row = i + 2
        ws.row_dimensions[row].height = 17
        bg_override, f_rank = _rank_style(S, i)
        bg = bg_override if bg_override else (S["EVEN"] if i % 2 == 0 else S["ODD"])

        vals = [
            i, lag["lag"], lag["gruppe"], lag["snitt_str"],
            lag["yngst_navn"], lag["yngst_alder"],
            lag["eldst_navn"], lag["eldst_alder"],
            f"{lag['antall']}/{lag['totalt']}",
        ]
        for col, (val, (_, _, al)) in enumerate(zip(vals, col_defs), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = bg; c.border = S["bot"]; c.alignment = al
            c.font = f_rank if (col == 1 and f_rank) else S["f_data"]

    _set_widths(ws, col_defs)
    print(f"  'Aldersfordeling': {len(lag_data)} lag skrevet")


# ─────────────────────────────────────────────────────────────────────────────
# ARK 2: Klubbdominans
# ─────────────────────────────────────────────────────────────────────────────

def bygg_klubbdominans() -> list[dict]:
    with open(CLUBS_JSON, encoding="utf-8") as f:
        clubs_data = json.load(f)

    klub_count: dict[str, int] = defaultdict(int)
    klub_land:  dict[str, str] = {}
    klub_lag:   dict[str, set] = defaultdict(set)

    for lagnavn, spillere in clubs_data.items():
        lag_no = no(lagnavn)
        for _, klubb in spillere.items():
            if not klubb:
                continue
            klubb = klubb.strip()
            klub_count[klubb] += 1
            klub_lag[klubb].add(lag_no)
            if klubb not in klub_land:
                en = CLUB_COUNTRY.get(klubb, "")
                klub_land[klubb] = LAND_NO.get(en, en) if en else "–"

    sortert = sorted(klub_count.items(), key=lambda x: -x[1])

    result = []
    for rang, (klubb, antall) in enumerate(sortert, 1):
        if antall < 2:
            break
        lag_liste = ", ".join(sorted(klub_lag[klubb]))
        result.append({
            "rang":   rang,
            "klubb":  klubb,
            "land":   klub_land.get(klubb, "–"),
            "antall": antall,
            "lag":    lag_liste,
        })
    return result


def skriv_klubbdominans(wb, klub_data: list[dict]) -> None:
    S = _styles()
    sheet = "Klubbdominans"
    if sheet in wb.sheetnames:
        del wb[sheet]
    ws = wb.create_sheet(sheet)

    col_defs = [
        ("#",       5,  S["ctr"]),
        ("Klubb",  28,  S["lft"]),
        ("Land",   18,  S["lft"]),
        ("Spillere",10, S["ctr"]),
        ("Lag i VM", 80, S["lft"]),
    ]

    _write_title(ws, S, 1, len(col_defs),
                 f"VM 2026 — Klubbdominans  (klubber med 2+ spillere)")
    _write_headers(ws, S, 2, col_defs)

    for i, k in enumerate(klub_data, 1):
        row = i + 2
        ws.row_dimensions[row].height = 17
        bg_override, f_rank = _rank_style(S, i)
        bg = bg_override if bg_override else (S["EVEN"] if i % 2 == 0 else S["ODD"])

        vals = [i, k["klubb"], k["land"], k["antall"], k["lag"]]
        for col, (val, (_, _, al)) in enumerate(zip(vals, col_defs), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = bg; c.border = S["bot"]; c.alignment = al
            c.font = f_rank if (col == 1 and f_rank) else \
                     (S["f_muted"] if col == 5 else S["f_data"])

    _set_widths(ws, col_defs)
    print(f"  'Klubbdominans': {len(klub_data)} klubber skrevet")


# ── Hovedprogram ──────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  VM 2026 — Statistikkark (aldersfordeling)")
    print("=" * 55)

    backup = Path(str(EXCEL_PATH) + ".bak")
    shutil.copy2(EXCEL_PATH, backup)
    print(f"\nBackup: {backup}")

    wb = load_workbook(EXCEL_PATH)

    print("\n[1/2] Bygger aldersfordeling...")
    lag_data = bygg_aldersfordeling()
    skriv_aldersfordeling(wb, lag_data)

    try:
        wb.save(EXCEL_PATH)
        print(f"\nLagret: {EXCEL_PATH}")
    except Exception as e:
        shutil.copy2(backup, EXCEL_PATH)
        sys.exit(f"FEIL — rullet tilbake: {e}")

    print("\nFerdig.")

if __name__ == "__main__":
    main()
