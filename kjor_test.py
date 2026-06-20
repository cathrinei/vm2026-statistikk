#!/usr/bin/env python3
"""
kjor_test.py
Full funksjonalitetstest: kjører oppdater.py --full + 3 manuelle scripts.
Lagrer resultat i "Test av funksjonalitet.xlsx" og skriver test_rapport.txt.

Bruk:
    python kjor_test.py
"""
import hashlib, io, shutil, subprocess, sys, threading, time
from datetime import datetime
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Mangler openpyxl: pip install openpyxl")

PROJECT_DIR  = Path(__file__).parent
EXCEL_PATH   = PROJECT_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
TEST_EXCEL   = PROJECT_DIR / "Test av funksjonalitet.xlsx"
GUARD_PATH   = PROJECT_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx.GUARD"
RAPPORT_PATH = PROJECT_DIR / "test_rapport.txt"
PYTHON       = sys.executable

FORVENTEDE_ARK = (
    [f"Gruppe {x}" for x in "ABCDEFGHIJKL"]
    + ["Alder", "Spillere etter klubbland", "Kort", "Toppscorere", "Assists",
       "Lagstatistikk", "Scoringstidspunkt", "Heatmap", "Klubber", "Nivå 2 og lavere"]
)


# ── Hjelpefunksjoner ──────────────────────────────────────────────────────────

def md5_fil(path: Path) -> str:
    h = hashlib.md5()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _tee_reader(stream, collector):
    for line in iter(stream.readline, ""):
        sys.stdout.write(line)
        sys.stdout.flush()
        collector.append(line)


def kjor_med_tee(script: str, args: list = None) -> tuple[int, list[str], float]:
    """Kjør script, vis output i sanntid og samle linjer. Returnerer (returncode, lines, elapsed)."""
    args = args or []
    start = time.time()
    proc = subprocess.Popen(
        [PYTHON, script] + args,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        cwd=str(PROJECT_DIR),
    )
    lines: list[str] = []
    t = threading.Thread(target=_tee_reader, args=(proc.stdout, lines))
    t.start()
    t.join()
    proc.wait()
    elapsed = time.time() - start
    return proc.returncode, lines, elapsed


def gjenopprett_guard():
    if GUARD_PATH.exists():
        shutil.copy2(GUARD_PATH, EXCEL_PATH)
        print(f"\n  Hoved-Excel gjenopprettet fra GUARD-kopi.")


# ── Validering ────────────────────────────────────────────────────────────────

def valider_excel(size_before: int, mtime_before: float, md5_before: str) -> list[tuple[bool, str]]:
    resultater = []

    def sjekk(ok: bool, tekst: str):
        resultater.append((ok, tekst))

    # 1. Fil finnes
    sjekk(EXCEL_PATH.exists(), "Excel-fil finnes")
    if not EXCEL_PATH.exists():
        return resultater

    stat = EXCEL_PATH.stat()
    size_after  = stat.st_size
    mtime_after = stat.st_mtime
    md5_after   = md5_fil(EXCEL_PATH)

    # 2–4. Fil-metadata
    sjekk(mtime_after > mtime_before, "Fil ble oppdatert (mtime)")
    sjekk(md5_after != md5_before,    "Innhold endret (md5)")
    sjekk(size_after > 50_000,        f"Filstørrelse rimelig (>50 KB, faktisk {size_after:,} bytes)")

    # 5. openpyxl åpner filen
    wb = None
    try:
        wb = load_workbook(EXCEL_PATH, read_only=True)
        sjekk(True, "openpyxl åpner filen")
    except Exception as e:
        sjekk(False, f"openpyxl åpner filen (FEIL: {e})")
        return resultater

    # 6. Ark finnes
    faktiske_ark = wb.sheetnames
    mangler = [a for a in FORVENTEDE_ARK if a not in faktiske_ark]
    if mangler:
        sjekk(False, f"Alle forventede ark finnes — mangler: {', '.join(mangler)}")
    else:
        sjekk(True, f"Alle forventede ark finnes ({len(faktiske_ark)} ark totalt)")

    # 7. Spillertelling = 1248
    FORVENTET_SPILLERE = 1248
    antall = 0
    BLOCKS = [(17, 19, 44), (46, 48, 73), (75, 77, 102), (104, 106, 131)]
    for gruppe in [f"Gruppe {x}" for x in "ABCDEFGHIJKL"]:
        if gruppe not in faktiske_ark:
            continue
        ws = wb[gruppe]
        for _, start_row, end_row in BLOCKS:
            for row in range(start_row, end_row + 1):
                if isinstance(ws.cell(row=row, column=12).value, int):
                    antall += 1
    wb.close()
    sjekk(antall == FORVENTET_SPILLERE,
          f"Antall spillere = {FORVENTET_SPILLERE} (talt: {antall})")

    return resultater, size_after, md5_after


# ── Rapport ───────────────────────────────────────────────────────────────────

def skriv_rapport(
    start_tid: datetime,
    total_elapsed: float,
    status: str,
    script_resultater: list[tuple[str, int, float]],
    valideringer: list[tuple[bool, str]],
    all_output: list[str],
    size_before: int,
    size_after: int,
    md5_endret: bool,
    testkopi_laget: bool,
):
    sep = "=" * 80
    linjer = [
        sep,
        "VM 2026 — Testrapport",
        sep,
        f"Kjørt:        {start_tid.strftime('%d.%m.%Y  %H:%M:%S')}",
        f"Kjøretid:     {int(total_elapsed)} sekunder",
        f"Status:       {status}",
        "",
        "FILINFO",
        f"  Hoved-Excel:  {EXCEL_PATH.name}",
        f"  Testkopi:     {'✓ ' + TEST_EXCEL.name if testkopi_laget else 'IKKE LAGET'}",
        f"  Størrelse:    {size_before:,} → {size_after:,} bytes",
        f"  MD5 endret:   {'ja' if md5_endret else 'nei'}",
        "",
        "SCRIPTS KJØRT",
    ]
    for navn, rc, elapsed in script_resultater:
        ok_str = "✓ OK " if rc == 0 else "✗ FEIL"
        linjer.append(f"  [{ok_str}]  {navn:<35}  ({elapsed:.0f}s)")

    if valideringer:
        linjer += ["", "VALIDERINGER"]
        for ok, tekst in valideringer:
            linjer.append(f"  [{'OK  ' if ok else 'FEIL'}]  {tekst}")

    linjer += ["", "TERMINAL-OUTPUT", "─" * 80]
    linjer += [l.rstrip("\n") for l in all_output]
    linjer.append(sep)

    with open(RAPPORT_PATH, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(linjer) + "\n")

    print(f"\n  Rapport skrevet: {RAPPORT_PATH.name}")


# ── Hovedflyt ─────────────────────────────────────────────────────────────────

def main():
    start_tid     = datetime.now()
    total_start   = time.time()
    all_output    = []
    script_res    = []
    testkopi_laget = False

    print("=" * 60)
    print("  VM 2026 — Full funksjonalitetstest")
    print(f"  {start_tid.strftime('%d.%m.%Y  %H:%M:%S')}")
    print("=" * 60)

    # ── Forsjekk ─────────────────────────────────────────────────────────────
    if not EXCEL_PATH.exists():
        sys.exit(f"FEIL: {EXCEL_PATH.name} finnes ikke — avbryter uten endringer.")

    # ── Pre-snapshot ──────────────────────────────────────────────────────────
    stat_before = EXCEL_PATH.stat()
    size_before  = stat_before.st_size
    mtime_before = stat_before.st_mtime
    md5_before   = md5_fil(EXCEL_PATH)

    # ── GUARD-kopi ────────────────────────────────────────────────────────────
    print(f"\nLager GUARD-kopi ...")
    try:
        shutil.copy2(EXCEL_PATH, GUARD_PATH)
        print(f"  {GUARD_PATH.name} ({size_before:,} bytes)")
    except Exception as e:
        sys.exit(f"FEIL: Kunne ikke lage GUARD-kopi: {e}")

    try:
        # ── Steg 1: oppdater.py --full ────────────────────────────────────────
        print(f"\n{'─'*60}")
        print("  [1/4] oppdater.py --full")
        print(f"{'─'*60}")
        rc, lines, elapsed = kjor_med_tee("oppdater.py", ["--full"])
        all_output += [f"\n{'─'*60}\n[1/4] oppdater.py --full\n{'─'*60}\n"] + lines
        script_res.append(("oppdater.py --full", rc, elapsed))

        if rc != 0:
            gjenopprett_guard()
            valideringer = [(False, "oppdater.py --full feilet — se terminal-output")]
            skriv_rapport(start_tid, time.time() - total_start,
                          "FEIL (oppdater.py returnerte != 0)",
                          script_res, valideringer, all_output,
                          size_before, size_before, False, False)
            sys.exit(1)

        # ── Steg 2–4: manuelle scripts ────────────────────────────────────────
        for idx, script in enumerate(
            ["add_country_sheet.py", "add_clubs_sheet.py", "lag_nivå_tabell.py"],
            start=2
        ):
            print(f"\n{'─'*60}")
            print(f"  [{idx}/4] {script}")
            print(f"{'─'*60}")
            rc_m, lines_m, elapsed_m = kjor_med_tee(script)
            all_output += [f"\n{'─'*60}\n[{idx}/4] {script}\n{'─'*60}\n"] + lines_m
            script_res.append((script, rc_m, elapsed_m))
            if rc_m != 0:
                print(f"\n  ADVARSEL: {script} returnerte {rc_m} — fortsetter likevel.")

        # ── Post-validering ───────────────────────────────────────────────────
        print(f"\n{'─'*60}")
        print("  Validering")
        print(f"{'─'*60}")
        val_resultat = valider_excel(size_before, mtime_before, md5_before)
        if isinstance(val_resultat, tuple):
            valideringer, size_after, md5_after = val_resultat
        else:
            valideringer = val_resultat
            size_after = EXCEL_PATH.stat().st_size if EXCEL_PATH.exists() else size_before
            md5_after  = md5_before

        alle_ok = all(ok for ok, _ in valideringer)
        for ok, tekst in valideringer:
            print(f"  [{'OK  ' if ok else 'FEIL'}]  {tekst}")

        # ── Testkopi ──────────────────────────────────────────────────────────
        print(f"\n  Lager testkopi: {TEST_EXCEL.name} ...")
        try:
            shutil.copy2(EXCEL_PATH, TEST_EXCEL)
            testkopi_laget = True
            print(f"  Ferdig ({TEST_EXCEL.stat().st_size:,} bytes)")
        except Exception as e:
            print(f"  FEIL ved kopiering: {e}")
            valideringer.append((False, f"Testkopi feilet: {e}"))
            alle_ok = False

        # ── Rapport ───────────────────────────────────────────────────────────
        md5_endret   = md5_after != md5_before
        total_elapsed = time.time() - total_start
        status = "OK" if alle_ok else "FEIL (se valideringer)"

        skriv_rapport(
            start_tid, total_elapsed, status,
            script_res, valideringer, all_output,
            size_before, size_after, md5_endret, testkopi_laget,
        )

        print(f"\n{'='*60}")
        print(f"  Status: {status}")
        print(f"  Kjøretid: {int(total_elapsed)}s")
        print(f"  Testkopi: {TEST_EXCEL.name}")
        print(f"  Rapport:  {RAPPORT_PATH.name}")
        print("=" * 60)

        if not alle_ok:
            sys.exit(2)

    except KeyboardInterrupt:
        print("\n\n  Avbrutt av bruker — gjenoppretter hoved-Excel ...")
        gjenopprett_guard()
        sys.exit(130)

    finally:
        if GUARD_PATH.exists():
            GUARD_PATH.unlink()


if __name__ == "__main__":
    main()
