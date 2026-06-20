#!/usr/bin/env python3
"""
add_minutter_column.py
Legger til "Min" (minutter spilt) som kolonne U (21) i gruppearkene.
Kilde: FIFA /topseasonplayerstatistics — ActualMinutesPlayed per spiller.
Kjøres automatisk av oppdater.py ved nye kamper.
"""
import io, re, shutil, sys, unicodedata
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from pathlib import Path
BASE_DIR = Path(__file__).parent

try:
    import requests
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as gcl
except ImportError as e:
    sys.exit(f"Mangler pakke: {e}")

EXCEL_PATH   = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
_FIFA_BASE   = "https://api.fifa.com/api/v3"
_FIFA_SEASON = "285023"
_HEADERS     = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json",
}

GROUPS = [f"Gruppe {x}" for x in "ABCDEFGHIJKL"]
BLOCKS = [(17, 19, 44), (46, 48, 73), (75, 77, 102), (104, 106, 131)]
COL_NAVN = 13   # M
COL_MIN  = 21   # U


_REPLACE = str.maketrans({
    'ø': 'o', 'Ø': 'o', 'æ': 'a', 'Æ': 'a',
    'å': 'a', 'Å': 'a', 'ı': 'i', 'İ': 'i',
    'ğ': 'g', 'Ğ': 'g', 'ş': 's', 'Ş': 's',
    'ç': 'c', 'Ç': 'c', 'ß': 'ss', '-': '', '’': '', "'": '',
})

def _norm(s: str) -> str:
    import unicodedata as _ud
    s = (s or '').lower().translate(_REPLACE)
    s = _ud.normalize('NFD', s)
    s = ''.join(c for c in s if _ud.category(c) != 'Mn')
    return re.sub(r'[^a-z0-9]', '', s)


# Excel-navn → normalisert API-navn (for mismatchar som ikkje løysast av _norm)
_ALIASES: dict[str, str] = {
    "alissonbecker":         "alisson",
    "edersonmilitao":        "ederson",
    "neymar":                "neymarjr",
    "daniloluiz":            "danilo",
    "gleisonbremer":         "bremer",
    "diney":                 "dineyborges",
    "joaopaulofernandes":    "joaopaulo",
    "antoniorudiger":        "antonioruediger",
    "alexandernubel":        "alexandernuebel",
    "pascalgross":           "pascalgross",   # ß→ss gjer at "Groß"→"gross" no
    "idrissagueye":          "idrissaganagueye",
    "jenscastrop":           "castropjens",   # koreansk namneformat
    "mohammadmohebi":        "mohammadmohebbi",
    "ehsanhajsafi":          "ehsanhajisafi",
    "shojaekhalilzadeh":     "shojakhalilzadeh",   # Excel: Shojae / API: Shoja
    "amirmohammadrazzaghinia": "amirmohammadrazaghinia",
    "mostafaziko":           "mostafazico",
    "mohanadlasheen":        "mohanadlashin",
    "meschakelia":           "meschackelia",
    "odiljonhamrobekov":     "odiljonxamrobekov",
    "behruzkarimov":         "behruzjonkarimov",
    "avazbekulmasaliev":     "avazbekulmasaliyev",
    "lawrenceatiizigi":      "lawrenceatizigi",
    "princekwabenaadu":      "princeadu",
    "yoelbarcenas":          "edgaryoelbarcenas",
    "alessandroschopf":      "alessandroschoepf",  # fix: var feil target (mangla 'o')
    "nourbaniatiyah":        "nourbaniateyah",
    "maximilianoaraujo":     "maxiaraujo",
    "marwanattia":           "marawanattia",
    "nicolasgonzalez":       "nicogonzalez",
    # Gruppe B
    "homamalamin":           "homamahmed",          # Homam Al-Amin → HOMAM AHMED
    "alhashmialhussain":     "alhashmialhussein",   # Al-Hussain → ALHUSSEIN
    "tahsinjamshid":         "tahsinmohammed",      # Tahsin Jamshid → TAHSIN MOHAMMED
    "eraycomert":            "eraycoemert",          # ö→oe
    # Gruppe C
    "ademarous":             "adamarous",            # Adem/Adam
    # Gruppe D
    "cammydevlin":           "camerondevlin",        # Cammy = Cameron
    # Gruppe E
    "jurgenlocadia":         "juergenlocadia",       # ü→ue
    # Gruppe F
    "koitakura":             "kouitakura",           # Ko → Kou
    "hadjmahmoud":           "mohamedhadjmahmoud",  # mangler Mohamed-prefiks
    # Gruppe G Egypt
    "mohamedabdelmonem":     "mohamedabdelmoneim",  # monem → moneim
    "elmahdysoliman":        "mahdysoliman",         # mangler El-prefiks
    "mostafashobeir":        "mostafashoubir",       # shobeir → shoubir
    # Gruppe G Iran
    "rouzbehcheshmi":        "roozbehcheshmi",       # Rouzbeh → Roozbeh
    "shahriyarmoghanlou":    "shahriyarmoghanloo",   # moghanlou → moghanloo
    "ariayousefi":           "aryayousefi",           # Aria → Arya
    "danialeiri":            "danialiri",             # eiri → iri
    "hosseinkanaanizadegan": "hosseinkanani",         # forkorta etternamn i API
    # Gruppe H Saudi-Arabia
    "aymanyahya":            "aimanyahya",            # Ayman → AIMAN
    "nawafboushal":          "nawafbuwashl",          # boushal → buwashl
    "hassankadesh":          "hassankadish",          # kadesh → kadish
    "alaaalhejji":           "alaalhajji",            # alaa→ala, hejji→hajji
    "abdullahalhamdan":      "abdullahalhamddan",    # hamdan → hamddan
    "mohammedkanno":         "mohamedkanno",          # Mohammed → Mohamed
    "jehadthakri":           "jehadthikri",           # thakri → thikri
    # Gruppe I Irak
    "ahmedmaknzi":           "ahmedmaknazi",          # maknzi → maknazi
    "zaidismail":            "zaidismael",            # ismail → ismael
    # Gruppe J Algerie
    "yacinetitraoui":        "yassinetitraoui",       # Yacine → Yassine
    # Gruppe J Østerrike — tropp-endringar (verifisert manuelt via draktnr)
    "phillippaffengruber":   "davidaffengruber",      # nr 2: Phillipp → David
    "patrickwiegele":        "florianwiegele",         # nr 12: Patrick → Florian
    "lovroljubicic":         "dejanljubicic",          # nr 19: Lovro → Dejan
    "maximilianprass":       "alexanderprass",         # nr 22: Maximilian → Alexander
    "marcowanner":           "paulwanner",             # nr 24: Marco → Paul
    "leopoldsvoboda":        "michaelsvoboda",         # nr 25: Leopold → Michael
    "phillippmwene":         "phillipmwene",           # dobbel-l → enkel-l
    # Gruppe J Jordan
    "mohammadabuhashish":    "mohammadabuhasheesh",  # hashish → abuhasheesh
    "odehalfakhouri":        "odehfakhoury",          # fakhouri → fakhoury
    "nourbaniattiah":        "nourbaniateyah",        # attiah → ateyah
    "moabualnadi":           "mohammadabualnadi",     # Mo → Mohammad
    "salimobaid":            "saleemobaid",           # Salim → Saleem
    "abdallahalfakhouri":    "abdallahalfakhori",    # fakhouri → fakhori
    "ihsanhaddad":           "ehsanhaddad",           # Ihsan → Ehsan
    # Gruppe K
    "juanfernandoquintero":  "juanquintero",          # fullt namn → kortform i API
    # Gruppe L
    "abdulrahmanbaba":       "babarahman",            # namn-rekkefølgje snudd i API
    "michaelamirmurillo":    "amirmurillo",           # Michael Amir → berre Amir i API
    # Funne via draktnummer (lineup-API)
    "ahmedfathi":            "homamahmed",            # Qatar nr 14 → HOMAM AHMED
    "mohamedalmannai":       "mohamedmanai",          # Al-Mannai → MANAI
    "munirmohamedi":         "munirelkajoui",         # Marokko nr 12, same fornavn
    "derricketiennejr":      "derricketienne",        # Jr. fjerna i API
    "dukelacroix":           "markhuslacroix",        # Duke = Markhus (kallenamn)
    "alejandrozendejas":     "alexzendejas",          # Alejandro → Alex
    "kaku":                  "alejandroromerogamarra", # Kaku = kallenamn
    "mohamedaminebenhamida": "mohamedaminebenhmida",  # Ben Hamida → BEN HMIDA
    "nabilemad":             "nabildonga",            # Egypt nr 18 (same fornavn)
    "denniseckert":          "dennisdargahi",         # Iran nr 24 (same fornavn)
    "robertolopes":          "picolopes",             # Pico = kallenamn for Roberto
    "firasalburaikan":       "ferasalbrikan",         # Firas→FERAS, Al-Buraikan→ALBRIKAN
    "manafyounis":           "munafyounus",           # Manaf→MUNAF, Younis→YOUNUS
    "mohammadabuzrayq":      "mohammadabuzraiq",      # Zrayq → ABUZRAIQ
    "musaaltaamari":         "mousaaltamari",         # Musa→MOUSA, Taamari→TAMARI
    "mohammadaldawoud":      "mohammadaldaoud",       # Dawoud → DAOUD
    "mohammadtaha":          "mohammadabughoush",     # to offisielle namn (Jordan)

}


def hent_minutter() -> dict[str, int]:
    """Henter ActualMinutesPlayed per spiller fra FIFA API."""
    r = requests.get(
        f"{_FIFA_BASE}/topseasonplayerstatistics/season/{_FIFA_SEASON}/topscorers"
        "?count=2000&language=en-GB",
        headers=_HEADERS, timeout=20,
    )
    r.raise_for_status()
    minutter: dict[str, int] = {}
    for p in r.json().get("PlayerStatsList", []):
        pi = p.get("PlayerInfo") or {}
        name_raw = pi.get("PlayerName") or []
        if isinstance(name_raw, list):
            name = next((n["Description"] for n in name_raw if n.get("Locale") == "en-GB"), "")
        else:
            name = str(name_raw)
        mins = p.get("ActualMinutesPlayed")
        if name and mins is not None:
            minutter[_norm(name)] = int(mins)
    return minutter


def skriv_minutter(minutter: dict[str, int]) -> None:
    backup = Path(str(EXCEL_PATH) + ".bak")
    shutil.copy2(EXCEL_PATH, backup)
    wb = load_workbook(EXCEL_PATH)

    NAVY = PatternFill("solid", fgColor="0F2044")
    BLUE = PatternFill("solid", fgColor="1A3C6B")
    bot  = Border(bottom=Side(style="thin", color="E2E8F0"))
    ctr  = Alignment(horizontal="center", vertical="center")

    totalt_treff = 0

    for g in GROUPS:
        if g not in wb.sheetnames:
            continue
        ws = wb[g]
        ws.column_dimensions[gcl(COL_MIN)].width = 6

        for (h_row, s_row, e_row) in BLOCKS:
            # "Min"-header i kolonneoverskriftsraden (h_row+1) — blå, bold
            c = ws.cell(row=h_row + 1, column=COL_MIN, value="Min")
            c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
            c.fill = BLUE
            c.alignment = ctr

            # Spillerrader
            for row in range(s_row, e_row + 1):
                navn = ws.cell(row=row, column=COL_NAVN).value
                k    = _norm(str(navn)) if navn else None
                mins = minutter.get(_ALIASES.get(k, k)) if k else None

                # Arv bakgrunn fra nabocelle (bevarer MINT for scorere)
                ref_fill = ws.cell(row=row, column=COL_NAVN).fill
                bg = PatternFill("solid", fgColor=ref_fill.fgColor.rgb) if ref_fill.fgColor.type == "rgb" else PatternFill("solid", fgColor="FFFFFF")

                c2 = ws.cell(row=row, column=COL_MIN, value=mins)
                c2.font      = Font(name="Calibri", size=10,
                                    color="1A1A2E" if mins else "C0C0C0")
                c2.fill      = bg
                c2.border    = bot
                c2.alignment = ctr
                if mins:
                    totalt_treff += 1

    try:
        wb.save(EXCEL_PATH)
        print(f"  {totalt_treff} spillere med minutter skrevet til 12 gruppeark")
    except Exception as e:
        shutil.copy2(backup, EXCEL_PATH)
        sys.exit(f"FEIL — rullet tilbake: {e}")


def main():
    print("=" * 55)
    print("  VM 2026 — Minutter spilt per spiller")
    print("=" * 55)

    print("\n[1/2] Henter minutter spilt fra FIFA API...")
    minutter = hent_minutter()
    print(f"  {len(minutter)} spillere med data")

    print("\n[2/2] Skriver til gruppearkene...")
    skriv_minutter(minutter)
    print("\nFerdig.")


if __name__ == "__main__":
    main()
