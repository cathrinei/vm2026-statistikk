"""
Genererer docs/historikk_rekorder.html fra historikk_data.json + live VM 2026-data.

Dynamiske seksjoner:
  - Toppscorere: pre-2026 mål fra JSON + 2026-mål fra Excel
  - Flest kamper: pre-2026 kamper fra JSON + 2026-kamper fra kamper_resultater.json
  - Rekordresultater: sjekker om 2026-kamper brøt historiske rekorder
  - Yngste/eldste scorere: pre-2026-baseline fra JSON + alder-ved-scoring beregnet
    fra fødselsdato (player_alder.json) og måldato (avvik_timeline_cache.json).
    Et 2026-mål telles bare hvis antall mål-hendelser i timeline-cachen stemmer
    nøyaktig med det verifiserte måltallet fra Excel (se les_2026_alder_kandidater) —
    dette filtrerer bort kjente datakvalitetsfeil i FIFA sin live-cache
    (f.eks. mål som dukker opp i to ulike kamper, eller mål uten dekning i
    "Scorere og assists"-arket).

Statiske seksjoner (bare fra JSON):
  - Raskeste mål

Kjøres av daglig_oppdatering.yml etter oppdater.py.
"""

import io
import json
import os
import re
import sys
import unicodedata
from datetime import date, datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

BASE_DIR       = Path(__file__).parent
JSON_FIL       = BASE_DIR / "historikk_data.json"
EXCEL_FIL      = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
KAMPER_FIL     = BASE_DIR / "kamper_resultater.json"
TIMELINE_FIL   = BASE_DIR / "avvik_timeline_cache.json"
FODSELSDATO_FIL = BASE_DIR / "player_alder.json"
NAVN_CACHE_FIL = BASE_DIR / "player_names_cache.json"
TEAM_ID_FIL    = BASE_DIR / "team_id_cache.json"
UTDATA         = BASE_DIR / "docs" / "historikk_rekorder.html"

BLOCKS        = [(17, 19, 44), (46, 48, 73), (75, 77, 102), (104, 106, 131)]
COL_NAVN      = 2
COL_MÅL       = 6

# FIFA-eventtyper i avvik_timeline_cache.json som representerer et mål scoret
# av IdPlayer selv (Type 34 = selvmål, telles IKKE som spillerens eget mål).
MÅL_TYPER     = (0, 41)
SELVMÅL_TYPE  = 34


# ── Datainnlesing ──────────────────────────────────────────────────────────────

def les_2026_mål() -> dict[str, int]:
    """Leser mål per spiller fra alle gruppeark i Excel."""
    wb   = load_workbook(EXCEL_FIL, read_only=True, data_only=True)
    mål  = {}
    for ark in wb.sheetnames:
        if not ark.startswith("Gruppe"):
            continue
        ws = wb[ark]
        for _, start, slutt in BLOCKS:
            for r in range(start, slutt + 1):
                navn = ws.cell(row=r, column=COL_NAVN).value
                m    = ws.cell(row=r, column=COL_MÅL).value
                if navn and isinstance(m, int) and m > 0:
                    mål[navn] = mål.get(navn, 0) + m
    wb.close()
    return mål


def les_2026_kamper_per_lag() -> dict[str, int]:
    """Teller antall spilte kamper per lag fra kamper_resultater.json."""
    with open(KAMPER_FIL, encoding="utf-8") as f:
        cache = json.load(f)
    teller: dict[str, int] = {}
    for kamper in cache.values():
        for k in kamper:
            if not k.get("spilt"):
                continue
            for felt in ("hjemme_en", "hjemme"):
                lag = k.get(felt)
                if lag:
                    teller[lag] = teller.get(lag, 0) + 1
                    break
            for felt in ("borte_en", "borte"):
                lag = k.get(felt)
                if lag:
                    teller[lag] = teller.get(lag, 0) + 1
                    break
    return teller


def les_2026_kamper_for_rekorder() -> list[dict]:
    """Returnerer alle spilte 2026-kamper med scoreinformasjon."""
    with open(KAMPER_FIL, encoding="utf-8") as f:
        cache = json.load(f)
    kamper = []
    for gruppe, liste in cache.items():
        for k in liste:
            if not k.get("spilt"):
                continue
            sh = k.get("score_h")
            sa = k.get("score_a")
            if sh is None or sa is None:
                continue
            hjemme = k.get("hjemme") or k.get("hjemme_en", "?")
            borte  = k.get("borte")  or k.get("borte_en", "?")
            dato   = k.get("dato", "")
            kamper.append({
                "kamp":     f"{hjemme} — {borte}",
                "resultat": f"{sh}–{sa}",
                "totalt":   sh + sa,
                "diff":     abs(sh - sa),
                "ar":       int(dato[:4]) if dato else 2026,
            })
    return kamper


def norm(navn: str) -> str:
    """Normaliserer et navn for fuzzy-matching (fjerner aksenter/bindestrek/case)."""
    s = unicodedata.normalize("NFKD", navn)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[-'.]", " ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def parse_alder(alder_str: str) -> tuple[int, int]:
    m = re.match(r"(\d+)\s*år,\s*(\d+)\s*dager", alder_str)
    return (int(m.group(1)), int(m.group(2))) if m else (0, 0)


def format_alder(years: int, days: int) -> str:
    return f"{years} år, {days} dager"


def beregn_alder_ved(fodselsdato: date, ved_dato: date) -> tuple[int, int]:
    """Alder (år, dager) på en gitt dato, robust mot 29. februar-bursdager."""
    try:
        siste_bursdag = date(ved_dato.year, fodselsdato.month, fodselsdato.day)
    except ValueError:
        siste_bursdag = date(ved_dato.year, fodselsdato.month, 28)
    if siste_bursdag > ved_dato:
        try:
            siste_bursdag = date(ved_dato.year - 1, fodselsdato.month, fodselsdato.day)
        except ValueError:
            siste_bursdag = date(ved_dato.year - 1, fodselsdato.month, 28)
    years = siste_bursdag.year - fodselsdato.year
    days = (ved_dato - siste_bursdag).days
    return years, days


def les_2026_alder_kandidater(mål_2026: dict[str, int]) -> tuple[dict, dict]:
    """
    Beregner alder-ved-scoring for hvert VM 2026-mål og returnerer
    (eldste_kandidater, yngste_kandidater): navn → beste 2026-rad for spilleren.

    Bruker kun mål der antall mål-hendelser i timeline-cachen stemmer nøyaktig
    med det verifiserte måltallet fra Excel — se modul-docstring.
    """
    if not (TIMELINE_FIL.exists() and FODSELSDATO_FIL.exists() and TEAM_ID_FIL.exists()):
        return {}, {}

    with open(TIMELINE_FIL, encoding="utf-8") as f:
        timeline = json.load(f)
    with open(FODSELSDATO_FIL, encoding="utf-8") as f:
        fodselsdatoer = json.load(f)
    with open(TEAM_ID_FIL, encoding="utf-8") as f:
        team_id = json.load(f)
    with open(NAVN_CACHE_FIL, encoding="utf-8") as f:
        navn_cache = json.load(f) if NAVN_CACHE_FIL.exists() else {}
    with open(KAMPER_FIL, encoding="utf-8") as f:
        kamper_raw = json.load(f)

    kamp_for_id = {}
    for liste in kamper_raw.values():
        for k in liste:
            if k.get("spilt") and k.get("dato"):
                kamp_for_id[k["id"]] = k

    excel_oppslag = {}
    for navn, mål in mål_2026.items():
        if mål > 0:
            excel_oppslag[norm(navn)] = (navn, mål)

    # Alle mål-hendelser per spiller (med duplikater, for validering mot Excel)
    hendelser_per_spiller: dict[str, list[str]] = {}
    for matchid, events in timeline.items():
        if matchid not in kamp_for_id:
            continue
        for e in events:
            if e.get("Type") in MÅL_TYPER and e.get("IdPlayer"):
                hendelser_per_spiller.setdefault(e["IdPlayer"], []).append(matchid)

    eldste_kandidater: dict[str, dict] = {}
    yngste_kandidater: dict[str, dict] = {}

    for pid, matchids in hendelser_per_spiller.items():
        info = fodselsdatoer.get(pid)
        ascii_navn = info["name"] if info else navn_cache.get(pid)
        if not ascii_navn:
            continue
        match = excel_oppslag.get(norm(ascii_navn))
        if not match:
            continue
        excel_navn, excel_mål = match
        if len(matchids) != excel_mål:
            continue  # timeline stemmer ikke med verifisert måltall — hopp over
        if not info:
            continue
        try:
            fodselsdato = date.fromisoformat(info["birthdate"][:10])
        except ValueError:
            continue

        for matchid in set(matchids):
            kamp = kamp_for_id[matchid]
            try:
                måldato = date.fromisoformat(kamp["dato"])
            except ValueError:
                continue

            hjemme_id = team_id.get(kamp.get("hjemme")) or team_id.get(kamp.get("hjemme_en"))
            borte_id  = team_id.get(kamp.get("borte"))  or team_id.get(kamp.get("borte_en"))
            teamid = next(
                (e["IdTeam"] for e in timeline[matchid]
                 if e.get("Type") in MÅL_TYPER and e.get("IdPlayer") == pid),
                None,
            )
            if teamid == hjemme_id:
                land, motstander = kamp["hjemme"], kamp["borte"]
            elif teamid == borte_id:
                land, motstander = kamp["borte"], kamp["hjemme"]
            else:
                continue

            years, days = beregn_alder_ved(fodselsdato, måldato)
            rad = {
                "spiller": excel_navn,
                "land": land,
                "alder": format_alder(years, days),
                "motstander": motstander,
                "ar": 2026,
                "aktiv_2026": True,
                "_dager": (years, days),
            }

            gjeldende_eldst = eldste_kandidater.get(excel_navn)
            if gjeldende_eldst is None or rad["_dager"] > gjeldende_eldst["_dager"]:
                eldste_kandidater[excel_navn] = rad

            gjeldende_yngst = yngste_kandidater.get(excel_navn)
            if gjeldende_yngst is None or rad["_dager"] < gjeldende_yngst["_dager"]:
                yngste_kandidater[excel_navn] = dict(rad)

    return eldste_kandidater, yngste_kandidater


def slå_sammen_alder_rekord(baseline: list[dict], kandidater: dict, retning: str) -> list[dict]:
    """Slår sammen historisk baseline med 2026-kandidater, dedupliserer per spiller."""
    beste: dict[str, dict] = {}
    for rad in baseline:
        rad = dict(rad)
        rad["_dager"] = parse_alder(rad["alder"])
        beste[rad["spiller"]] = rad
    for navn, kandidat in kandidater.items():
        gjeldende = beste.get(navn)
        if gjeldende is None:
            beste[navn] = kandidat
        else:
            bedre = kandidat["_dager"] > gjeldende["_dager"] if retning == "eldste" \
                else kandidat["_dager"] < gjeldende["_dager"]
            if bedre:
                beste[navn] = kandidat
    resultat = list(beste.values())
    resultat.sort(key=lambda r: r["_dager"], reverse=(retning == "eldste"))
    for r in resultat:
        r.pop("_dager", None)
    return resultat[:10]


def bygg_alder_note(liste: list[dict], fast_intro: str, kilder: str, felles_navn: str = None) -> str:
    nye = [(i, s) for i, s in enumerate(liste, 1) if s.get("aktiv_2026")]
    linjer = [fast_intro]
    if nye:
        innslag = ", ".join(f'{s["spiller"].split()[-1]} (#{i})' for i, s in nye)
        linjer.append(f"VM 2026 bidro med {len(nye)} innslag på topp 10: {innslag}.")
    if felles_navn:
        linjer.append(felles_navn)
    linjer.append(f"Kilder: {kilder}.")
    return " ".join(linjer)


# ── Beregninger ────────────────────────────────────────────────────────────────

def beregn_toppscorere(data: dict, mål_2026: dict) -> list[dict]:
    """Slår sammen historiske mål med 2026-mål og sorterer."""
    resultat = []
    for spiller in data["toppscorere"]:
        navn_2026 = None
        mal_2026  = 0
        if "aktiv_2026_lag" in spiller:
            # prøv direkte navn, eller lag-basert oppslag
            mal_2026 = mål_2026.get(spiller["spiller"], 0)
            navn_2026 = spiller["aktiv_2026_lag"]
        total = spiller["mal_pre2026"] + mal_2026
        resultat.append({**spiller, "mal_2026": mal_2026, "mal_total": total})
    resultat.sort(key=lambda x: (-x["mal_total"], x["spiller"]))
    return resultat


def beregn_flest_kamper(data: dict, kamper_2026: dict) -> list[dict]:
    """Slår sammen historiske kamper med 2026-kamper og sorterer."""
    # Lag-navnkart (norsk) → engelsk nøkkel i kamper_resultater.json
    lag_map = {
        "Argentina": ["Argentina"],
        "Portugal":  ["Portugal"],
        "Frankrike": ["France", "Frankrike"],
        "England":   ["England"],
    }
    resultat = []
    for spiller in data["flest_kamper"]:
        k2026 = 0
        if "aktiv_2026_lag" in spiller:
            lag = spiller["aktiv_2026_lag"]
            kandidater = lag_map.get(lag, [lag])
            k2026 = max((kamper_2026.get(l, 0) for l in kandidater), default=0)
        total = spiller["kamper_pre2026"] + k2026
        resultat.append({**spiller, "kamper_2026": k2026, "kamper_total": total})
    resultat.sort(key=lambda x: (-x["kamper_total"], x["spiller"]))
    return resultat


def beregn_rekordresultater(data: dict, kamper_2026: list) -> tuple[list, list]:
    """Kombinerer historiske rekorder med eventuelle 2026-kamper som brøt dem."""
    hist_score = data["høyest_scorende_historisk"]
    hist_seier = data["størst_seier_historisk"]

    min_totalt = min(r["totalt"] for r in hist_score)
    min_diff   = min(r["diff"]   for r in hist_seier)

    ekstra_score = [k for k in kamper_2026 if k["totalt"] >= min_totalt]
    ekstra_seier = [k for k in kamper_2026 if k["diff"]   >= min_diff]

    alle_score = hist_score + ekstra_score
    alle_seier = hist_seier + ekstra_seier

    alle_score.sort(key=lambda x: (-x["totalt"], x["ar"]))
    alle_seier.sort(key=lambda x: (-x["diff"], x["ar"]))

    # Behold bare kamper på nivå med rekorden (topp 5, kun de med høyest totalt/diff)
    if alle_score:
        grense_score = sorted(set(r["totalt"] for r in alle_score), reverse=True)
        cutoff = grense_score[min(4, len(grense_score)-1)]
        alle_score = [r for r in alle_score if r["totalt"] >= cutoff]
    if alle_seier:
        grense_seier = sorted(set(r["diff"] for r in alle_seier), reverse=True)
        cutoff = grense_seier[0]
        alle_seier = [r for r in alle_seier if r["diff"] >= cutoff]

    return alle_score, alle_seier


# ── HTML-generering ───────────────────────────────────────────────────────────

def rang_liste(sorterte: list, felt: str) -> list[tuple[int, dict]]:
    """Gir (rang, rad) med delte rangeringer."""
    resultat = []
    rang = 1
    for i, rad in enumerate(sorterte):
        if i > 0 and rad[felt] < sorterte[i-1][felt]:
            rang = i + 1
        resultat.append((rang, rad))
    return resultat


def tr_spiller(rang: int, rad: dict, felt_total: str, felt_vis: str = None) -> str:
    aktiv = "aktiv_2026_lag" in rad
    cls   = ' class="row-active"' if aktiv else ""
    badge = '<span class="active-badge">Aktiv</span>' if aktiv else ""
    verdi = rad[felt_total]
    vis   = rad.get(felt_vis, verdi) if felt_vis else verdi
    return (
        f'      <tr{cls}>\n'
        f'        <td class="rank-num">{rang}</td>\n'
        f'        <td class="player-name">{rad["spiller"]}</td>\n'
        f'        <td>{rad["land"]}</td>\n'
        f'        <td class="goals-num center">{vis}</td>\n'
        f'        <td>{rad["turneringer"]}</td>\n'
        f'        <td>{badge}</td>\n'
        f'      </tr>'
    )


def tr_kamp_rekord(rang: int, rad: dict, felt: str) -> str:
    ar_str = str(rad["ar"]) if rad["ar"] != 2026 else '<span class="active-badge" style="background:#0F6E56;">2026</span>'
    return (
        f'      <tr>\n'
        f'        <td class="rank-num">{rang}</td>\n'
        f'        <td class="player-name">{rad["kamp"]}</td>\n'
        f'        <td class="center">{rad["resultat"]}</td>\n'
        f'        <td class="goals-num center">{rad[felt]}</td>\n'
        f'        <td>{ar_str}</td>\n'
        f'      </tr>'
    )


def bygg_toppscorer_note(scorere: list, dato: str) -> str:
    aktive = [s for s in scorere if "aktiv_2026_lag" in s and s["mal_2026"] > 0]
    linjer = []
    for s in aktive:
        linjer.append(f'{s["spiller"]} scorte {s["mal_2026"]} mål i VM 2026 (totalt {s["mal_total"]}).')
    delt_10_pluss = [s["spiller"] for _, s in rang_liste(scorere, "mal_total") if _ >= 10]
    if delt_10_pluss:
        antall = len(delt_10_pluss)
        linjer.append(f'{antall} spillere deler plass 10 og lavere med 10 mål.')
    linjer.append(f'Oppdatert {dato}. Kilder: FIFA.com, Olympics.com, ESPN, NBC Sports.')
    return " ".join(linjer)


def bygg_kamper_note(kamper: list, dato: str) -> str:
    aktive = [s for s in kamper if "aktiv_2026_lag" in s]
    linjer = []
    for s in aktive:
        linjer.append(f'{s["spiller"]} hadde {s["kamper_pre2026"]} kamper inn i 2026 og har nå {s["kamper_total"]}.')
    linjer.append("Cafu er den eneste spilleren som har spilt i tre VM-finaler.")
    linjer.append(f'Oppdatert {dato}. Kilder: FIFA.com, Olympics.com, NBC Sports.')
    return " ".join(linjer)


def generer_html(
    toppscorere: list,
    flest_kamper: list,
    data: dict,
    høyest_score: list,
    størst_seier: list,
    yngste_scorere: list,
    eldste_scorere: list,
) -> str:
    now  = datetime.now(timezone.utc)
    dato_raw = now.strftime("%B %Y")
    dato = f"{now.day}. {dato_raw}".lower()
    dato = dato[0].upper() + dato[1:]

    # Metrics
    top1 = toppscorere[0]
    top1_kamper = flest_kamper[0]
    aktive_scorere = [s for s in toppscorere if "aktiv_2026_lag" in s]
    aktive_navn = ", ".join(s["spiller"].split()[-1] for s in aktive_scorere[:4])

    # Toppscorer-rader
    top_rader = "\n".join(tr_spiller(r, s, "mal_total") for r, s in rang_liste(toppscorere, "mal_total"))

    # Flest kamper-rader — vis "N+" for aktive
    def kamper_vis(rad):
        t = rad["kamper_total"]
        return f"{t}+" if "aktiv_2026_lag" in rad else str(t)

    kamper_rader = []
    for rang, rad in rang_liste(flest_kamper, "kamper_total"):
        aktiv = "aktiv_2026_lag" in rad
        cls   = ' class="row-active"' if aktiv else ""
        badge = '<span class="active-badge">Aktiv</span>' if aktiv else ""
        kamper_rader.append(
            f'      <tr{cls}>\n'
            f'        <td class="rank-num">{rang}</td>\n'
            f'        <td class="player-name">{rad["spiller"]}</td>\n'
            f'        <td>{rad["land"]}</td>\n'
            f'        <td class="goals-num center">{kamper_vis(rad)}</td>\n'
            f'        <td>{rad["turneringer"]}</td>\n'
            f'        <td>{badge}</td>\n'
            f'      </tr>'
        )
    kamper_rader_str = "\n".join(kamper_rader)

    # Yngste scorere
    yngste_rader = []
    for i, s in enumerate(yngste_scorere, 1):
        aktiv = s.get("aktiv_2026")
        cls   = ' class="row-active"' if aktiv else ""
        yngste_rader.append(
            f'      <tr{cls}>\n'
            f'        <td class="rank-num">{i}</td>\n'
            f'        <td class="player-name">{s["spiller"]}</td>\n'
            f'        <td>{s["land"]}</td>\n'
            f'        <td>{s["alder"]}</td>\n'
            f'        <td>{s["motstander"]}</td>\n'
            f'        <td>{s["ar"]}</td>\n'
            f'      </tr>'
        )
    yngste_rader_str = "\n".join(yngste_rader)

    # Eldste scorere
    eldste_rader = []
    for i, s in enumerate(eldste_scorere, 1):
        aktiv = s.get("aktiv_2026")
        cls   = ' class="row-active"' if aktiv else ""
        eldste_rader.append(
            f'      <tr{cls}>\n'
            f'        <td class="rank-num">{i}</td>\n'
            f'        <td class="player-name">{s["spiller"]}</td>\n'
            f'        <td>{s["land"]}</td>\n'
            f'        <td>{s["alder"]}</td>\n'
            f'        <td>{s["motstander"]}</td>\n'
            f'        <td>{s["ar"]}</td>\n'
            f'      </tr>'
        )
    eldste_rader_str = "\n".join(eldste_rader)

    # Notes for yngste/eldste — flagg spillere som finnes på begge lister
    yngste_navn = {s["spiller"] for s in yngste_scorere}
    eldste_navn = {s["spiller"] for s in eldste_scorere}
    felles = yngste_navn & eldste_navn
    felles_note = None
    if felles:
        navn = ", ".join(sorted(felles))
        verb = "er unik" if len(felles) == 1 else "er unike"
        felles_note = f"{navn} {verb} — finnes på begge listene (yngste og eldste)."

    yngste_note = bygg_alder_note(
        yngste_scorere,
        "Pelé er den eneste spilleren i VM-historien som scoret før fylte 18 år.",
        "FIFA.com, beIN Sports, ESPN",
    )
    eldste_note = bygg_alder_note(
        eldste_scorere,
        "Roger Millas rekord fra 1994 er trolig uslåelig — han ble faktisk bragt tilbake av Kamerun etter å ha pensjonert seg.",
        "Opta Analyst, ESPN, FIFA.com",
        felles_navn=felles_note,
    )

    # Rekordresultater
    score_rader = "\n".join(tr_kamp_rekord(r, k, "totalt") for r, k in rang_liste(høyest_score, "totalt"))
    def ar_celle(ar):
        if ar == 2026:
            return '<span class="active-badge" style="background:#0F6E56;">2026</span>'
        return str(ar)

    seier_rader = "\n".join(
        f'      <tr>\n'
        f'        <td class="rank-num">{r}</td>\n'
        f'        <td class="player-name">{k["kamp"]}</td>\n'
        f'        <td class="center">{k["resultat"]}</td>\n'
        f'        <td class="goals-num center">+{k["diff"]}</td>\n'
        f'        <td>{ar_celle(k["ar"])}</td>\n'
        f'      </tr>'
        for r, k in rang_liste(størst_seier, "diff")
    )

    # Raskeste mål
    rask_rader = []
    for i, s in enumerate(data["raskeste_mal"], 1):
        rask_rader.append(
            f'      <tr>\n'
            f'        <td class="rank-num">{i}</td>\n'
            f'        <td class="player-name">{s["spiller"]}</td>\n'
            f'        <td>{s["land"]}</td>\n'
            f'        <td style="font-weight:500;">{s["tid"]}</td>\n'
            f'        <td>{s["motstander"]}</td>\n'
            f'        <td>{s["ar"]}</td>\n'
            f'      </tr>'
        )
    rask_rader_str = "\n".join(rask_rader)

    # Notes
    top_note    = bygg_toppscorer_note(toppscorere, dato)
    kamper_note = bygg_kamper_note(flest_kamper, dato)

    return f"""<!DOCTYPE html>
<html lang="no">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>VM 2026 — All-time rekorder</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

  :root {{
    --bg: #ffffff;
    --bg-secondary: #f5f4f0;
    --text-primary: #2c2c2a;
    --text-secondary: #5f5e5a;
    --border: rgba(0,0,0,0.12);
    --radius: 8px;
    --radius-lg: 12px;
    --active-bg: #E6F1FB;
    --active-color: #185FA5;
  }}

  @media (prefers-color-scheme: dark) {{
    :root {{
      --bg: #1e1e1c;
      --bg-secondary: #2a2a28;
      --text-primary: #e8e7e2;
      --text-secondary: #a0a09a;
      --border: rgba(255,255,255,0.10);
      --active-bg: #0c2a4a;
      --active-color: #85B7EB;
    }}
  }}

  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: var(--bg);
    color: var(--text-primary);
    padding: 32px 24px;
    max-width: 860px;
    margin: 0 auto;
    line-height: 1.5;
  }}

  header {{ margin-bottom: 28px; }}
  header h1 {{ font-size: 20px; font-weight: 500; margin-bottom: 4px; }}
  header p  {{ font-size: 13px; color: var(--text-secondary); }}

  .metrics {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
    gap: 10px;
    margin-bottom: 24px;
  }}

  .metric-card {{
    background: var(--bg-secondary);
    border-radius: var(--radius);
    padding: 14px 16px;
    border: 0.5px solid var(--border);
  }}
  .metric-label {{ font-size: 11px; color: var(--text-secondary); text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 6px; }}
  .metric-value {{ font-size: 26px; font-weight: 500; line-height: 1; margin-bottom: 4px; }}
  .metric-sub   {{ font-size: 12px; color: var(--text-secondary); }}

  .metric-card.blue  .metric-value {{ color: #185FA5; }}
  .metric-card.green .metric-value {{ color: #0F6E56; }}
  .metric-card.amber .metric-value {{ color: #854F0B; }}
  .metric-card.coral .metric-value {{ color: #993C1D; }}

  @media (prefers-color-scheme: dark) {{
    .metric-card.blue  .metric-value {{ color: #85B7EB; }}
    .metric-card.green .metric-value {{ color: #5DCAA5; }}
    .metric-card.amber .metric-value {{ color: #FAC775; }}
    .metric-card.coral .metric-value {{ color: #F0997B; }}
  }}

  .section {{
    background: var(--bg);
    border: 0.5px solid var(--border);
    border-radius: var(--radius-lg);
    padding: 20px;
    margin-bottom: 16px;
  }}
  .section-header {{
    font-size: 14px;
    font-weight: 500;
    color: var(--text-primary);
    margin-bottom: 14px;
    padding-bottom: 10px;
    border-bottom: 0.5px solid var(--border);
  }}
  .section-sub {{
    font-size: 12px;
    color: var(--text-secondary);
    font-weight: 400;
    margin-left: 6px;
  }}

  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  thead tr {{ border-bottom: 0.5px solid var(--border); }}
  th {{
    font-size: 11px;
    font-weight: 500;
    color: var(--text-secondary);
    text-align: left;
    padding: 6px 10px;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    white-space: nowrap;
  }}
  th.center, td.center {{ text-align: center; }}
  td {{
    padding: 9px 10px;
    color: var(--text-primary);
    border-bottom: 0.5px solid var(--border);
    white-space: nowrap;
  }}
  tr:last-child td {{ border-bottom: none; }}
  tbody tr:hover td {{ background: var(--bg-secondary); }}

  .rank-num {{
    font-size: 12px;
    color: var(--text-secondary);
    font-weight: 500;
  }}
  .player-name {{ font-weight: 500; }}
  .goals-num {{
    font-size: 15px;
    font-weight: 500;
    color: var(--text-primary);
  }}
  .row-active td {{ background: var(--active-bg); }}
  .row-active td {{ color: var(--active-color); }}
  .row-active:hover td {{ background: var(--active-bg); filter: brightness(0.97); }}

  .active-badge {{
    display: inline-block;
    background: #185FA5;
    color: #fff;
    font-size: 10px;
    font-weight: 500;
    padding: 2px 7px;
    border-radius: 10px;
    vertical-align: middle;
    white-space: nowrap;
  }}

  .note {{ font-size: 11px; color: var(--text-secondary); margin-top: 14px; line-height: 1.7; }}

  .tab-bar {{
    display: flex;
    gap: 6px;
    margin-bottom: 20px;
  }}
  .tab-btn {{
    padding: 7px 18px;
    border: 0.5px solid var(--border);
    border-radius: 20px;
    background: var(--bg-secondary);
    color: var(--text-secondary);
    font-size: 13px;
    font-family: inherit;
    cursor: pointer;
    transition: background 0.12s, color 0.12s;
  }}
  .tab-btn.aktiv {{
    background: #185FA5;
    color: #fff;
    border-color: #185FA5;
  }}
  @media (prefers-color-scheme: dark) {{
    .tab-btn.aktiv {{ background: #85B7EB; color: #1e1e1c; border-color: #85B7EB; }}
  }}

  @media (max-width: 480px) {{
    body {{ padding: 20px 16px; }}
    header h1 {{ font-size: 17px; }}
    .metric-card {{ padding: 10px 12px; }}
    .metric-value {{ font-size: 22px; }}
    .section {{ padding: 14px; }}
    th, td {{ padding: 6px 7px; font-size: 12px; }}
    .active-badge {{ font-size: 9px; padding: 1px 5px; }}
    .tab-btn {{ font-size: 12px; padding: 6px 14px; }}
  }}
</style>
</head>
<body>

<header>
  <a href="index.html" style="display:inline-flex;align-items:center;gap:5px;font-size:12px;color:var(--text-secondary);text-decoration:none;margin-bottom:10px;padding:4px 0;">← Tilbake til oversikten</a>
  <h1>All-time VM-rekorder</h1>
  <p>Toppscorere, flest kamper, yngste/eldste scorere og rekordresultater &mdash; oppdatert {dato}</p>
</header>

<div class="tab-bar">
  <button class="tab-btn aktiv" onclick="byttFane('person', this)">Personrekorder</button>
  <button class="tab-btn" onclick="byttFane('lag', this)">Lagrekorder</button>
</div>

<div class="metrics">
  <div class="metric-card blue">
    <div class="metric-label">All-time toppscorer</div>
    <div class="metric-value">{top1["spiller"].split()[-1]}</div>
    <div class="metric-sub">{top1["mal_total"]} mål ({top1["turneringer"]})</div>
  </div>
  <div class="metric-card green">
    <div class="metric-label">Flest kamper</div>
    <div class="metric-value">{top1_kamper["spiller"].split()[-1]}</div>
    <div class="metric-sub">{top1_kamper["kamper_total"]}{"+" if "aktiv_2026_lag" in top1_kamper else ""} kamper ({"2026 pågår" if "aktiv_2026_lag" in top1_kamper else top1_kamper["turneringer"]})</div>
  </div>
  <div class="metric-card amber">
    <div class="metric-label">Rekord i ett VM</div>
    <div class="metric-value">13 mål</div>
    <div class="metric-sub">Just Fontaine, Frankrike 1958</div>
  </div>
  <div class="metric-card coral">
    <div class="metric-label">Aktive i VM 2026</div>
    <div class="metric-value">{len(aktive_scorere)}</div>
    <div class="metric-sub">{aktive_navn} (scorere)</div>
  </div>
</div>

<!-- TOPPSCORERE -->
<div class="section" data-category="person">
  <div class="section-header">
    Toppscorere alle tider
    <span class="section-sub">— inkl. VM 2026</span>
  </div>
  <table>
    <thead>
      <tr>
        <th style="width:36px;">#</th>
        <th>Spiller</th>
        <th>Land</th>
        <th class="center" style="width:60px;">Mål</th>
        <th>Turneringer</th>
        <th>VM 2026</th>
      </tr>
    </thead>
    <tbody>
{top_rader}
    </tbody>
  </table>
  <p class="note">
    {top_note}
  </p>
</div>

<!-- FLEST KAMPER -->
<div class="section" data-category="person">
  <div class="section-header">
    Flest kamper alle tider
    <span class="section-sub">— topp 15, inkl. VM 2026</span>
  </div>
  <table>
    <thead>
      <tr>
        <th style="width:36px;">#</th>
        <th>Spiller</th>
        <th>Land</th>
        <th class="center" style="width:70px;">Kamper</th>
        <th>Turneringer</th>
        <th>VM 2026</th>
      </tr>
    </thead>
    <tbody>
{kamper_rader_str}
    </tbody>
  </table>
  <p class="note">
    {kamper_note}
  </p>
</div>

<!-- YNGSTE SCORERE -->
<div class="section" data-category="person">
  <div class="section-header">
    Yngste scorere alle tider
    <span class="section-sub">— topp 10</span>
  </div>
  <table>
    <thead>
      <tr>
        <th style="width:36px;">#</th>
        <th>Spiller</th>
        <th>Land</th>
        <th>Alder ved scoring</th>
        <th>Motstander</th>
        <th>År</th>
      </tr>
    </thead>
    <tbody>
{yngste_rader_str}
    </tbody>
  </table>
  <p class="note">
    {yngste_note}
  </p>
</div>

<!-- ELDSTE SCORERE -->
<div class="section" data-category="person">
  <div class="section-header">
    Eldste scorere alle tider
    <span class="section-sub">— topp 10</span>
  </div>
  <table>
    <thead>
      <tr>
        <th style="width:36px;">#</th>
        <th>Spiller</th>
        <th>Land</th>
        <th>Alder ved scoring</th>
        <th>Motstander</th>
        <th>År</th>
      </tr>
    </thead>
    <tbody>
{eldste_rader_str}
    </tbody>
  </table>
  <p class="note">
    {eldste_note}
  </p>
</div>

<!-- REKORDRESULTATER -->
<div class="section" id="rekordresultater" data-category="lag">
  <div class="section-header">
    Rekordresultater
  </div>

  <p style="font-size:12px;font-weight:500;color:var(--text-secondary);margin-bottom:8px;text-transform:uppercase;letter-spacing:0.04em;">Høyest scorende kamper (flest mål totalt)</p>
  <table style="margin-bottom:20px;">
    <thead>
      <tr>
        <th style="width:36px;">#</th>
        <th>Kamp</th>
        <th class="center" style="width:56px;">Resultat</th>
        <th class="center" style="width:60px;">Totalt</th>
        <th style="width:60px;">År</th>
      </tr>
    </thead>
    <tbody>
{score_rader}
    </tbody>
  </table>

  <p style="font-size:12px;font-weight:500;color:var(--text-secondary);margin-bottom:8px;text-transform:uppercase;letter-spacing:0.04em;">Største seire (målforskjell)</p>
  <table style="margin-bottom:20px;">
    <thead>
      <tr>
        <th style="width:36px;">#</th>
        <th>Kamp</th>
        <th class="center" style="width:56px;">Resultat</th>
        <th class="center" style="width:60px;">+/−</th>
        <th style="width:60px;">År</th>
      </tr>
    </thead>
    <tbody>
{seier_rader}
    </tbody>
  </table>

  <p style="font-size:12px;font-weight:500;color:var(--text-secondary);margin-bottom:8px;text-transform:uppercase;letter-spacing:0.04em;">Raskeste mål i VM-historien</p>
  <table>
    <thead>
      <tr>
        <th style="width:36px;">#</th>
        <th>Spiller</th>
        <th>Land</th>
        <th style="width:70px;">Tid</th>
        <th>Motstander</th>
        <th style="width:60px;">År</th>
      </tr>
    </thead>
    <tbody>
{rask_rader_str}
    </tbody>
  </table>
  <p class="note">
    Şükür scoret i bronsefinalen mot Sør-Korea etter kun 11 sekunder — anerkjent av FIFA og Guinness World Records. Østerrike–Sveits 1954 kalles «Heat Battle in Lausanne» og ble spilt i ekstrem varme. Ungarn dominerte 1954-VM med tre av de fem høyest scorende kampene.
    <br>Kilder: FIFA.com, Olympics.com, Opta Analyst, Guinness World Records.
  </p>
</div>


<script>
function byttFane(kategori, knapp) {{
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('aktiv'));
  knapp.classList.add('aktiv');
  document.querySelectorAll('.section[data-category]').forEach(s => {{
    s.style.display = s.dataset.category === kategori ? '' : 'none';
  }});
}}
</script>

</body>
</html>
"""


def main():
    print("=" * 50)
    print("  Genererer historikk_rekorder.html")
    print("=" * 50)

    with open(JSON_FIL, encoding="utf-8") as f:
        data = json.load(f)

    print("[1/5] Leser 2026-mål fra Excel...")
    mål_2026 = les_2026_mål()
    aktive_mål = {k: v for k, v in mål_2026.items() if v > 0}
    print(f"  {len(aktive_mål)} spillere med mål i 2026")

    print("[2/5] Leser kampresultater...")
    kamper_per_lag  = les_2026_kamper_per_lag()
    kamper_for_rek  = les_2026_kamper_for_rekorder()
    print(f"  {len(kamper_for_rek)} spilte kamper")

    print("[3/5] Beregner rangeringer...")
    toppscorere  = beregn_toppscorere(data, mål_2026)
    flest_kamper = beregn_flest_kamper(data, kamper_per_lag)
    høyest, seier = beregn_rekordresultater(data, kamper_for_rek)

    print("[4/5] Beregner alder ved scoring (yngste/eldste)...")
    eldste_kand, yngste_kand = les_2026_alder_kandidater(mål_2026)
    yngste_scorere = slå_sammen_alder_rekord(data["yngste_scorere"], yngste_kand, "yngste")
    eldste_scorere = slå_sammen_alder_rekord(data["eldste_scorere"], eldste_kand, "eldste")
    print(f"  {len(yngste_kand)} validerte 2026-kandidater (yngste), {len(eldste_kand)} (eldste)")

    print("[5/5] Genererer HTML...")
    html = generer_html(toppscorere, flest_kamper, data, høyest, seier, yngste_scorere, eldste_scorere)
    UTDATA.write_text(html, encoding="utf-8")
    print(f"  → {UTDATA} skrevet ({len(html):,} tegn)")

    # Rapport
    print()
    top5 = toppscorere[:5]
    print("  Toppscorere topp 5:")
    for r, s in rang_liste(top5, "mal_total"):
        aktiv = " (Aktiv)" if "aktiv_2026_lag" in s else ""
        print(f"    {r}. {s['spiller']}: {s['mal_total']} mål{aktiv}")
    print()
    top3_k = flest_kamper[:3]
    print("  Flest kamper topp 3:")
    for r, s in rang_liste(top3_k, "kamper_total"):
        aktiv = "+" if "aktiv_2026_lag" in s else ""
        print(f"    {r}. {s['spiller']}: {s['kamper_total']}{aktiv} kamper")

    print()
    nye_eldste = [(i, s) for i, s in enumerate(eldste_scorere, 1) if s.get("aktiv_2026")]
    if nye_eldste:
        print("  Nye/oppdaterte 2026-innslag i 'Eldste scorere':")
        for r, s in nye_eldste:
            print(f"    #{r}. {s['spiller']}: {s['alder']} mot {s['motstander']}")
    nye_yngste = [(i, s) for i, s in enumerate(yngste_scorere, 1) if s.get("aktiv_2026")]
    if nye_yngste:
        print("  Nye/oppdaterte 2026-innslag i 'Yngste scorere':")
        for r, s in nye_yngste:
            print(f"    #{r}. {s['spiller']}: {s['alder']} mot {s['motstander']}")

    print()
    print("=" * 50)
    print("  Ferdig")
    print("=" * 50)


if __name__ == "__main__":
    main()
