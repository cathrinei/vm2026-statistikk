#!/usr/bin/env python3
"""
add_sluttspill_sheet.py
Lager ark for sluttspillrundene i VM 2026:
  16-delsfinale, 8-delsfinale, Kvartfinaler, Semifinaler, Bronsefinale, Finale

Henter kampdata fra FIFA API. Fremtidige kamper vises med "TBA" som lagnavn.
Cache: sluttspill_cache.json
"""
import io, json, shutil, sys, time
from datetime import datetime
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
BASE_DIR = Path(__file__).parent

try:
    import requests
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    sys.exit(f"Mangler pakke: {e}")

EXCEL_PATH       = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
CACHE_PATH       = BASE_DIR / "sluttspill_cache.json"
TILSKUERE_CACHE  = BASE_DIR / "tilskuere_cache.json"

_FIFA_BASE   = "https://api.fifa.com/api/v3"
_FIFA_COMP   = "17"
_FIFA_SEASON = "285023"
_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json",
}

NORSK = {
    "Mexico": "Mexico", "South Africa": "Sør-Afrika", "South Korea": "Sør-Korea",
    "Korea Republic": "Sør-Korea", "Czechia": "Tsjekkia", "Czech Republic": "Tsjekkia",
    "Canada": "Canada", "Bosnia and Herzegovina": "Bosnia-Hercegovina",
    "Qatar": "Qatar", "Switzerland": "Sveits", "Brazil": "Brasil", "Morocco": "Marokko",
    "Haiti": "Haiti", "Scotland": "Skottland", "United States": "USA", "USA": "USA",
    "Paraguay": "Paraguay", "Australia": "Australia",
    "Türkiye": "Tyrkia", "Turkey": "Tyrkia",
    "Germany": "Tyskland", "Curacao": "Curaçao", "Curaçao": "Curaçao",
    "Ivory Coast": "Elfenbenskysten", "Côte d'Ivoire": "Elfenbenskysten",
    "Ecuador": "Ecuador", "Sweden": "Sverige", "Japan": "Japan",
    "New Zealand": "New Zealand", "Tunisia": "Tunisia", "Spain": "Spania",
    "Cape Verde": "Kapp Verde", "Cabo Verde": "Kapp Verde",
    "Saudi Arabia": "Saudi-Arabia", "Uruguay": "Uruguay",
    "Netherlands": "Nederland", "Belgium": "Belgia", "Egypt": "Egypt", "Jordan": "Jordan",
    "Norway": "Norge", "France": "Frankrike", "Senegal": "Senegal", "Colombia": "Colombia",
    "Argentina": "Argentina", "Panama": "Panama", "England": "England", "Algeria": "Algerie",
    "Croatia": "Kroatia", "Iran": "Iran", "IR Iran": "Iran", "Iraq": "Irak",
    "Portugal": "Portugal", "Ghana": "Ghana", "DR Congo": "DR Kongo", "Congo DR": "DR Kongo",
    "Uzbekistan": "Usbekistan", "Austria": "Østerrike",
    "Serbia": "Serbia", "Costa Rica": "Costa Rica",
}

def til_norsk(name: str) -> str:
    return NORSK.get(name, name)

# FIFA API bruker engelske rundenavn — mapper til norske arknavn
RUNDE_MAP: dict[str, str] = {
    "round of 32":           "16-delsfinale",
    "round of 32 - leg 1":   "16-delsfinale",
    "round of 16":           "8-delsfinale",
    "round of 16 - leg 1":   "8-delsfinale",
    "quarter-final":         "Kvartfinaler",
    "quarter-finals":        "Kvartfinaler",
    "semi-final":            "Semifinaler",
    "semi-finals":           "Semifinaler",
    "third place play-off":  "Bronsefinale",
    "play-off for third place": "Bronsefinale",
    "third place":           "Bronsefinale",
    "final":                 "Finale",
}

RUNDE_ORDER = ["16-delsfinale", "8-delsfinale", "Kvartfinaler", "Semifinaler", "Bronsefinale", "Finale"]

RUNDE_TITLER = {
    "16-delsfinale": "VM 2026 — 16-delsfinale",
    "8-delsfinale":  "VM 2026 — 8-delsfinale",
    "Kvartfinaler":  "VM 2026 — Kvartfinaler",
    "Semifinaler":   "VM 2026 — Semifinaler",
    "Bronsefinale":  "VM 2026 — Bronsefinale",
    "Finale":        "VM 2026 — Finale",
}

GROUP_PREFIXES = {"group a","group b","group c","group d","group e","group f",
                  "group g","group h","group i","group j","group k","group l"}


def _loc(items, locale="en-GB") -> str:
    if not items:
        return ""
    if isinstance(items, list):
        return next((x.get("Description", "") for x in items if x.get("Locale") == locale), "")
    return str(items)


def hent_sluttspill() -> dict[str, list[dict]]:
    """Henter sluttspill-kamper fra FIFA API og returnerer dict per runde."""
    print("  Henter kampdata fra FIFA API...")
    try:
        r = requests.get(
            f"{_FIFA_BASE}/calendar/matches"
            f"?idCompetition={_FIFA_COMP}&idSeason={_FIFA_SEASON}&count=200&language=en-GB",
            headers=_HEADERS, timeout=20
        )
        r.raise_for_status()
        alle = r.json().get("Results", [])
    except Exception as e:
        print(f"  ADVARSEL: Kunne ikke hente FIFA-data: {e}")
        # Prøv cache
        if Path(CACHE_PATH).exists():
            with open(CACHE_PATH, encoding="utf-8") as f:
                return json.load(f)
        return {r: [] for r in RUNDE_ORDER}

    runder: dict[str, list[dict]] = {r: [] for r in RUNDE_ORDER}
    ukjente_runder: set[str] = set()
    antall_sluttspill = 0

    for m in alle:
        # Hopp over gruppekamper
        group_name = _loc(m.get("GroupName")).lower()
        if any(group_name.startswith(p) for p in GROUP_PREFIXES):
            continue

        # Finn rundenavn fra StageName eller GroupName
        stage = _loc(m.get("StageName")).strip()
        runde = RUNDE_MAP.get(stage.lower())
        if not runde:
            runde = RUNDE_MAP.get(group_name)
        if not runde:
            if stage:
                ukjente_runder.add(stage)
            continue

        antall_sluttspill += 1

        home_en = _loc((m.get("Home") or {}).get("TeamName"))
        away_en = _loc((m.get("Away") or {}).get("TeamName"))
        home_no = til_norsk(home_en) if home_en else "TBA"
        away_no = til_norsk(away_en) if away_en else "TBA"

        dato_raw = (m.get("LocalDate") or m.get("Date") or "")[:10]
        spilt    = m.get("MatchStatus") == 0
        score_h  = (m.get("Home") or {}).get("Score") if spilt else None
        score_a  = (m.get("Away") or {}).get("Score") if spilt else None

        # Hent dommer
        officials = m.get("Officials") or []
        dommer = next(
            (_loc(o.get("NameShort")) for o in officials if o.get("OfficialType") == 1),
            ""
        )

        runder[runde].append({
            "id":      m.get("IdMatch", ""),
            "dato":    dato_raw,
            "hjemme":  home_no,
            "borte":   away_no,
            "score_h": score_h,
            "score_a": score_a,
            "spilt":   spilt,
            "dommer":  dommer,
        })

    for runde in runder:
        runder[runde].sort(key=lambda x: x["dato"] or "9999")

    if ukjente_runder:
        print(f"  ADVARSEL: Ukjente rundenavn fra FIFA API: {ukjente_runder}")
        print("            Legg dem til i RUNDE_MAP om nødvendig.")

    print(f"  {antall_sluttspill} sluttspill-kamper hentet fra FIFA API")
    for runde, kamper in runder.items():
        if kamper:
            spilt_n = sum(1 for k in kamper if k["spilt"])
            print(f"    {runde}: {len(kamper)} kamper ({spilt_n} spilt)")

    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(runder, f, ensure_ascii=False, indent=2)

    return runder


def _hent_tilskuere(runder: dict) -> tuple[dict, dict]:
    """Henter tilskuertall og stadionnavn fra cache (deles med gruppearkene)."""
    raw: dict = {}
    if Path(TILSKUERE_CACHE).exists():
        with open(TILSKUERE_CACHE, encoding="utf-8") as f:
            raw = json.load(f)

    # Migrer gammelt format
    cache: dict[str, dict] = {}
    for mid, val in raw.items():
        cache[mid] = val if isinstance(val, dict) else {"att": val, "stadion": None}

    # Hent data for spilte kamper som mangler stadion
    needed = {k["id"] for kamper in runder.values() for k in kamper
              if k.get("spilt") and k.get("id")
              and (k["id"] not in cache or not cache[k["id"]].get("stadion"))}

    for mid in needed:
        try:
            r = requests.get(f"{_FIFA_BASE}/live/football/{mid}",
                             headers=_HEADERS, timeout=10)
            if r.status_code == 200:
                data = r.json() or {}
                att  = data.get("Attendance")
                name_list = (data.get("Stadium") or {}).get("Name") or []
                stadion   = next((n["Description"] for n in name_list
                                  if n.get("Locale") == "en-GB"), "")
                cache[mid] = {"att": int(att) if att else None, "stadion": stadion}
            else:
                cache[mid] = {"att": None, "stadion": ""}
        except Exception:
            cache[mid] = {"att": None, "stadion": ""}
        time.sleep(0.15)

    with open(TILSKUERE_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)

    tilskuere = {mid: v.get("att") if isinstance(v, dict) else v for mid, v in cache.items()}
    stadioner = {mid: (v.get("stadion", "") if isinstance(v, dict) else "") for mid, v in cache.items()}
    return tilskuere, stadioner


def _fmt_dato(dato_str: str) -> str:
    try:
        return datetime.strptime(dato_str[:10], "%Y-%m-%d").strftime("%d.%m")
    except Exception:
        return dato_str or ""

def _fmt_resultat(k: dict) -> str:
    if k["spilt"]:
        h = k["score_h"] if k["score_h"] is not None else "?"
        a = k["score_a"] if k["score_a"] is not None else "?"
        return f"{h} – {a}"
    if k["dato"]:
        try:
            if datetime.strptime(k["dato"][:10], "%Y-%m-%d").date() <= datetime.now().date():
                return "– – –"
        except Exception:
            pass
    return ""

def _tittel(s: str) -> str:
    return " ".join(w.capitalize() for w in (s or "").split())


def skriv_ark(runder: dict, tilskuere: dict, stadioner: dict) -> None:
    S = {
        "NAVY":    PatternFill("solid", fgColor="0F2044"),
        "BLUE":    PatternFill("solid", fgColor="1A3C6B"),
        "PLAYED":  PatternFill("solid", fgColor="EBF8F2"),
        "EVEN":    PatternFill("solid", fgColor="F0F5FB"),
        "ODD":     PatternFill("solid", fgColor="FFFFFF"),
        "GRAY":    PatternFill("solid", fgColor="D9D9D9"),
        "f_title": Font(name="Calibri", bold=True,  size=13, color="FFFFFF"),
        "f_hdr":   Font(name="Calibri", bold=True,  size=10, color="FFFFFF"),
        "f_data":  Font(name="Calibri",              size=10, color="1A1A2E"),
        "f_muted": Font(name="Calibri",              size=10, color="6B7A99"),
        "f_score": Font(name="Calibri", bold=True,  size=10, color="0F5132"),
        "f_tba":   Font(name="Calibri", italic=True, size=10, color="9AA5B4"),
        "ctr":     Alignment(horizontal="center", vertical="center"),
        "lft":     Alignment(horizontal="left",   vertical="center"),
        "bot":     Border(bottom=Side(style="thin", color="E2E8F0")),
    }

    COL_WIDTHS = [8, 20, 10, 20, 12, 24, 18]  # Dato, Hjemme, Res, Borte, Tilskuere, Stadion, Dommer
    HDR_LABELS = ["Dato", "Hjemmelag", "Resultat", "Bortelag", "Tilskuere", "Stadion", "Dommer"]

    backup = Path(str(EXCEL_PATH) + ".bak")
    shutil.copy2(EXCEL_PATH, backup)
    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception:
        raise

    # Finn innsettingspunkt: rett etter Gruppe L
    insert_after = "Gruppe L"
    if insert_after in wb.sheetnames:
        insert_idx = wb.sheetnames.index(insert_after) + 1
    else:
        insert_idx = 12  # fallback

    skrevet = 0
    for runde in RUNDE_ORDER:
        kamper = runder.get(runde, [])

        # Slett gammelt ark (inkl. eventuelle utdaterte navn)
        GAMLE_NAVN = {"16-delsfinale": ["32-delsfinale"]}
        for gammelt in GAMLE_NAVN.get(runde, []):
            if gammelt in wb.sheetnames:
                del wb[gammelt]
        if runde in wb.sheetnames:
            del wb[runde]

        ws = wb.create_sheet(runde, insert_idx)
        insert_idx += 1

        # Kolonnebredder
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ncols = len(HDR_LABELS)

        # Tittelrad
        ws.row_dimensions[1].height = 28
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c = ws.cell(row=1, column=1, value=RUNDE_TITLER[runde])
        c.font = S["f_title"]; c.fill = S["NAVY"]; c.alignment = S["lft"]

        # Kolonneoverskrifter
        ws.row_dimensions[2].height = 20
        for col, label in enumerate(HDR_LABELS, 1):
            c = ws.cell(row=2, column=col, value=label)
            c.font = S["f_hdr"]; c.fill = S["BLUE"]
            c.alignment = S["lft"] if col in (2, 4, 6, 7) else S["ctr"]

        if not kamper:
            # Ingen kamper ennå — tom rad med melding
            ws.row_dimensions[3].height = 17
            c = ws.cell(row=3, column=1, value="Kampene er ikke fastsatt ennå")
            c.font = S["f_tba"]; c.fill = S["EVEN"]
            ws.merge_cells(f"A3:{get_column_letter(ncols)}3")
            c.alignment = S["ctr"]
            skrevet += 1
            continue

        for i, k in enumerate(kamper):
            row = i + 3
            ws.row_dimensions[row].height = 17
            spilt = k["spilt"]
            tba_h = k["hjemme"] == "TBA"
            tba_a = k["borte"]  == "TBA"
            bg    = S["PLAYED"] if spilt else (S["EVEN"] if i % 2 == 0 else S["ODD"])
            res   = _fmt_resultat(k)
            mid   = k.get("id", "")
            att   = tilskuere.get(mid)
            att_str     = f"{att:,}".replace(",", " ") if att else ""
            stadion_str = stadioner.get(mid, "") or ""
            dommer_str  = _tittel(k.get("dommer", "") or "")

            vals = [
                _fmt_dato(k["dato"]) if k["dato"] else "",
                k["hjemme"],
                res,
                k["borte"],
                att_str,
                stadion_str,
                dommer_str,
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill   = bg
                c.border = S["bot"]
                c.alignment = S["lft"] if col in (2, 4, 6, 7) else S["ctr"]
                if col == 3 and spilt:
                    c.font = S["f_score"]
                elif col in (2, 4) and (tba_h or tba_a):
                    c.font = S["f_tba"]
                elif col in (1, 5):
                    c.font = S["f_muted"]
                else:
                    c.font = S["f_data"]

        skrevet += 1

    try:
        wb.save(EXCEL_PATH)
    except Exception:
        shutil.copy2(backup, EXCEL_PATH)
        raise

    print(f"  → {skrevet} sluttspill-ark skrevet til Excel")


def main():
    print("=" * 55)
    print("  VM 2026 — Sluttspill-ark")
    print("=" * 55)

    print("\n[1/3] Henter sluttspill-kamper fra FIFA API...")
    runder = hent_sluttspill()

    totalt_kamper = sum(len(k) for k in runder.values())
    totalt_spilt  = sum(1 for kamper in runder.values() for k in kamper if k["spilt"])

    print(f"\n[2/3] Henter tilskuertall for spilte kamper...")
    tilskuere, stadioner = _hent_tilskuere(runder)

    print(f"\n[3/3] Skriver ark til Excel...")
    skriv_ark(runder, tilskuere, stadioner)

    print(f"\n  Totalt: {totalt_kamper} kamper funnet, {totalt_spilt} spilt")
    print("\nFerdig.")


if __name__ == "__main__":
    main()
