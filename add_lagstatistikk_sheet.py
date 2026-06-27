#!/usr/bin/env python3
"""
add_lagstatistikk_sheet.py
Legger til "Lagstatistikk"-ark med sju seksjoner:
  1. Mål per kamp — snitt per lag + mest målrike enkeltmatcher
  2. Clean sheet — lag med flest kamper uten baklengsmål
  3. Straffespark — tilkjente og scorede straffespark per lag
  4. Selvmål — selvmål fordelt på lag og spiller
  5. Skudd — totale skudd og skudd på mål per lag
  6. Formasjoner — hvilken formasjon hvert lag har brukt
  7. Spillerbytter — bytte-timing per lag (snitt-minutt, tidligste bytte)

Kilde: kamper_resultater.json (seksjon 1+2) + FIFA API timelines (seksjon 3+4)
       + FIFA /topseasonplayerstatistics (seksjon 5, aggregert per lag)
       + FIFA /live/football/{id} (seksjon 6+7, cachet i live_cache.json).
Cache: lagstatistikk_cache.json — lagres for å unngå gjentatte API-kall.
"""
import io, json, shutil, sys, time
from collections import defaultdict
from pathlib import Path
BASE_DIR = Path(__file__).parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    import requests
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter as gcl
except ImportError as e:
    sys.exit(f"Mangler pakke: {e}")

EXCEL_PATH       = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
CACHE_PATH       = BASE_DIR / "kamper_resultater.json"
SLUTTSPILL_CACHE = BASE_DIR / "sluttspill_cache.json"
STAT_CACHE       = BASE_DIR / "lagstatistikk_cache.json"
LIVE_CACHE       = BASE_DIR / "live_cache.json"

_FIFA_BASE   = "https://api.fifa.com/api/v3"
_FIFA_COMP   = "17"
_FIFA_SEASON = "285023"
_HEADERS     = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json",
}

NORSK: dict[str, str] = {
    "Mexico": "Mexico", "South Africa": "Sør-Afrika", "South Korea": "Sør-Korea",
    "Korea Republic": "Sør-Korea", "Republic of Korea": "Sør-Korea",
    "Czechia": "Tsjekkia", "Czech Republic": "Tsjekkia",
    "Canada": "Canada", "Bosnia and Herzegovina": "Bosnia-Hercegovina",
    "Qatar": "Qatar", "Switzerland": "Sveits",
    "Brazil": "Brasil", "Morocco": "Marokko", "Haiti": "Haiti", "Scotland": "Skottland",
    "United States": "USA", "United States of America": "USA", "USA": "USA",
    "Paraguay": "Paraguay", "Australia": "Australia",
    "Turkey": "Tyrkia", "Türkiye": "Tyrkia",
    "Germany": "Tyskland", "Curacao": "Curaçao", "Curaçao": "Curaçao",
    "Ivory Coast": "Elfenbenskysten", "Côte d'Ivoire": "Elfenbenskysten",
    "Ecuador": "Ecuador", "Sweden": "Sverige", "Japan": "Japan",
    "New Zealand": "New Zealand", "Tunisia": "Tunisia",
    "Spain": "Spania", "Cape Verde": "Kapp Verde", "Cabo Verde": "Kapp Verde",
    "Saudi Arabia": "Saudi-Arabia", "Uruguay": "Uruguay",
    "Netherlands": "Nederland", "Belgium": "Belgia", "Egypt": "Egypt", "Jordan": "Jordan",
    "Norway": "Norge", "France": "Frankrike", "Senegal": "Senegal", "Colombia": "Colombia",
    "Argentina": "Argentina", "Panama": "Panama", "England": "England", "Algeria": "Algerie",
    "Croatia": "Kroatia", "Iran": "Iran", "IR Iran": "Iran",
    "Iraq": "Irak", "Portugal": "Portugal",
    "Ghana": "Ghana", "DR Congo": "DR Kongo", "Congo DR": "DR Kongo",
    "Democratic Republic of the Congo": "DR Kongo",
    "Uzbekistan": "Usbekistan", "Austria": "Østerrike",
    "Sør-Korea": "Sør-Korea", "Tsjekkia": "Tsjekkia",
}

def no(name): return NORSK.get(name, name)


# ── Seksjon 1 + 2: Mål og rene nullere fra cache ────────────────────────────

def bygg_maal_og_nullere() -> tuple[list[dict], list[dict], list[dict]]:
    with open(CACHE_PATH, encoding="utf-8") as f:
        gruppe_cache = json.load(f)

    sluttspill_kamper: list = []
    if SLUTTSPILL_CACHE.exists():
        with open(SLUTTSPILL_CACHE, encoding="utf-8") as f:
            for kampeliste in json.load(f).values():
                sluttspill_kamper.extend(kampeliste)

    lag_stats: dict[str, dict] = {}
    kamper: list[dict] = []

    def get_lag(navn):
        if navn not in lag_stats:
            lag_stats[navn] = {"lag": navn, "kamper": 0, "mf": 0, "mm": 0, "nullere": 0}
        return lag_stats[navn]

    # Gruppespill
    for kampeliste in gruppe_cache.values():
        for k in kampeliste:
            if not k.get("spilt"):
                continue
            h = k["hjemme"]; b = k["borte"]
            sh = k["score_h"]; sa = k["score_a"]
            get_lag(h)["kamper"] += 1
            get_lag(h)["mf"]     += sh
            get_lag(h)["mm"]     += sa
            if sa == 0: get_lag(h)["nullere"] += 1
            get_lag(b)["kamper"] += 1
            get_lag(b)["mf"]     += sa
            get_lag(b)["mm"]     += sh
            if sh == 0: get_lag(b)["nullere"] += 1
            kamper.append({"hjemme": h, "borte": b,
                           "score": f"{sh}–{sa}", "totalt": sh + sa, "dato": k.get("dato", "")})

    # Sluttspill
    for k in sluttspill_kamper:
        if not k.get("spilt"):
            continue
        h = k["hjemme"]; b = k["borte"]
        sh = k["score_h"]; sa = k["score_a"]
        if sh is None or sa is None:
            continue
        get_lag(h)["kamper"] += 1
        get_lag(h)["mf"]     += sh
        get_lag(h)["mm"]     += sa
        if sa == 0: get_lag(h)["nullere"] += 1
        get_lag(b)["kamper"] += 1
        get_lag(b)["mf"]     += sa
        get_lag(b)["mm"]     += sh
        if sh == 0: get_lag(b)["nullere"] += 1
        kamper.append({"hjemme": h, "borte": b,
                       "score": f"{sh}–{sa}", "totalt": sh + sa, "dato": k.get("dato", "")})

    for s in lag_stats.values():
        k = s["kamper"] or 1
        s["mf_snitt"] = s["mf"] / k
        s["mm_snitt"] = s["mm"] / k
        s["mf_str"]   = f"{s['mf_snitt']:.2f}".replace(".", ",")
        s["mm_str"]   = f"{s['mm_snitt']:.2f}".replace(".", ",")

    maal_sorted   = sorted(lag_stats.values(), key=lambda x: (-x["mf"], -x["mf_snitt"]))
    nuller_sorted = sorted(lag_stats.values(), key=lambda x: (-x["nullere"], -x["kamper"]))
    kamper_sorted = sorted(kamper, key=lambda x: -x["totalt"])

    return maal_sorted, nuller_sorted, kamper_sorted


# ── Felles: hent og cache timeline-events fra FIFA API ───────────────────────

def _hent_og_cache_events():
    """Henter alle spilte kamper og cacher timeline-events (inkl. spillernavn og minutt).
    Returnerer (played, team_names, stat_cache, gruppe_for_lag, match_info)."""

    stat_cache = {}
    if Path(STAT_CACHE).exists():
        with open(STAT_CACHE, encoding="utf-8") as f:
            stat_cache = json.load(f)

    try:
        r = requests.get(
            f"{_FIFA_BASE}/calendar/matches?idCompetition={_FIFA_COMP}"
            f"&idSeason={_FIFA_SEASON}&count=200&language=en-GB",
            headers=_HEADERS, timeout=20
        )
        r.raise_for_status()
        played = [m for m in r.json().get("Results", [])
                  if m.get("MatchStatus") == 0
                  and "2026-06-11" <= (m.get("LocalDate") or "")[:10] <= "2026-07-19"]
    except Exception as e:
        print(f"  ADVARSEL: Kunne ikke hente kamper: {e}")
        return [], {}, stat_cache, {}, {}

    def _loc(ev):
        tl = ev.get("TypeLocalized") or []
        if isinstance(tl, list):
            return next((x.get("Description","") for x in tl if x.get("Locale") == "en-GB"), "")
        return str(tl)

    def _player_name(ev):
        # Navn hentes fra EventDescription: "Damian BOBADILLA (Paraguay) scores an own goal."
        descs = ev.get("EventDescription") or []
        if isinstance(descs, list):
            desc = next((x.get("Description","") for x in descs if x.get("Locale") == "en-GB"), "")
        else:
            desc = str(descs) if descs else ""
        if desc and " (" in desc:
            return desc.split(" (")[0].strip()
        return ""

    # Bygg lagnavn-oppslag og match_info
    team_names: dict[str, str] = {}
    match_info: dict[str, dict] = {}
    for m in played:
        home_id = m.get("Home", {}).get("IdTeam", "")
        away_id = m.get("Away", {}).get("IdTeam", "")
        for side in ("Home", "Away"):
            tid = m.get(side, {}).get("IdTeam", "")
            if tid and tid not in team_names:
                name_list = m.get(side, {}).get("TeamName", [])
                name = next((n["Description"] for n in name_list if n.get("Locale") == "en-GB"), "")
                team_names[tid] = no(name) if name else tid
        match_info[m["IdMatch"]] = {
            "home": team_names.get(home_id, home_id),
            "away": team_names.get(away_id, away_id),
        }

    # Ugyldiggjør cache-oppføringer der PlayerName mangler selv om IdPlayer finnes
    def _is_stale(evts):
        return any("PlayerName" not in ev or
                   (ev.get("IdPlayer") and not ev.get("PlayerName"))
                   for ev in evts)
    stale = [mid for mid, evts in stat_cache.items() if _is_stale(evts)]
    for mid in stale:
        del stat_cache[mid]

    nye = [m for m in played if m["IdMatch"] not in stat_cache]
    print(f"  {len(stat_cache)} kamper i cache, {len(nye)} nye å hente...")

    for m in played:
        mid = m["IdMatch"]
        if mid not in stat_cache:
            try:
                r2 = requests.get(
                    f"{_FIFA_BASE}/timelines/{mid}?language=en-GB",
                    headers=_HEADERS, timeout=15
                )
                if r2.status_code != 200:
                    stat_cache[mid] = []
                else:
                    events = (r2.json() or {}).get("Event") or []
                    stat_cache[mid] = [
                        {
                            "Type":             ev.get("Type"),
                            "TypeLocalized":    _loc(ev),
                            "IdTeam":           ev.get("IdTeam", ""),
                            "IdPlayer":         ev.get("IdPlayer", ""),
                            "PlayerName":       _player_name(ev),
                            "MatchMinute":      ev.get("MatchMinute", ""),
                            "Period":           ev.get("Period"),
                            "HomeGoals":        ev.get("HomeGoals"),
                            "AwayGoals":        ev.get("AwayGoals"),
                            "HomePenaltyGoals": ev.get("HomePenaltyGoals"),
                            "AwayPenaltyGoals": ev.get("AwayPenaltyGoals"),
                        }
                        for ev in events
                    ]
            except Exception:
                stat_cache[mid] = []
            time.sleep(0.1)

    with open(STAT_CACHE, "w", encoding="utf-8") as f:
        json.dump(stat_cache, f, ensure_ascii=False)

    gruppe_for_lag: dict[str, str] = {}
    with open(CACHE_PATH, encoding="utf-8") as f:
        kamp_cache = json.load(f)
    for gruppe, kamper in kamp_cache.items():
        for k in kamper:
            gruppe_for_lag[k["hjemme"]] = gruppe
            gruppe_for_lag[k["borte"]]  = gruppe

    return played, team_names, stat_cache, gruppe_for_lag, match_info


# ── Seksjon 3: Straffespark ───────────────────────────────────────────────────

# Period-verdier i FIFA API (2026): 3=1.omgang, 5=2.omgang, 6=ET 1.omgang,
# 7=ET 2.omgang, 9=straffesparkkonkurranse, 10=slutt.
# Eksisterende cache-oppføringer uten Period-felt regnes som kampstraffer.
_SHOOTOUT_PERIOD = 9


def _er_shootout(ev: dict) -> bool:
    p = ev.get("Period")
    return p is not None and p >= _SHOOTOUT_PERIOD


def bygg_straffespark(played, team_names, stat_cache) -> tuple[list[dict], list[dict], list[dict]]:
    """Returnerer (lag_kamp, spiller_kamp, spiller_shootout).
    Kampstraffer = straffer i ordinær tid + ekstraomganger (Period < 9).
    Shootout = straffesparkkonkurransen (Period >= 9).
    """
    lag_kamp: dict[str, dict] = defaultdict(lambda: {"tilkjente": 0, "scoret": 0})
    spiller_kamp: dict[str, dict] = defaultdict(lambda: defaultdict(lambda: {"scoret": 0, "brent": 0}))
    spiller_shootout: dict[str, dict] = defaultdict(lambda: defaultdict(lambda: {"scoret": 0, "brent": 0}))

    for m in played:
        mid = m["IdMatch"]
        for ev in stat_cache.get(mid, []):
            loc     = (ev.get("TypeLocalized") or "").lower()
            ev_team = ev.get("IdTeam", "")
            lag_no  = team_names.get(ev_team, ev_team)
            spiller = " ".join(w.capitalize() for w in (ev.get("PlayerName") or "").split())
            shootout = _er_shootout(ev)

            if loc == "penalty awarded" and not shootout:
                lag_kamp[lag_no]["tilkjente"] += 1
            elif loc == "penalty goal":
                if shootout:
                    if spiller:
                        spiller_shootout[lag_no][spiller]["scoret"] += 1
                else:
                    lag_kamp[lag_no]["scoret"] += 1
                    if spiller:
                        spiller_kamp[lag_no][spiller]["scoret"] += 1
            elif loc in ("penalty missed", "penalty saved"):
                if shootout:
                    if spiller:
                        spiller_shootout[lag_no][spiller]["brent"] += 1
                else:
                    if spiller:
                        spiller_kamp[lag_no][spiller]["brent"] += 1

    def _lag_result(lag_dict):
        result = []
        for lag, s in lag_dict.items():
            if s["tilkjente"] == 0:
                continue
            ikke_scoret = s["tilkjente"] - s["scoret"]
            result.append({
                "lag":         lag,
                "tilkjente":   s["tilkjente"],
                "scoret":      s["scoret"],
                "ikke_scoret": ikke_scoret,
                "andel":       f"{s['scoret']}/{s['tilkjente']}",
            })
        return sorted(result, key=lambda x: (-x["tilkjente"], -x["scoret"]))

    def _spiller_result(spiller_dict):
        result = []
        for lag, spillere in spiller_dict.items():
            for spiller, tall in spillere.items():
                tot = tall["scoret"] + tall["brent"]
                result.append({
                    "lag":     lag,
                    "spiller": spiller,
                    "scoret":  tall["scoret"],
                    "brent":   tall["brent"],
                    "forsøkt": tot,
                    "andel":   f"{tall['scoret']}/{tot}",
                })
        return sorted(result, key=lambda x: (-x["forsøkt"], -x["scoret"], x["lag"]))

    return _lag_result(lag_kamp), _spiller_result(spiller_kamp), _spiller_result(spiller_shootout)


# ── Seksjon 4: Selvmål ────────────────────────────────────────────────────────

def bygg_selvmål(played, team_names, stat_cache, match_info) -> list[dict]:
    selvmål = []

    for m in played:
        mid     = m["IdMatch"]
        mi      = match_info.get(mid, {})
        kamp_str = f"{mi.get('home', '')} – {mi.get('away', '')}"

        for ev in stat_cache.get(mid, []):
            loc = (ev.get("TypeLocalized") or "").lower()
            if "own goal" not in loc:
                continue
            ev_team = ev.get("IdTeam", "")
            lag     = team_names.get(ev_team, ev_team)
            spiller = ev.get("PlayerName", "") or "?"
            minutt  = ev.get("MatchMinute", "")
            selvmål.append({
                "lag":     lag,
                "spiller": spiller,
                "kamp":    kamp_str,
                "minutt":  f"{minutt}'" if minutt else "",
            })

    lag_count: dict[str, int] = defaultdict(int)
    for s in selvmål:
        lag_count[s["lag"]] += 1

    selvmål.sort(key=lambda x: (-lag_count[x["lag"]], x["lag"], x["minutt"]))
    return selvmål


# ── Live-data: formasjoner per kamp ─────────────────────────────────────────

def _hent_live_data(played: list, team_names: dict) -> dict:
    """Henter formasjoner per kamp fra FIFA live-API, cacher i live_cache.json."""
    live_cache: dict = {}
    if Path(LIVE_CACHE).exists():
        with open(LIVE_CACHE, encoding="utf-8") as f:
            live_cache = json.load(f)

    nye = [m for m in played if m["IdMatch"] not in live_cache]
    print(f"  {len(live_cache)} kamper i live-cache, {len(nye)} nye å hente...")

    for m in played:
        mid = m["IdMatch"]
        if mid in live_cache:
            continue
        home_id = m.get("Home", {}).get("IdTeam", "")
        away_id = m.get("Away", {}).get("IdTeam", "")
        try:
            r = requests.get(
                f"{_FIFA_BASE}/live/football/{mid}?language=en-GB",
                headers=_HEADERS, timeout=15,
            )
            if r.status_code != 200:
                live_cache[mid] = {}
            else:
                data = r.json() or {}
                live_cache[mid] = {
                    "home_id":      home_id,
                    "away_id":      away_id,
                    "home_tactics": (data.get("HomeTeam") or {}).get("Tactics", ""),
                    "away_tactics": (data.get("AwayTeam") or {}).get("Tactics", ""),
                }
        except Exception:
            live_cache[mid] = {}
        time.sleep(0.1)

    with open(LIVE_CACHE, "w", encoding="utf-8") as f:
        json.dump(live_cache, f, ensure_ascii=False)

    return live_cache


# ── Seksjon 5: Skudd ─────────────────────────────────────────────────────────

def bygg_skudd(team_names: dict, kamper_per_lag: dict) -> list[dict]:
    """Henter spillerstatistikk fra FIFA og aggregerer skudd/skudd-på-mål per lag."""
    try:
        r = requests.get(
            f"{_FIFA_BASE}/topseasonplayerstatistics/season/{_FIFA_SEASON}/topscorers"
            "?count=2000&language=en-GB",
            headers=_HEADERS, timeout=20
        )
        r.raise_for_status()
        players = r.json().get("PlayerStatsList", [])
    except Exception as e:
        print(f"  ADVARSEL: Kunne ikke hente skudddata: {e}")
        return []

    lag_skudd: dict[str, dict] = defaultdict(lambda: {"skudd": 0, "på_mål": 0})

    for p in players:
        pi      = p.get("PlayerInfo") or {}
        team_id = pi.get("IdTeam", "")
        lag     = team_names.get(team_id)
        if not lag:
            continue
        lag_skudd[lag]["skudd"]  += p.get("TotalAttempts")    or 0
        lag_skudd[lag]["på_mål"] += p.get("AttemptsOnTarget") or 0

    result = []
    for lag, s in lag_skudd.items():
        kamper = kamper_per_lag.get(lag, 1)
        skudd  = s["skudd"]
        på_mål = s["på_mål"]
        result.append({
            "lag":         lag,
            "kamper":      kamper,
            "skudd":       skudd,
            "skudd_snitt": f"{skudd/kamper:.1f}".replace(".", ",") if kamper else "0,0",
            "på_mål":      på_mål,
            "treff_pst":   f"{100*på_mål/skudd:.0f}%" if skudd else "–",
        })

    return sorted(result, key=lambda x: (-x["skudd"], -x["på_mål"]))


# ── Seksjon 6: Formasjoner ────────────────────────────────────────────────────

def bygg_formasjoner(played: list, team_names: dict, live_cache: dict) -> list[dict]:
    """Per lag: hvilke formasjoner de har brukt (fra live-API Tactics-felt)."""
    from collections import Counter
    lag_form: dict[str, Counter] = defaultdict(Counter)

    for m in played:
        mid     = m["IdMatch"]
        lc      = live_cache.get(mid) or {}
        home_id = m.get("Home", {}).get("IdTeam", "")
        away_id = m.get("Away", {}).get("IdTeam", "")

        ht = (lc.get("home_tactics") or "").strip()
        at = (lc.get("away_tactics") or "").strip()

        home_lag = team_names.get(home_id)
        away_lag = team_names.get(away_id)
        if home_lag and ht:
            lag_form[home_lag][ht] += 1
        if away_lag and at:
            lag_form[away_lag][at] += 1

    result = []
    for lag, counter in lag_form.items():
        primary = counter.most_common(1)[0][0]
        alle    = " · ".join(
            f"{f} (x{c})" if c > 1 else f for f, c in counter.most_common()
        )
        result.append({
            "lag":       lag,
            "kamper":    sum(counter.values()),
            "primær":    primary,
            "formasjon": alle,
        })

    return sorted(result, key=lambda x: (x["primær"], x["lag"]))


# ── Seksjon 7: Spillerbytter ──────────────────────────────────────────────────

def _parse_min(s: str):
    """'58\'' → 58, '90\'+2\'' → 92"""
    s = (s or "").replace("'", "").strip()
    if "+" in s:
        try:
            a, b = s.split("+")
            return int(a.strip()) + int(b.strip())
        except Exception:
            return None
    try:
        return int(s)
    except Exception:
        return None


def bygg_spillerbytter(played: list, team_names: dict, stat_cache: dict) -> list[dict]:
    """Per lag: total bytter, bytter/kamp, snitt-minutt, tidligste bytte."""
    lag_minutter: dict[str, list] = defaultdict(list)
    lag_kamper:   dict[str, int]  = defaultdict(int)

    for m in played:
        home_id = m.get("Home", {}).get("IdTeam", "")
        away_id = m.get("Away", {}).get("IdTeam", "")
        for tid in (home_id, away_id):
            lag = team_names.get(tid)
            if lag:
                lag_kamper[lag] += 1

    for m in played:
        mid = m["IdMatch"]
        for ev in stat_cache.get(mid, []):
            loc = (ev.get("TypeLocalized") or "").lower()
            if "substitu" not in loc:
                continue
            ev_team = ev.get("IdTeam", "")
            lag     = team_names.get(ev_team)
            if not lag:
                continue
            minutt = _parse_min(ev.get("MatchMinute", ""))
            if minutt is not None:
                lag_minutter[lag].append(minutt)

    result = []
    for lag, minutter in lag_minutter.items():
        kamper = lag_kamper.get(lag, 1)
        tot    = len(minutter)
        snitt  = sum(minutter) / tot if minutter else 0
        tidlig = min(minutter) if minutter else None
        result.append({
            "lag":          lag,
            "kamper":       kamper,
            "bytter":       tot,
            "bytter_snitt": f"{tot/kamper:.1f}".replace(".", ","),
            "snitt_min":    f"{snitt:.0f}",
            "tidligste":    f"{tidlig}'" if tidlig is not None else "–",
            "_snitt_sort":  snitt,
        })

    return sorted(result, key=lambda x: x["_snitt_sort"])


# ── Excel-skriving ────────────────────────────────────────────────────────────

def _styles():
    bot  = Border(bottom=Side(style="thin", color="E2E8F0"))
    ctr  = Alignment(horizontal="center", vertical="center")
    lft  = Alignment(horizontal="left",   vertical="center")
    return {
        "NAVY":    PatternFill("solid", fgColor="0F2044"),
        "BLUE":    PatternFill("solid", fgColor="1A3C6B"),
        "GOLD":    PatternFill("solid", fgColor="FFFBE6"),
        "SILV":    PatternFill("solid", fgColor="F8F8F8"),
        "BRNZ":    PatternFill("solid", fgColor="FFF4EC"),
        "EVEN":    PatternFill("solid", fgColor="F0F5FB"),
        "ODD":     PatternFill("solid", fgColor="FFFFFF"),
        "MINT":    PatternFill("solid", fgColor="EBF8F2"),
        "f_title": Font(name="Calibri", bold=True, size=13, color="FFFFFF"),
        "f_hdr":   Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
        "f_data":  Font(name="Calibri", size=10, color="1A1A2E"),
        "f_muted": Font(name="Calibri", size=10, color="6B7A99"),
        "f_bold":  Font(name="Calibri", bold=True, size=10, color="1A1A2E"),
        "f_rank1": Font(name="Calibri", bold=True, size=10, color="92680A"),
        "f_rank2": Font(name="Calibri", bold=True, size=10, color="5C5C5C"),
        "f_rank3": Font(name="Calibri", bold=True, size=10, color="8B4513"),
        "ctr": ctr, "lft": lft, "bot": bot,
    }

def _title_row(ws, S, row, ncols, text):
    ws.row_dimensions[row].height = 26
    ws.merge_cells(f"A{row}:{gcl(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = S["f_title"]; c.fill = S["NAVY"]; c.alignment = S["lft"]

def _header_row(ws, S, row, col_defs):
    ws.row_dimensions[row].height = 18
    for col, (label, _, al) in enumerate(col_defs, 1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = S["f_hdr"]; c.fill = S["BLUE"]; c.alignment = al

def _data_row(ws, S, row, vals, col_defs, rang=None):
    ws.row_dimensions[row].height = 16
    if rang == 1:   bg, f0 = S["GOLD"], S["f_rank1"]
    elif rang == 2: bg, f0 = S["SILV"], S["f_rank2"]
    elif rang == 3: bg, f0 = S["BRNZ"], S["f_rank3"]
    else:           bg, f0 = (S["EVEN"] if (row % 2 == 0) else S["ODD"]), S["f_muted"]
    for col, (val, (_, _, al)) in enumerate(zip(vals, col_defs), 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = bg; c.border = S["bot"]; c.alignment = al
        c.font = f0 if col == 1 else S["f_data"]


def skriv_lagstatistikk(maal, nullere, kamper, straffe, straffe_spillere, shootout_spillere, selvmål, skudd, formasjoner, bytter):
    S = _styles()
    ncols = 8

    backup = Path(str(EXCEL_PATH) + ".bak")
    shutil.copy2(EXCEL_PATH, backup)
    wb = load_workbook(EXCEL_PATH)

    sheet = "Lagstatistikk"
    if sheet in wb.sheetnames:
        del wb[sheet]
    ws = wb.create_sheet(sheet)

    for col, w in enumerate([5, 18, 9, 8, 14, 9, 8, 9], 1):
        ws.column_dimensions[gcl(col)].width = w

    row = 1

    # ── SEKSJON 1: Mål per kamp ───────────────────────────────────────────────

    maal_defs = [
        ("#",        5,  S["ctr"]),
        ("Lag",     22,  S["lft"]),
        ("Kamper",   8,  S["ctr"]),
        ("Mål for",  8,  S["ctr"]),
        ("MF/kamp",  9,  S["ctr"]),
        ("Mål mot",  8,  S["ctr"]),
        ("MM/kamp",  9,  S["ctr"]),
    ]
    total_maal   = sum(k["totalt"] for k in kamper)
    straffe_maal = sum(s["scoret"] for s in straffe)
    selvmaal_ant = len(selvmål)
    spillermaal  = total_maal - straffe_maal - selvmaal_ant

    _title_row(ws, S, row, ncols, f"VM 2026 — Mål per kamp   {total_maal} mål totalt"); row += 1

    ws.row_dimensions[row].height = 15
    ws.merge_cells(f"A{row}:{gcl(ncols)}{row}")
    c = ws.cell(row=row, column=1,
                value=f"Spillermål: {spillermaal}     Straffer: {straffe_maal}     Selvmål: {selvmaal_ant}")
    c.font      = Font(name="Calibri", size=10, color="1A1A2E")
    c.fill      = PatternFill("solid", fgColor="E8EEF6")
    c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    _header_row(ws, S, row, maal_defs); row += 1

    rang = 1
    for i, lag in enumerate(maal):
        if i > 0 and lag["mf"] != maal[i-1]["mf"]:
            rang = i + 1
        _data_row(ws, S, row,
                  [rang, lag["lag"], lag["kamper"],
                   lag["mf"], lag["mf_str"], lag["mm"], lag["mm_str"]],
                  maal_defs, rang if rang <= 3 else None)
        row += 1

    # Mest målrike enkeltmatcher (topp 10)
    row += 1
    sub_defs = [
        ("#",        5,  S["ctr"]),
        ("Hjemmelag",22, S["lft"]),
        ("Resultat", 8,  S["ctr"]),
        ("Bortelag", 22, S["lft"]),
        ("Dato",     9,  S["ctr"]),
        ("Totalt",   8,  S["ctr"]),
    ]
    _title_row(ws, S, row, ncols, "VM 2026 — Mest målrike enkeltmatcher"); row += 1
    _header_row(ws, S, row, sub_defs); row += 1

    topp = [k for k in kamper if k["totalt"] > 0][:10]
    rang = 1
    for i, k in enumerate(topp):
        if i > 0 and k["totalt"] != topp[i-1]["totalt"]:
            rang = i + 1
        dato_fmt = k["dato"][8:10] + "." + k["dato"][5:7] if len(k.get("dato","")) >= 10 else ""
        _data_row(ws, S, row,
                  [rang, k["hjemme"], k["score"], k["borte"], dato_fmt, k["totalt"]],
                  sub_defs, rang if rang <= 3 else None)
        row += 1

    # ── SEKSJON 2: Clean sheet ────────────────────────────────────────────────

    row += 1
    null_defs = [
        ("#",            5,  S["ctr"]),
        ("Lag",         22,  S["lft"]),
        ("Kamper",       8,  S["ctr"]),
        ("Clean sheet", 13, S["ctr"]),
        ("Andel",        9,  S["ctr"]),
    ]
    _title_row(ws, S, row, ncols, "VM 2026 — Clean sheet"); row += 1
    _header_row(ws, S, row, null_defs); row += 1

    rang = 1
    for i, lag in enumerate(nullere):
        if lag["nullere"] == 0 and i > 0:
            break
        if i > 0 and lag["nullere"] != nullere[i-1]["nullere"]:
            rang = i + 1
        andel = f"{lag['nullere']}/{lag['kamper']}"
        _data_row(ws, S, row,
                  [rang, lag["lag"], lag["kamper"], lag["nullere"], andel],
                  null_defs, rang if rang <= 3 else None)
        row += 1

    # ── SEKSJON 3: Straffespark ───────────────────────────────────────────────

    row += 1
    str_defs = [
        ("#",           5,  S["ctr"]),
        ("Lag",        22,  S["lft"]),
        ("Tilkjente",   9,  S["ctr"]),
        ("Scoret",      8,  S["ctr"]),
        ("Ikke scoret", 11, S["ctr"]),
        ("Andel",       8,  S["ctr"]),
    ]
    str_spiller_defs = [
        ("#",        5,  S["ctr"]),
        ("Spiller", 26,  S["lft"]),
        ("Lag",     22,  S["lft"]),
        ("Scoret",   8,  S["ctr"]),
        ("Forsøkt",  8,  S["ctr"]),
        ("Andel",    8,  S["ctr"]),
    ]

    if straffe:
        _title_row(ws, S, row, ncols, "VM 2026 — Straffespark per lag (ordinær tid + ekstraomganger)"); row += 1
        _header_row(ws, S, row, str_defs); row += 1
        rang = 1
        for i, s in enumerate(straffe):
            if i > 0 and s["tilkjente"] != straffe[i-1]["tilkjente"]:
                rang = i + 1
            _data_row(ws, S, row,
                      [rang, s["lag"], s["tilkjente"], s["scoret"], s["ikke_scoret"], s["andel"]],
                      str_defs, rang if rang <= 3 else None)
            row += 1
        if straffe_spillere:
            row += 1
            _title_row(ws, S, row, ncols, "VM 2026 — Kampstraffer per spiller (ordinær tid + ekstraomganger)"); row += 1
            _header_row(ws, S, row, str_spiller_defs); row += 1
            rang = 1
            for i, s in enumerate(straffe_spillere):
                if i > 0 and s["forsøkt"] != straffe_spillere[i-1]["forsøkt"]:
                    rang = i + 1
                _data_row(ws, S, row,
                          [rang, s["spiller"], s["lag"], s["scoret"], s["forsøkt"], s["andel"]],
                          str_spiller_defs, rang if rang <= 3 else None)
                row += 1
        if shootout_spillere:
            row += 1
            _title_row(ws, S, row, ncols, "VM 2026 — Straffesparkkonkurranse per spiller"); row += 1
            _header_row(ws, S, row, str_spiller_defs); row += 1
            rang = 1
            for i, s in enumerate(shootout_spillere):
                if i > 0 and s["forsøkt"] != shootout_spillere[i-1]["forsøkt"]:
                    rang = i + 1
                _data_row(ws, S, row,
                          [rang, s["spiller"], s["lag"], s["scoret"], s["forsøkt"], s["andel"]],
                          str_spiller_defs, rang if rang <= 3 else None)
                row += 1
    else:
        _title_row(ws, S, row, ncols, "VM 2026 — Straffespark"); row += 1
        ws.cell(row=row, column=1, value="Straffespark: ingen data tilgjengelig fra FIFA API.")
        ws.cell(row=row, column=1).font = S["f_muted"]
        row += 1

    # ── SEKSJON 4: Selvmål ────────────────────────────────────────────────────

    row += 1
    og_defs = [
        ("#",        5,  S["ctr"]),
        ("Lag",     22,  S["lft"]),
        ("Spiller", 26,  S["lft"]),
        ("Kamp",    30,  S["lft"]),
        ("Minutt",   8,  S["ctr"]),
    ]

    _title_row(ws, S, row, ncols, "VM 2026 — Selvmål"); row += 1

    if selvmål:
        _header_row(ws, S, row, og_defs); row += 1

        lag_count: dict[str, int] = defaultdict(int)
        for s in selvmål:
            lag_count[s["lag"]] += 1
        sorted_lags = sorted(lag_count, key=lambda l: -lag_count[l])
        lag_rang = {l: i + 1 for i, l in enumerate(sorted_lags)}

        for s in selvmål:
            r_lag = lag_rang[s["lag"]]
            spiller_fmt = " ".join(w.capitalize() for w in s["spiller"].split())
            _data_row(ws, S, row,
                      [r_lag, s["lag"], spiller_fmt, s["kamp"], s["minutt"]],
                      og_defs, r_lag if r_lag <= 3 else None)
            row += 1
    else:
        c = ws.cell(row=row, column=1, value="Ingen selvmål registrert.")
        c.font = S["f_muted"]
        row += 1

    # ── SEKSJON 5: Skudd ─────────────────────────────────────────────────────────

    row += 1
    skudd_defs = [
        ("#",           5,  S["ctr"]),
        ("Lag",        22,  S["lft"]),
        ("Kamper",      8,  S["ctr"]),
        ("Skudd",       8,  S["ctr"]),
        ("Skudd/kamp",  9,  S["ctr"]),
        ("På mål",      8,  S["ctr"]),
        ("Treff%",      9,  S["ctr"]),
    ]

    _title_row(ws, S, row, ncols, "VM 2026 — Skudd per lag"); row += 1

    if skudd:
        _header_row(ws, S, row, skudd_defs); row += 1
        rang = 1
        for i, s in enumerate(skudd):
            if i > 0 and s["skudd"] != skudd[i-1]["skudd"]:
                rang = i + 1
            _data_row(ws, S, row,
                      [rang, s["lag"], s["kamper"],
                       s["skudd"], s["skudd_snitt"], s["på_mål"], s["treff_pst"]],
                      skudd_defs, rang if rang <= 3 else None)
            row += 1
    else:
        c = ws.cell(row=row, column=1, value="Ingen skudddata tilgjengelig.")
        c.font = S["f_muted"]
        row += 1

    # ── SEKSJON 6: Formasjoner ────────────────────────────────────────────────

    row += 1
    form_defs = [
        ("#",            5,  S["ctr"]),
        ("Lag",         22,  S["lft"]),
        ("Kamper",       8,  S["ctr"]),
        ("Formasjon(er)",14, S["lft"]),
    ]

    _title_row(ws, S, row, ncols, "VM 2026 — Formasjoner per lag"); row += 1

    if formasjoner:
        _header_row(ws, S, row, form_defs); row += 1
        rang = 1
        prev_form = None
        for i, f in enumerate(formasjoner):
            if prev_form is not None and f["primær"] != prev_form:
                rang = i + 1
            _data_row(ws, S, row,
                      [rang, f["lag"], f["kamper"], f["formasjon"]],
                      form_defs, None)
            prev_form = f["primær"]
            row += 1
    else:
        c = ws.cell(row=row, column=1, value="Ingen formasjonsdata tilgjengelig.")
        c.font = S["f_muted"]
        row += 1

    # ── SEKSJON 7: Spillerbytter ──────────────────────────────────────────────

    row += 1
    bytte_defs = [
        ("#",         5,  S["ctr"]),
        ("Lag",      22,  S["lft"]),
        ("Kamper",    8,  S["ctr"]),
        ("Bytter",   14,  S["ctr"]),
        ("B/kamp",    9,  S["ctr"]),
        ("Snitt-min", 8,  S["ctr"]),
        ("Tidligst",  9,  S["ctr"]),
    ]

    _title_row(ws, S, row, ncols, "VM 2026 — Spillerbytte-timing (sortert etter snitt-minutt)"); row += 1

    if bytter:
        _header_row(ws, S, row, bytte_defs); row += 1
        for i, b in enumerate(bytter):
            _data_row(ws, S, row,
                      [i + 1, b["lag"], b["kamper"],
                       b["bytter"], b["bytter_snitt"], b["snitt_min"], b["tidligste"]],
                      bytte_defs, None)
            row += 1
    else:
        c = ws.cell(row=row, column=1, value="Ingen bytte-data tilgjengelig (sjekk TypeLocalized i stat_cache).")
        c.font = S["f_muted"]
        row += 1

    try:
        wb.save(EXCEL_PATH)
        print(f"  → 'Lagstatistikk'-sheet skrevet (rad 1–{row})")
    except Exception as e:
        shutil.copy2(backup, EXCEL_PATH)
        sys.exit(f"FEIL — rullet tilbake: {e}")


# ── Hovedprogram ──────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  VM 2026 — Lagstatistikk-ark")
    print("=" * 55)

    print("\n[1/7] Beregner mål og rene nullere fra cache...")
    maal, nullere, kamper = bygg_maal_og_nullere()
    print(f"  {len(maal)} lag, {len(kamper)} kamper")

    print("\n[2/7] Henter timeline-data fra FIFA API...")
    played, team_names, stat_cache, _, match_info = _hent_og_cache_events()

    print("\n[3/7] Prosesserer straffespark og selvmål...")
    straffe, straffe_spillere, shootout_spillere = bygg_straffespark(played, team_names, stat_cache)
    selvmål = bygg_selvmål(played, team_names, stat_cache, match_info)
    print(f"  {len(straffe)} lag med kampstraffer, {len(straffe_spillere)} spiller(e), "
          f"{len(shootout_spillere)} shootout-spiller(e), {len(selvmål)} selvmål")

    print("\n[4/7] Henter skuddstatistikk fra FIFA API...")
    kamper_per_lag = {m["lag"]: m["kamper"] for m in maal}
    skudd = bygg_skudd(team_names, kamper_per_lag)
    print(f"  {len(skudd)} lag med skudddata")

    print("\n[5/7] Henter formasjonsdata fra FIFA live-API...")
    live_cache   = _hent_live_data(played, team_names)
    formasjoner  = bygg_formasjoner(played, team_names, live_cache)
    print(f"  {len(formasjoner)} lag med formasjonsdata")

    print("\n[6/7] Beregner spillerbytte-timing...")
    bytter = bygg_spillerbytter(played, team_names, stat_cache)
    print(f"  {len(bytter)} lag med bytte-data")

    print("\n[7/7] Skriver Excel-ark...")
    skriv_lagstatistikk(maal, nullere, kamper, straffe, straffe_spillere, shootout_spillere, selvmål, skudd, formasjoner, bytter)

    print("\nFerdig.")


if __name__ == "__main__":
    main()
