#!/usr/bin/env python3
"""
check_avvik.py
Sjekker avvik mellom Excel-data og offisiell VM-statistikk (mål/assist).
Kilde: FIFA API  https://api.fifa.com/api/v3/
"""

import argparse, io, re, shutil, sys, time, unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
BASE_DIR = Path(__file__).parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import requests
except ImportError:
    sys.exit("Mangler requests: pip install requests")
try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Mangler openpyxl: pip install openpyxl")

# ── Konfig ────────────────────────────────────────────────────────────────────

EXCEL_PATH       = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
REPORT_PATH      = BASE_DIR / "avvik_rapport.md"
TIMELINE_CACHE   = BASE_DIR / "avvik_timeline_cache.json"
PLAYER_NAME_CACHE= BASE_DIR / "player_names_cache.json"

_FIFA_BASE   = "https://api.fifa.com/api/v3"
_FIFA_COMP   = "17"
_FIFA_SEASON = "285023"

GROUPS = [f"Gruppe {x}" for x in "ABCDEFGHIJKL"]
BLOCKS = [(1, 3, 28), (30, 32, 57), (59, 61, 86), (88, 90, 115)]

COL_MÅL    = 17
COL_ASSIST = 18
COL_GULT   = 19
COL_RØDT   = 20

# ── Normalisering ─────────────────────────────────────────────────────────────

SPECIAL = str.maketrans({
    "ø":"o","Ø":"o","æ":"ae","Æ":"ae","å":"a","Å":"a",
    "ß":"ss","ð":"d","þ":"th","ı":"i",
})
_JR_RE = re.compile(r"\bjr\.?$", re.IGNORECASE)

def normalize(s: str) -> str:
    s = (s or "").translate(SPECIAL)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[-'']", "", s)
    s = " ".join(s.lower().split())
    s = _JR_RE.sub("junior", s)
    return s

# ── Navnemapping (Excel-navn → FIFA-navn, begge normalisert) ──────────────────
# Legg til nye par ved behov: normalize("Excel-navn"): normalize("FIFA-navn")

NAME_ALIASES: dict[str, str] = {
    normalize(excel): normalize(fifa)
    for excel, fifa in [
        ("Nicolás González",           "Nico Gonzalez"),
        ("Mohammad Mohebi",            "Mohammad Mohebbi"),
        ("Maximiliano Araújo",         "Maxi Araujo"),
        ("Derrick Etienne Jr.",        "Derrick Etienne"),
        ("Alejandro Zendejas",         "Alex Zendejas"),
        ("Cammy Devlin",               "Cameron Devlin"),
        ("Ko Itakura",                 "Kou Itakura"),
        ("Adem Arous",                 "Adam Arous"),
        ("Mohamed Amine Ben Hamida",   "Mohamed Amine Ben Hmida"),
        ("Mostafa Ziko",               "Mostafa Zico"),
        ("Mohanad Lasheen",            "Mohanad Lashin"),
        ("Marwan Attia",               "Marawan Attia"),
        ("Mostafa Shobeir",            "Mostafa Shoubir"),
        ("Ayman Yahya",                "Aiman Yahya"),
        ("Jehad Thakri",               "Jehad Thikri"),
        ("Nawaf Boushal",              "Nawaf Bu Washl"),
        ("Idrissa Gueye",              "Idrissa Gana Gueye"),
        ("Yacine Titraoui",            "Yassine Titraoui"),
        ("Phillipp Mwene",             "Phillip Mwene"),
        ("Ehsan Hajsafi",              "Ehsan Hajisafi"),
        ("Shojae Khalilzadeh",         "Shoja Khalilzadeh"),
        ("Rouzbeh Cheshmi",            "Roozbeh Cheshmi"),
        ("Aria Yousefi",               "Arya Yousefi"),
        ("Amirmohammad Razzaghinia",   "Amirmohammad Razaghinia"),
        ("Shahriyar Moghanlou",        "Shahriyar Moghanloo"),
    ]
}

# ── Lagnavn FIFA engelsk → norsk ──────────────────────────────────────────────

NORSK: dict[str, str] = {
    "Mexico": "Mexico", "South Africa": "Sør-Afrika", "South Korea": "Sør-Korea",
    "Korea Republic": "Sør-Korea", "Czechia": "Tsjekkia", "Czech Republic": "Tsjekkia",
    "Canada": "Canada", "Bosnia and Herzegovina": "Bosnia-Hercegovina",
    "Qatar": "Qatar", "Switzerland": "Sveits",
    "Brazil": "Brasil", "Morocco": "Marokko", "Haiti": "Haiti", "Scotland": "Skottland",
    "United States": "USA", "USA": "USA", "Paraguay": "Paraguay",
    "Australia": "Australia", "Turkey": "Tyrkia", "Türkiye": "Tyrkia",
    "Germany": "Tyskland", "Curacao": "Curaçao", "Curaçao": "Curaçao",
    "Ivory Coast": "Elfenbenskysten", "Côte d'Ivoire": "Elfenbenskysten",
    "Ecuador": "Ecuador", "Sweden": "Sverige", "Japan": "Japan",
    "New Zealand": "New Zealand", "Tunisia": "Tunisia",
    "Spain": "Spania", "Cape Verde": "Kapp Verde", "Cabo Verde": "Kapp Verde",
    "Saudi Arabia": "Saudi-Arabia", "Uruguay": "Uruguay",
    "Netherlands": "Nederland", "Belgium": "Belgia", "Egypt": "Egypt", "Jordan": "Jordan",
    "Norway": "Norge", "France": "Frankrike", "Senegal": "Senegal", "Colombia": "Colombia",
    "Argentina": "Argentina", "Panama": "Panama", "England": "England", "Algeria": "Algerie",
    "Croatia": "Kroatia", "Iran": "Iran", "IR Iran": "Iran",
    "Iraq": "Irak", "Portugal": "Portugal",
    "Ghana": "Ghana", "DR Congo": "DR Kongo", "Congo DR": "DR Kongo",
    "Uzbekistan": "Usbekistan", "Austria": "Østerrike",
}

# ── Hent FIFA-data ────────────────────────────────────────────────────────────

def hent_fifa() -> tuple[list[dict] | None, str]:
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
        "Accept": "application/json",
    }

    # Step 1: Get all played matches (with home/away team IDs)
    try:
        r = requests.get(
            f"{_FIFA_BASE}/calendar/matches?idCompetition={_FIFA_COMP}&idSeason={_FIFA_SEASON}&count=100&language=en-GB",
            headers=headers, timeout=15
        )
        r.raise_for_status()
        played = [
            m for m in r.json().get("Results", [])
            if m.get("MatchStatus") == 0
            and "2026-06-11" <= (m.get("LocalDate") or "")[:10] <= "2026-07-19"
        ]
    except Exception as e:
        return None, f"Feil ved henting av kamper: {e}"

    if not played:
        return None, "Ingen spillte kamper funnet i FIFA API"

    # Step 1b: Build team name lookup from match data
    team_names: dict[str, str] = {}
    for m in played:
        for side in ("Home", "Away"):
            tid = m.get(side, {}).get("IdTeam", "")
            if tid and tid not in team_names:
                name_list = m.get(side, {}).get("TeamName", [])
                name = next((n["Description"] for n in name_list if n.get("Locale") == "en-GB"), "")
                if not name:
                    name = next((n["Description"] for n in name_list), "")
                team_names[tid] = NORSK.get(name, name)

    # Step 2: Parse timelines — les fra cache, hent kun nye kamper
    tl_cache: dict[str, list] = {}
    try:
        from pathlib import Path as _Path
        if _Path(TIMELINE_CACHE).exists():
            with open(TIMELINE_CACHE, encoding="utf-8") as _f:
                tl_cache = json.load(_f)
    except Exception:
        pass

    player_goals:      dict[str, int] = defaultdict(int)
    player_assists:    dict[str, int] = defaultdict(int)
    player_yellow:     dict[str, int] = defaultdict(int)
    player_red:        dict[str, int] = defaultdict(int)
    player_yellow_red: dict[str, int] = defaultdict(int)
    player_team:       dict[str, str] = {}
    player_ids: set[str] = set()
    goal_event_types: dict[int, int] = defaultdict(int)
    all_event_types:  dict[int, int] = defaultdict(int)

    nye_tl = [m for m in played if m["IdMatch"] not in tl_cache]
    print(f"      Timelines: {len(tl_cache)} i cache, {len(nye_tl)} nye å hente")

    for m in played:
        mid     = m["IdMatch"]
        home_id = m.get("Home", {}).get("IdTeam", "")
        away_id = m.get("Away", {}).get("IdTeam", "")

        if mid not in tl_cache:
            try:
                r2 = requests.get(f"{_FIFA_BASE}/timelines/{mid}?language=en-GB",
                                   headers=headers, timeout=15)
                if r2.status_code != 200:
                    tl_cache[mid] = []
                else:
                    raw = (r2.json() or {}).get("Event") or []
                    tl_cache[mid] = [
                        {"Type":       ev.get("Type"),
                         "IdPlayer":   ev.get("IdPlayer"),
                         "IdSubPlayer":ev.get("IdSubPlayer"),
                         "IdTeam":     ev.get("IdTeam", ""),
                         "HomeGoals":  ev.get("HomeGoals"),
                         "AwayGoals":  ev.get("AwayGoals")}
                        for ev in raw
                    ]
            except Exception:
                tl_cache[mid] = []
            time.sleep(0.1)

        prev_home, prev_away = 0, 0
        for ev in tl_cache.get(mid, []):
            ev_type = ev.get("Type")
            pid     = ev.get("IdPlayer")
            ev_team = ev.get("IdTeam", "")

            if ev_type is not None:
                all_event_types[ev_type] += 1

            # Card tracking — Type 2 = Yellow, Type 3 = Red
            if ev_type in (2, 3) and pid:
                if ev_type == 2:
                    player_yellow[pid] += 1
                else:
                    player_red[pid] += 1
                player_ids.add(pid)
                if pid not in player_team and ev_team:
                    player_team[pid] = ev_team

            # Goal tracking (score-change based)
            cur_home = ev.get("HomeGoals") if ev.get("HomeGoals") is not None else prev_home
            cur_away = ev.get("AwayGoals") if ev.get("AwayGoals") is not None else prev_away
            if cur_home <= prev_home and cur_away <= prev_away:
                prev_home, prev_away = cur_home, cur_away
                continue

            goal_event_types[ev_type] += 1
            scoring_team_is_home = cur_home > prev_home
            prev_home, prev_away = cur_home, cur_away

            sub_pid = ev.get("IdSubPlayer") if ev_type == 0 else None

            if scoring_team_is_home and ev_team == away_id:
                continue
            if not scoring_team_is_home and ev_team == home_id:
                continue

            if pid:
                player_goals[pid] += 1
                player_ids.add(pid)
                if pid not in player_team and ev_team:
                    player_team[pid] = ev_team
            if sub_pid:
                player_assists[sub_pid] += 1
                player_ids.add(sub_pid)
                if sub_pid not in player_team and ev_team:
                    player_team[sub_pid] = ev_team

    # Lagre oppdatert timeline-cache
    try:
        import json as _json
        with open(TIMELINE_CACHE, "w", encoding="utf-8") as _f:
            _json.dump(tl_cache, _f, ensure_ascii=False)
    except Exception:
        pass

    print("\n      === Alle event-typer i FIFA API ===")
    kort_typer = {1: "gult?", 2: "rødt?", 3: "gult→rødt?"}
    mål_typer  = {t for t in goal_event_types}
    for t, count in sorted(all_event_types.items()):
        merke = ""
        if t in mål_typer:
            merke = f"  ← MÅL ({goal_event_types[t]} stk)"
        elif t in kort_typer:
            merke = f"  ← {kort_typer[t]} (antatt)"
        print(f"      Type {t:>3}: {count:>4} events{merke}")
    print()

    # Step 3: Hent spillernavn — les fra cache, hent kun ukjente
    player_names: dict[str, str] = {}
    try:
        from pathlib import Path as _Path2
        if _Path2(PLAYER_NAME_CACHE).exists():
            with open(PLAYER_NAME_CACHE, encoding="utf-8") as _f2:
                player_names = json.load(_f2)
    except Exception:
        pass

    nye_pids = [pid for pid in player_ids if pid not in player_names]
    print(f"      Spillernavn: {len(player_names)} i cache, {len(nye_pids)} nye å hente")

    for pid in nye_pids:
        try:
            r3 = requests.get(f"{_FIFA_BASE}/players/{pid}?language=en-GB", headers=headers, timeout=10)
            if r3.status_code == 200:
                name_list = (r3.json() or {}).get("Name", [])
                name = next((n["Description"] for n in name_list if n.get("Locale") == "en-GB"), "")
                player_names[pid] = name.title() if name else pid
        except Exception:
            pass
        time.sleep(0.05)

    # Lagre oppdatert navne-cache
    try:
        with open(PLAYER_NAME_CACHE, "w", encoding="utf-8") as _f3:
            json.dump(player_names, _f3, ensure_ascii=False)
    except Exception:
        pass

    result = []
    kort_result = []
    for pid in player_ids:
        name = player_names.get(pid, pid)
        land = team_names.get(player_team.get(pid, ""), "")
        g  = player_goals.get(pid, 0)
        a  = player_assists.get(pid, 0)
        gy = player_yellow.get(pid, 0)
        gr = player_red.get(pid, 0)
        if g > 0 or a > 0:
            result.append({
                "name":    name,
                "goals":   g,
                "assists": a,
                "key":     normalize(name),
                "lag":     "",
                "land":    land,
            })
        if gy > 0 or gr > 0:
            kort_result.append({
                "name":      name,
                "land":      land,
                "gule":      gy,
                "rode":      gr,
                "gule_rode": 0,
            })

    if not result and not kort_result:
        return None, [], "Ingen spillere funnet i FIFA API"

    return result, kort_result, f"{len(played)} kamper"

# ── Felles designstiler ───────────────────────────────────────────────────────

def _xl_styles():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin  = Side(style="thin",   color="E2E8F0")
    bot   = Border(bottom=thin)
    none_ = Border()
    ctr   = Alignment(horizontal="center", vertical="center")
    lft   = Alignment(horizontal="left",   vertical="center")
    return {
        # Farger
        "NAVY":     PatternFill("solid", fgColor="0F2044"),
        "BLUE":     PatternFill("solid", fgColor="1A3C6B"),
        "GOLD_BG":  PatternFill("solid", fgColor="FFFBE6"),
        "SILV_BG":  PatternFill("solid", fgColor="F8F8F8"),
        "BRNZ_BG":  PatternFill("solid", fgColor="FFF4EC"),
        "EVEN":     PatternFill("solid", fgColor="F0F5FB"),
        "ODD":      PatternFill("solid", fgColor="FFFFFF"),
        "YLW_PILL": PatternFill("solid", fgColor="FFF3CD"),
        "RED_PILL":  PatternFill("solid", fgColor="FDECEA"),
        "PLAYED":   PatternFill("solid", fgColor="EBF8F2"),
        # Fonts
        "f_title":  Font(name="Calibri", bold=True,  size=13, color="FFFFFF"),
        "f_hdr":    Font(name="Calibri", bold=True,  size=10, color="FFFFFF"),
        "f_data":   Font(name="Calibri",              size=10, color="1A1A2E"),
        "f_muted":  Font(name="Calibri",              size=10, color="6B7A99"),
        "f_rank1":  Font(name="Calibri", bold=True,  size=10, color="92680A"),
        "f_rank2":  Font(name="Calibri", bold=True,  size=10, color="5C5C5C"),
        "f_rank3":  Font(name="Calibri", bold=True,  size=10, color="8B4513"),
        "f_ylw":    Font(name="Calibri", bold=True,  size=10, color="856404"),
        "f_red":    Font(name="Calibri", bold=True,  size=10, color="842029"),
        # Alignment
        "ctr": ctr, "lft": lft,
        # Borders
        "bot":  bot, "none": none_,
    }

# ── Skriv kortsheet ───────────────────────────────────────────────────────────

def skriv_kort_sheet(kort: list[dict]) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    S = _xl_styles()
    backup = EXCEL_PATH + ".bak"
    shutil.copy2(EXCEL_PATH, backup)
    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception:
        raise

    if "Kort" in wb.sheetnames:
        del wb["Kort"]
    ws = wb.create_sheet("Kort")

    # Tittelrad
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:F1")
    sortert = sorted(kort, key=lambda x: (-(x["rode"] + x["gule_rode"]), -x["gule"], x["name"]))
    c = ws.cell(row=1, column=1, value=f"VM 2026 — Kort  ({len(sortert)} spillere)")
    c.font = S["f_title"]; c.fill = S["NAVY"]; c.alignment = S["lft"]

    # Kolonneoverskrifter
    ws.row_dimensions[2].height = 20
    for col, label in enumerate(["#", "Spiller", "Lag", "Gule", "Røde", "Tot. røde"], 1):
        c = ws.cell(row=2, column=col, value=label)
        c.font = S["f_hdr"]; c.fill = S["BLUE"]
        c.alignment = S["lft"] if col == 2 else S["ctr"]

    # Beregn rang: lik rang når både røde og gule er identiske
    ranger, cur = [], 1
    for i, p in enumerate(sortert):
        tot = p["rode"] + p["gule_rode"]
        if i > 0:
            prev = sortert[i - 1]
            prev_tot = prev["rode"] + prev["gule_rode"]
            ranger.append(ranger[-1] if tot == prev_tot and p["gule"] == prev["gule"] else cur)
        else:
            ranger.append(cur)
        cur = i + 2

    for i, (p, rang) in enumerate(zip(sortert, ranger)):
        row  = i + 3
        totalt_rode = p["rode"] + p["gule_rode"]
        ws.row_dimensions[row].height = 17

        if rang == 1:
            bg, f_rank = S["GOLD_BG"], S["f_rank1"]
        elif rang == 2:
            bg, f_rank = S["SILV_BG"], S["f_rank2"]
        elif rang == 3:
            bg, f_rank = S["BRNZ_BG"], S["f_rank3"]
        else:
            bg, f_rank = (S["EVEN"] if i % 2 == 0 else S["ODD"]), S["f_muted"]

        data = [rang, p["name"], p["land"], p["gule"] or None, p["rode"] or None, totalt_rode or None]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = S["bot"]
            c.alignment = S["lft"] if col == 2 else S["ctr"]
            c.fill = bg
            c.font = S["f_data"]
            if col == 1:
                c.font = f_rank
            elif col == 4 and val:
                c.fill = S["YLW_PILL"]; c.font = S["f_ylw"]
            elif col in (5, 6) and val:
                c.fill = S["RED_PILL"]; c.font = S["f_red"]

    for col, w in enumerate([5, 28, 20, 8, 8, 10], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    try:
        wb.save(EXCEL_PATH)
    except Exception:
        shutil.copy2(backup, EXCEL_PATH)
        raise
    print(f"      → 'Kort'-sheet skrevet ({len(sortert)} spillere)")

# ── Les Excel ─────────────────────────────────────────────────────────────────

def les_excel() -> list[dict]:
    wb = load_workbook(EXCEL_PATH, data_only=True)
    spillere = []
    for g in GROUPS:
        ws = wb[g]
        for header_row, start, end in BLOCKS:
            lag = (ws.cell(row=header_row, column=12).value or "").strip()
            for r in range(start, end + 1):
                name = ws.cell(row=r, column=13).value
                if not name:
                    continue
                def _int(v):
                    try: return int(v or 0)
                    except (TypeError, ValueError): return 0

                maal  = ws.cell(row=r, column=COL_MÅL).value
                ast   = ws.cell(row=r, column=COL_ASSIST).value
                gult  = ws.cell(row=r, column=COL_GULT).value
                rødt  = ws.cell(row=r, column=COL_RØDT).value
                spillere.append({
                    "name":    name.strip(),
                    "lag":     lag,
                    "gruppe":  g,
                    "goals":   _int(maal),
                    "assists": _int(ast),
                    "gule":    _int(gult),
                    "rode":    _int(rødt),
                    "key":     normalize(name),
                    "_sheet":  g,
                    "_row":    r,
                })
    return spillere

# ── Sammenlign ────────────────────────────────────────────────────────────────

def sammenlign(excel: list[dict], fifa: list[dict], kort: list[dict]) -> list[dict]:
    avvik = []
    idx = {p["key"]: p for p in fifa}

    matched_keys: set[str] = set()

    for ep in excel:
        fp = idx.get(ep["key"])
        if fp is None:
            fp = idx.get(NAME_ALIASES.get(ep["key"], ""))
        if fp is None:
            fp = idx.get(" ".join(sorted(ep["key"].split())))

        if fp:
            matched_keys.add(fp["key"])

        if fp is None and (ep["goals"] > 0 or ep["assists"] > 0):
            avvik.append({"type": "mangler_i_fifa", "name": ep["name"],
                          "lag": ep["lag"], "land": "", "gruppe": ep["gruppe"],
                          "detalj": f"Excel: {ep['goals']} mål, {ep['assists']} assist — ikke godkjent av FIFA, nullstilles",
                          "_sheet": ep["_sheet"], "_row": ep["_row"],
                          "_col": COL_MÅL, "_ny_verdi": 0})
            if ep["assists"] > 0:
                avvik.append({"type": "mangler_i_fifa", "name": ep["name"],
                              "lag": ep["lag"], "land": "", "gruppe": ep["gruppe"],
                              "detalj": f"",  # allerede rapportert over
                              "_sheet": ep["_sheet"], "_row": ep["_row"],
                              "_col": COL_ASSIST, "_ny_verdi": 0})
        elif fp:
            if ep["goals"] != fp["goals"]:
                avvik.append({"type": "mål_avvik", "name": ep["name"],
                              "lag": ep["lag"], "land": fp.get("land", ""), "gruppe": ep["gruppe"],
                              "detalj": f"Mål: Excel={ep['goals']}, FIFA={fp['goals']}",
                              "_sheet": ep["_sheet"], "_row": ep["_row"],
                              "_col": COL_MÅL, "_ny_verdi": fp["goals"]})
            if ep["assists"] != fp["assists"]:
                avvik.append({"type": "assist_avvik", "name": ep["name"],
                              "lag": ep["lag"], "land": fp.get("land", ""), "gruppe": ep["gruppe"],
                              "detalj": f"Assist: Excel={ep['assists']}, FIFA={fp['assists']}",
                              "_sheet": ep["_sheet"], "_row": ep["_row"],
                              "_col": COL_ASSIST, "_ny_verdi": fp["assists"]})

    excel_keys = {p["key"] for p in excel}
    for fp in fifa:
        if (fp["goals"] > 0 or fp["assists"] > 0) and fp["key"] not in excel_keys and fp["key"] not in matched_keys:
            avvik.append({"type": "mangler_i_excel", "name": fp["name"],
                          "lag": fp["lag"], "land": fp.get("land", ""), "gruppe": "?",
                          "detalj": f"FIFA: {fp['goals']} mål, {fp['assists']} assist — mangler i Excel"})

    # Kortsammenligning
    kort_idx = {normalize(k["name"]): k for k in kort}
    for ep in excel:
        kp = kort_idx.get(ep["key"]) or kort_idx.get(NAME_ALIASES.get(ep["key"], ""))
        fifa_gule  = kp["gule"] if kp else 0
        fifa_røde  = (kp["rode"] + kp["gule_rode"]) if kp else 0
        if ep["gule"] != fifa_gule:
            avvik.append({"type": "gult_avvik", "name": ep["name"],
                          "lag": ep["lag"], "land": kp.get("land", "") if kp else "", "gruppe": ep["gruppe"],
                          "detalj": f"Gule: Excel={ep['gule']}, FIFA={fifa_gule}",
                          "_sheet": ep["_sheet"], "_row": ep["_row"],
                          "_col": COL_GULT, "_ny_verdi": fifa_gule})
        if ep["rode"] != fifa_røde:
            avvik.append({"type": "rødt_avvik", "name": ep["name"],
                          "lag": ep["lag"], "land": kp.get("land", "") if kp else "", "gruppe": ep["gruppe"],
                          "detalj": f"Røde: Excel={ep['rode']}, FIFA={fifa_røde}",
                          "_sheet": ep["_sheet"], "_row": ep["_row"],
                          "_col": COL_RØDT, "_ny_verdi": fifa_røde})
    return avvik

# ── Toppscorere / Assists sheet ───────────────────────────────────────────────

def skriv_toppscore_sheets(excel: list[dict]) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    S = _xl_styles()
    backup = EXCEL_PATH + ".bak"
    shutil.copy2(EXCEL_PATH, backup)
    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception:
        raise

    col_defs = [
        ("#",       5,  S["ctr"]),
        ("Spiller", 28, S["lft"]),
        ("Lag",     20, S["lft"]),
        ("Mål",     7,  S["ctr"]),
        ("Assist",  8,  S["ctr"]),
        ("Gule",    7,  S["ctr"]),
        ("Røde",    7,  S["ctr"]),
    ]
    ncols = len(col_defs)

    def _rang(rows, rank_key):
        rang, cur = [], 1
        for i, p in enumerate(rows):
            rang.append(rang[-1] if i > 0 and rank_key(p) == rank_key(rows[i-1]) else cur)
            cur = i + 2
        return rang

    def _write_sheet(name, rows, title, rank_key):
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)

        # Tittelrad
        ws.row_dimensions[1].height = 28
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c = ws.cell(row=1, column=1, value=title)
        c.font = S["f_title"]; c.fill = S["NAVY"]; c.alignment = S["lft"]

        # Kolonneoverskrifter
        ws.row_dimensions[2].height = 20
        for col, (label, _, al) in enumerate(col_defs, 1):
            c = ws.cell(row=2, column=col, value=label)
            c.font = S["f_hdr"]; c.fill = S["BLUE"]
            c.alignment = S["lft"] if col in (2, 3) else S["ctr"]

        ranger = _rang(rows, rank_key)
        for i, (p, rang) in enumerate(zip(rows, ranger)):
            r = i + 3
            ws.row_dimensions[r].height = 17
            if rang == 1:
                bg, f_rank = S["GOLD_BG"], S["f_rank1"]
            elif rang == 2:
                bg, f_rank = S["SILV_BG"], S["f_rank2"]
            elif rang == 3:
                bg, f_rank = S["BRNZ_BG"], S["f_rank3"]
            else:
                bg, f_rank = (S["EVEN"] if i % 2 == 0 else S["ODD"]), S["f_muted"]

            vals = [rang, p["name"], p["lag"],
                    p["goals"] or None, p["assists"] or None,
                    p["gule"] or None, p["rode"] or None]
            for col, (val, (_, _, _)) in enumerate(zip(vals, col_defs), 1):
                c = ws.cell(row=r, column=col, value=val)
                c.fill = bg
                c.border = S["bot"]
                c.alignment = S["lft"] if col in (2, 3) else S["ctr"]
                c.font = f_rank if col == 1 else S["f_data"]

        for col, (_, width, _) in enumerate(col_defs, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    med_statistikk = [p for p in excel if p["goals"] > 0 or p["assists"] > 0]

    scorere = sorted(
        [p for p in med_statistikk if p["goals"] > 0],
        key=lambda x: (-x["goals"], -x["assists"], x["name"]),
    )
    _write_sheet(
        "Toppscorere", scorere,
        f"VM 2026 — Toppscorere   {sum(p['goals'] for p in med_statistikk)} mål totalt",
        rank_key=lambda p: p["goals"],
    )

    assistere = sorted(
        [p for p in med_statistikk if p["assists"] > 0],
        key=lambda x: (-x["assists"], -x["goals"], x["name"]),
    )
    _write_sheet(
        "Assists", assistere,
        f"VM 2026 — Assists   {sum(p['assists'] for p in med_statistikk)} assist totalt",
        rank_key=lambda p: p["assists"],
    )

    try:
        wb.save(EXCEL_PATH)
    except Exception:
        shutil.copy2(backup, EXCEL_PATH)
        raise

    print(f"      → 'Toppscorere'-sheet skrevet ({len(scorere)} spillere)")
    print(f"      → 'Assists'-sheet skrevet ({len(assistere)} spillere)")

# ── Rapport ───────────────────────────────────────────────────────────────────

IKON = {"mål_avvik": "🔴", "assist_avvik": "🟡",
        "gult_avvik": "🟨", "rødt_avvik": "🟥",
        "mangler_i_fifa": "⚠️", "mangler_i_excel": "❌"}

def oppdater_excel(avvik: list[dict]) -> int:
    rettelser = [a for a in avvik if "_sheet" in a and "_row" in a and "_col" in a]
    if not rettelser:
        return 0
    backup = EXCEL_PATH + ".bak"
    shutil.copy2(EXCEL_PATH, backup)
    try:
        wb = load_workbook(EXCEL_PATH)
        for a in rettelser:
            ws = wb[a["_sheet"]]
            ws.cell(row=a["_row"], column=a["_col"], value=a["_ny_verdi"])
        wb.save(EXCEL_PATH)
    except Exception:
        shutil.copy2(backup, EXCEL_PATH)
        raise
    return len(rettelser)

def skriv_rapport(excel: list[dict], fifa: list[dict] | None, avvik: list[dict], status: str) -> None:
    kilde_link = f"[FIFA API]({_FIFA_BASE}/calendar/matches)"
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    L = [
        "# VM 2026 — Avviksrapport",
        f"*Generert: {now}*",
        "",
        f"**Kilde:** {kilde_link}",
        f"**Status:** {status}",
        f"**Spillere i Excel (mål/assist > 0):** {len([p for p in excel if p['goals'] > 0 or p['assists'] > 0])}",
        f"**Spillere hos FIFA:** {len(fifa) if fifa else 'N/A'}",
        "",
    ]
    if fifa is None:
        L += ["## ⚠️ Kunne ikke hente FIFA-data", ""]
    else:
        if avvik:
            L += [f"## 🔍 Avvik — {len(avvik)} stk", ""]
            for a in sorted(avvik, key=lambda x: x["type"]):
                land_str = f" 🌍 {a['land']}" if a.get("land") else ""
                L.append(f"{IKON.get(a['type'],'?')}  **{a['name']}**{land_str} ({a['lag']}, {a['gruppe']}) — {a['detalj']}")
            L.append("")
        else:
            L += ["## ✅ Ingen avvik", "", "Excel stemmer med FIFA.", ""]

        fifa_idx = {fp["key"]: fp for fp in fifa} if fifa else {}
        L += ["## Registrert i Excel", "",
              "| # | Spiller | Land | Lag | Gruppe | Mål | Assist |",
              "|---|---------|------|-----|--------|-----|--------|"]
        for i, p in enumerate(sorted(
                [p for p in excel if p["goals"] > 0 or p["assists"] > 0],
                key=lambda x: (-x["goals"], -x["assists"])), 1):
            fp = fifa_idx.get(p["key"]) or fifa_idx.get(NAME_ALIASES.get(p["key"], ""))
            land = fp.get("land", "") if fp else ""
            L.append(f"| {i} | {p['name']} | {land} | {p['lag']} | {p['gruppe']} | {p['goals']} | {p['assists']} |")

    L += ["", "---", f"*Kilde: {kilde_link}*"]
    Path(REPORT_PATH).write_text("\n".join(L), encoding="utf-8")
    print(f"\n  → Rapport: {REPORT_PATH}")

# ── Hovedprogram ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="VM 2026 avvikssjekk mot FIFA API")
    parser.add_argument("--fix", action="store_true",
                        help="Rett opp avvik direkte i Excel-filen")
    args = parser.parse_args()

    print("=" * 60)
    print("  VM 2026 Avvikssjekk")
    print("  Kilde: FIFA API (offisiell)")
    if args.fix:
        print("  Modus: --fix (avvik skrives til Excel)")
    else:
        print("  Modus: kun rapport (bruk --fix for å rette Excel)")
    print("=" * 60)

    print("\n[1/3] Leser Excel...")
    excel = les_excel()
    excel_med_statistikk = [p for p in excel if p["goals"] > 0 or p["assists"] > 0]
    print(f"      {len(excel)} spillere lest, {len(excel_med_statistikk)} med mål/assist")
    for p in sorted(excel_med_statistikk, key=lambda x: (-x["goals"], -x["assists"])):
        print(f"      {p['name']:<30} mål={p['goals']}  assist={p['assists']}  ({p['lag']})")

    print("\n[2/3] Henter FIFA-statistikk (kan ta 30–60 sek)...")
    fifa, kort, err = hent_fifa()
    if fifa:
        status = f"✅ {len(fifa)} spillere hentet ({err})"
        print(f"      {status}")
        for p in sorted(fifa, key=lambda x: (-x["goals"], -x["assists"]))[:20]:
            print(f"      {p['name']:<30} mål={p['goals']}  assist={p['assists']}  ({p.get('land', '')})")
        print(f"      {len(kort)} spillere med kort")
    else:
        status = f"❌ {err}"
        print(f"      {status}")
        kort = []

    print("\n[3/3] Sammenligner...")
    if fifa:
        avvik = sammenlign(excel, fifa, kort)
        if avvik:
            print(f"      {len(avvik)} avvik:")
            for a in avvik:
                print(f"      {IKON.get(a['type'],'?')} {a['name']} — {a['detalj']}")
        else:
            print("      ✅ Ingen avvik funnet")
    else:
        avvik = []

    skriv_rapport(excel, fifa, avvik, status)

    if args.fix and fifa:
        print("\n[Fix] Oppdaterer Excel...")
        n = oppdater_excel(avvik)
        print(f"      ✅ {n} celler oppdatert (mål/assist/gule/røde kort)" if n else "      Ingen celler å oppdatere")
        if kort:
            skriv_kort_sheet(kort)
        excel_oppdatert = les_excel()
        skriv_toppscore_sheets(excel_oppdatert)

    print("\nFerdig.")

if __name__ == "__main__":
    main()
