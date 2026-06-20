#!/usr/bin/env python3
"""
add_country_sheet.py
Legger til "Spillere etter klubbland"-ark i Excel.
"""
import io, json, shutil, sys
from collections import defaultdict
from pathlib import Path
BASE_DIR = Path(__file__).parent

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    sys.exit(f"Mangler pakke: {e}")

from club_country_map import CLUB_COUNTRY
from top_division import TOP_DIVISION_CLUBS
from level_map import LEVEL_MAP, LAND_NO

EXCEL_PATH  = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
CLUBS_JSON  = BASE_DIR / "clubs_new.json"
FORVENTET   = 1248


def _xl_styles():
    thin = Side(style="thin", color="E2E8F0")
    bot  = Border(bottom=thin)
    ctr  = Alignment(horizontal="center", vertical="center")
    lft  = Alignment(horizontal="left",   vertical="center")
    return {
        "NAVY":    PatternFill("solid", fgColor="0F2044"),
        "BLUE":    PatternFill("solid", fgColor="1A3C6B"),
        "GOLD_BG": PatternFill("solid", fgColor="FFFBE6"),
        "SILV_BG": PatternFill("solid", fgColor="F8F8F8"),
        "BRNZ_BG": PatternFill("solid", fgColor="FFF4EC"),
        "EVEN":    PatternFill("solid", fgColor="F0F5FB"),
        "ODD":     PatternFill("solid", fgColor="FFFFFF"),
        "f_title": Font(name="Calibri", bold=True, size=13, color="FFFFFF"),
        "f_hdr":   Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
        "f_data":  Font(name="Calibri",             size=10, color="1A1A2E"),
        "f_muted": Font(name="Calibri",             size=10, color="6B7A99"),
        "f_rank1": Font(name="Calibri", bold=True,  size=10, color="92680A"),
        "f_rank2": Font(name="Calibri", bold=True,  size=10, color="5C5C5C"),
        "f_rank3": Font(name="Calibri", bold=True,  size=10, color="8B4513"),
        "f_bold":  Font(name="Calibri", bold=True,  size=10, color="1A1A2E"),
        "ctr": ctr, "lft": lft, "bot": bot,
    }


def bygg_land_data():
    with open(CLUBS_JSON, encoding="utf-8") as f:
        data = json.load(f)

    country_total = defaultdict(int)
    country_lvl   = defaultdict(lambda: defaultdict(int))

    for team, players in data.items():
        for player, club in players.items():
            if not club:
                continue
            club = club.strip()
            land_en = CLUB_COUNTRY.get(club, "Unknown")
            land_no = LAND_NO.get(land_en, land_en)
            country_total[land_no] += 1

            if club in TOP_DIVISION_CLUBS:
                country_lvl[land_no][1] += 1
            elif club in LEVEL_MAP:
                lvl = LEVEL_MAP[club][0]
                country_lvl[land_no][lvl] += 1

    ranked = sorted(country_total.items(), key=lambda x: -x[1])
    return ranked, country_lvl


def skriv_country_sheet(ranked, country_lvl):
    S = _xl_styles()

    col_defs = [
        ("#",        6,  S["ctr"]),
        ("Land",    22,  S["lft"]),
        ("Spillere", 9,  S["ctr"]),
        ("Nivå 1",   8,  S["ctr"]),
        ("Nivå 2",   8,  S["ctr"]),
        ("Nivå 3",   8,  S["ctr"]),
        ("Nivå 4",   8,  S["ctr"]),
        ("Nivå 5",   8,  S["ctr"]),
        ("Nivå 6",   8,  S["ctr"]),
    ]
    ncols = len(col_defs)

    backup = Path(str(EXCEL_PATH) + ".bak")
    shutil.copy2(EXCEL_PATH, backup)
    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception:
        raise

    sheet_name = "Spillere etter klubbland"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"

    # Kolonneoverskrifter rad 1
    ws.row_dimensions[1].height = 20
    for col, (label, _, al) in enumerate(col_defs, 1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = S["f_hdr"]; c.fill = S["BLUE"]; c.alignment = al

    # Datarader fra rad 2
    rang = 1
    for i, (land, total) in enumerate(ranked):
        row = i + 2
        ws.row_dimensions[row].height = 17
        lvls = country_lvl.get(land, {})

        if i > 0 and total == ranked[i-1][1]:
            pass
        else:
            rang = i + 1

        if rang == 1:
            bg, f_rang = S["GOLD_BG"], S["f_rank1"]
        elif rang == 2:
            bg, f_rang = S["SILV_BG"], S["f_rank2"]
        elif rang == 3:
            bg, f_rang = S["BRNZ_BG"], S["f_rank3"]
        else:
            bg, f_rang = (S["EVEN"] if i % 2 == 0 else S["ODD"]), S["f_muted"]

        vals = [rang, land, total,
                lvls.get(1, "") or "", lvls.get(2, "") or "",
                lvls.get(3, "") or "", lvls.get(4, "") or "",
                lvls.get(5, "") or "", lvls.get(6, "") or ""]

        for col, (val, (_, _, al)) in enumerate(zip(vals, col_defs), 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = bg
            c.border = S["bot"]
            c.alignment = al
            if col == 1:
                c.font = f_rang
            elif col == 3:
                c.font = S["f_bold"]
            else:
                c.font = S["f_data"]

    for col, (_, width, _) in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    try:
        wb.save(EXCEL_PATH)
    except Exception:
        shutil.copy2(backup, EXCEL_PATH)
        raise

    print(f"  → 'Spillere etter klubbland'-sheet skrevet ({len(ranked)} land)")


def main():
    print("=" * 55)
    print("  VM 2026 — Spillere etter klubbland")
    print("=" * 55)

    print("\n[1/2] Bygger landdata fra clubs_new.json...")
    ranked, country_lvl = bygg_land_data()
    total_spillere = sum(v for _, v in ranked)
    sjekk = "✓" if total_spillere == FORVENTET else f"ADVARSEL: forventet {FORVENTET}"
    print(f"  {len(ranked)} land, {total_spillere} spillere ({sjekk})")

    print("\n[2/2] Skriver Excel-ark...")
    skriv_country_sheet(ranked, country_lvl)

    print("\nFerdig.")


if __name__ == "__main__":
    main()
