"""
Konverterer VM2026_avansert_gruppetabeller_og_sluttspill.xlsx til HTML
for publisering via GitHub Pages.

Kjør: python excel_til_html.py
Utdata: docs/index.html
"""

import os
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.fills import PatternFill

EXCEL_FIL = "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
UTDATA_DIR = "docs"
UTDATA_FIL = os.path.join(UTDATA_DIR, "index.html")

# Ark som skal hoppes over (interne/tomme)
HOPP_OVER_ARK = set()

# Disse arkene vises alltid sist, i denne rekkefølgen
SIST = ["Klubbdominans", "Spillere etter klubbland", "Klubber", "Nivå 2 og lavere", "Aldersfordeling", "Alder"]

# Maks kolonner å vise per ark (None = alle)
MAKS_KOLS = None


def rgb_fra_fill(fill):
    """Returnerer hex-fargekode fra en PatternFill, eller None."""
    if fill is None:
        return None
    if fill.fill_type not in (None, "none", "solid"):
        return None
    if fill.fgColor is None:
        return None
    c = fill.fgColor
    if c.type == "rgb":
        rgb = c.rgb
        if rgb and len(rgb) == 8 and rgb != "00000000" and rgb.upper() != "FFFFFFFF":
            return "#" + rgb[2:]  # fjern alpha-prefix
    return None


def rgb_fra_font_farge(color):
    """Returnerer hex fra font.color, eller None."""
    if color is None:
        return None
    if color.type == "rgb":
        rgb = color.rgb
        if rgb and len(rgb) == 8 and rgb != "00000000":
            return "#" + rgb[2:]
    return None


def escape(tekst):
    if tekst is None:
        return ""
    s = str(tekst)
    s = s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    return s


def formater_verdi(celle):
    """Returnerer celleverdien som string, med tallformatering."""
    v = celle.value
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return f"{v:.2f}".rstrip("0").rstrip(".")
    return str(v)


def bygg_tabell(ws, legg_til_anker=False):
    """Bygger HTML-tabell fra et ark. Med legg_til_anker=True returneres (html, seksjoner)."""
    # Finn brukt område
    min_rad = ws.min_row or 1
    max_rad = ws.max_row or 1
    min_kol = ws.min_column or 1
    max_kol = ws.max_column or 1

    if max_rad < min_rad or max_kol < min_kol:
        if legg_til_anker:
            return "<p><em>Tomt ark</em></p>", []
        return "<p><em>Tomt ark</em></p>"

    seksjon_nr = 0
    seksjoner = []

    # Bygg merged cell-kart: (rad, kol) → (rowspan, colspan) for master-celler
    # og sett med (rad, kol) for celler som skal hoppes over
    merged_info = {}   # (r, c) → {"rowspan": n, "colspan": n}
    skip_celler = set()

    for merge in ws.merged_cells.ranges:
        r1, c1, r2, c2 = merge.min_row, merge.min_col, merge.max_row, merge.max_col
        merged_info[(r1, c1)] = {
            "rowspan": r2 - r1 + 1,
            "colspan": c2 - c1 + 1,
        }
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if not (r == r1 and c == c1):
                    skip_celler.add((r, c))

    # Kolonnebredder (approksimert)
    col_widths = {}
    for i in range(min_kol, max_kol + 1):
        col_letter = get_column_letter(i)
        cd = ws.column_dimensions.get(col_letter)
        if cd and cd.width:
            col_widths[i] = max(30, int(cd.width * 7))
        else:
            col_widths[i] = 80

    html = ['<div class="tabell-wrapper"><table>']

    # Kolgroup for bredder
    html.append("<colgroup>")
    for i in range(min_kol, max_kol + 1):
        html.append(f'<col style="width:{col_widths.get(i, 80)}px">')
    html.append("</colgroup>")

    for r in range(min_rad, max_rad + 1):
        tr_id = ""
        if legg_til_anker:
            første_celle = ws.cell(row=r, column=min_kol)
            bg_første = rgb_fra_fill(første_celle.fill)
            val_første = formater_verdi(første_celle)
            if bg_første and bg_første.lstrip("#").upper() == "0F2044" and val_første:
                seksjon_nr += 1
                anchor_id = f"lagstat-seksjon-{seksjon_nr}"
                raw = val_første.split("—")[-1].strip() if "—" in val_første else val_første.strip()
                tittel = raw.split("   ")[0].strip()
                seksjoner.append((anchor_id, tittel))
                tr_id = f' id="{anchor_id}"'
        html.append(f"<tr{tr_id}>")
        for c in range(min_kol, max_kol + 1):
            if (r, c) in skip_celler:
                continue

            celle = ws.cell(row=r, column=c)
            verdi = formater_verdi(celle)

            stiler = []
            klasser = []

            # Bakgrunnsfarge
            bg = rgb_fra_fill(celle.fill)
            if bg:
                stiler.append(f"background:{bg}")

            # Font
            font = celle.font
            if font:
                if font.bold:
                    stiler.append("font-weight:bold")
                if font.italic:
                    stiler.append("font-style:italic")
                farge = rgb_fra_font_farge(font.color)
                if farge:
                    stiler.append(f"color:{farge}")
                if font.sz and font.sz != 10:
                    stiler.append(f"font-size:{int(font.sz)}pt")

            # Justering
            al = celle.alignment
            if al:
                if al.horizontal in ("center", "general") and verdi and all(
                    ch.isdigit() or ch in ".,%-+" for ch in verdi.replace(" ", "")
                ):
                    stiler.append("text-align:center")
                elif al.horizontal == "right":
                    stiler.append("text-align:right")
                elif al.horizontal == "center":
                    stiler.append("text-align:center")

            # Merged cell-attributter
            merge = merged_info.get((r, c), {})
            attrs = ""
            if merge.get("rowspan", 1) > 1:
                attrs += f' rowspan="{merge["rowspan"]}"'
            if merge.get("colspan", 1) > 1:
                attrs += f' colspan="{merge["colspan"]}"'

            stil_str = ";".join(stiler)
            if stil_str:
                attrs += f' style="{stil_str}"'

            tag = "td"
            html.append(f"<{tag}{attrs}>{escape(verdi)}</{tag}>")

        html.append("</tr>")

    html.append("</table></div>")
    if legg_til_anker:
        return "\n".join(html), seksjoner
    return "\n".join(html)


def generer_lagstat_nav(seksjoner):
    """Genererer en navigasjonsbar med lenker til seksjoner i Lagstatistikk."""
    if not seksjoner:
        return ""
    lenker = "".join(
        f'<a href="#{anchor_id}">{escape(tittel)}</a>'
        for anchor_id, tittel in seksjoner
    )
    return f'<nav class="lagstat-nav">{lenker}</nav>'


# Kolonnene: (wc_rang, fifa_rang, land, gruppe, delta)
# wc_rang   = rangering blant de 48 VM-lagene (live, pr. 26. juni 2026)
# fifa_rang  = live FIFA-verdensrangering (pr. 26. juni 2026, ikke offisielt godkjent)
# delta     = endring fra offisiell rangering 11. juni 2026; positivt = opp
# Kilde: https://inside.fifa.com/fifa-world-ranking/men
FIFA_RANKING = [
    (1,  1,  "Argentina",          "J",  0),
    (2,  2,  "Frankrike",          "I", +1),
    (3,  3,  "Spania",             "G", -1),
    (4,  4,  "England",            "J",  0),
    (5,  5,  "Brasil",             "C", +1),
    (6,  6,  "Marokko",            "C", +1),
    (7,  7,  "Nederland",          "H", +1),
    (8,  8,  "Portugal",           "K", -3),
    (9,  9,  "Mexico",             "A", +5),
    (10, 10, "Belgia",             "H", -1),
    (11, 11, "Colombia",           "I", +2),
    (12, 12, "Tyskland",           "E", -2),
    (13, 13, "Kroatia",            "K", -2),
    (14, 15, "USA",                "D", +2),
    (15, 16, "Sveits",             "B", +3),
    (16, 17, "Japan",              "F", +1),
    (17, 18, "Uruguay",            "G", -2),
    (18, 19, "Senegal",            "I", -4),
    (19, 21, "Iran",               "K", -1),
    (20, 22, "Norge",              "I", +9),
    (21, 23, "Østerrike",          "L", +1),
    (22, 24, "Ecuador",            "E", -1),
    (23, 26, "Egypt",              "H", +3),
    (24, 27, "Tyrkia",             "D", -5),
    (25, 28, "Australia",          "D", -1),
    (26, 29, "Algerie",            "J", -1),
    (27, 30, "Elfenbenskysten",    "E", +3),
    (28, 31, "Sør-Korea",          "A", -6),
    (29, 32, "Canada",             "B", -2),
    (30, 36, "Sverige",            "F", +2),
    (31, 37, "Paraguay",           "D", +4),
    (32, 41, "Skottland",          "C", +1),
    (33, 42, "Panama",             "J", -8),
    (34, 46, "DR Kongo",           "L",  0),
    (35, 48, "Tsjekkia",           "A", -8),
    (36, 54, "Sør-Afrika",         "A", +6),
    (37, 57, "Usbekistan",         "L", -7),
    (38, 58, "Saudi-Arabia",       "G", +3),
    (39, 59, "Tunisia",            "F", -14),
    (40, 60, "Irak",               "K", -3),
    (41, 61, "Qatar",              "B", -5),
    (42, 62, "Bosnia-Hercegovina", "B", +2),
    (43, 64, "Kapp Verde",         "G", +3),
    (44, 65, "Ghana",              "L", +8),
    (45, 72, "Jordan",             "H", -9),
    (46, 82, "Curaçao",            "E",  0),
    (47, 84, "New Zealand",        "F", +1),
    (48, 88, "Haiti",              "C", -5),
]


IKKE_KVALIFISERT_TOPP48 = [
    (12, "Italia"),
    (21, "Danmark"),
    (26, "Nigeria"),
    (32, "Ukraina"),
    (35, "Russland"),
    (36, "Polen"),
    (37, "Wales"),
    (39, "Ungarn"),
    (43, "Serbia"),
    (44, "Kamerun"),
    (47, "Slovakia"),
    (48, "Hellas"),
]


def generer_fifa_ranking_seksjon():
    rader = []
    for i, (wc_rang, fifa_rang, land, gruppe, delta) in enumerate(FIFA_RANKING):
        bg = "#FFFFFF" if i % 2 == 0 else "#F0F5FB"
        if delta > 0:
            delta_html = f'<span style="color:#1a7a3a;font-weight:bold">▲ {delta}</span>'
        elif delta < 0:
            delta_html = f'<span style="color:#c0392b;font-weight:bold">▼ {abs(delta)}</span>'
        else:
            delta_html = '<span style="color:#888">—</span>'
        rader.append(
            f'<tr>'
            f'<td style="background:{bg};font-weight:bold;color:#1A1A2E;text-align:center">{wc_rang}</td>'
            f'<td style="background:{bg};font-weight:bold;color:#1A1A2E;text-align:center">{fifa_rang}</td>'
            f'<td style="background:{bg};color:#1A1A2E">{escape(land)}</td>'
            f'<td style="background:{bg};text-align:center">{delta_html}</td>'
            f'<td style="background:{bg};color:#1A1A2E;text-align:center">{gruppe}</td>'
            f'</tr>'
        )
    rader_html = "\n".join(rader)

    ikke_kval_rader = []
    for i, (fifa_rang, land) in enumerate(IKKE_KVALIFISERT_TOPP48):
        bg = "#FFFFFF" if i % 2 == 0 else "#F0F5FB"
        ikke_kval_rader.append(
            f'<tr>'
            f'<td style="background:{bg};font-weight:bold;color:#1A1A2E;text-align:center">{fifa_rang}</td>'
            f'<td style="background:{bg};color:#1A1A2E">{escape(land)}</td>'
            f'</tr>'
        )
    ikke_kval_html = "\n".join(ikke_kval_rader)

    return (
        '<div id="ark_FIFA_ranking" class="ark-innhold" style="display:none">\n'
        '<h2>FIFA-ranking</h2>\n'
        '<div class="tabell-wrapper"><table>\n'
        '<colgroup>'
        '<col style="width:50px">'
        '<col style="width:90px">'
        '<col style="width:200px">'
        '<col style="width:70px">'
        '<col style="width:80px">'
        '</colgroup>\n'
        '<tr><td colspan="5" style="background:#0F2044;font-weight:bold;color:#FFFFFF;font-size:13pt">'
        'VM 2026 — FIFA-verdensranking (48 lag)</td></tr>\n'
        '<tr>'
        '<td style="background:#1A3C6B;font-weight:bold;color:#FFFFFF;text-align:center">#</td>'
        '<td style="background:#1A3C6B;font-weight:bold;color:#FFFFFF;text-align:center">FIFA-rang</td>'
        '<td style="background:#1A3C6B;font-weight:bold;color:#FFFFFF">Land</td>'
        '<td style="background:#1A3C6B;font-weight:bold;color:#FFFFFF;text-align:center">Endring</td>'
        '<td style="background:#1A3C6B;font-weight:bold;color:#FFFFFF;text-align:center">Gruppe</td>'
        '</tr>\n'
        + rader_html + '\n'
        '<tr><td colspan="5" style="font-style:italic;color:#6B7A99;font-size:9pt">'
        '* FIFA-verdensrangering pr. 11. juni 2026. Endring = live-bevegelse under VM (pr. 26. juni 2026) vs. offisiell rangering 11. juni. '
        'Kilde: inside.fifa.com/fifa-world-ranking/men. # = rangering blant VM-lagene.</td></tr>\n'
        '</table></div>\n'
        '<div class="tabell-wrapper" style="margin-top:18px"><table>\n'
        '<colgroup>'
        '<col style="width:90px">'
        '<col style="width:200px">'
        '</colgroup>\n'
        '<tr><td colspan="2" style="background:#0F2044;font-weight:bold;color:#FFFFFF;font-size:13pt">'
        'Topp 48 på FIFA-rankingen — ikke kvalifisert for VM 2026</td></tr>\n'
        '<tr>'
        '<td style="background:#1A3C6B;font-weight:bold;color:#FFFFFF;text-align:center">FIFA-rang</td>'
        '<td style="background:#1A3C6B;font-weight:bold;color:#FFFFFF">Land</td>'
        '</tr>\n'
        + ikke_kval_html + '\n'
        '<tr><td colspan="2" style="font-style:italic;color:#6B7A99;font-size:9pt">'
        '* Russland er utestengt av FIFA.</td></tr>\n'
        '</table></div>\n'
        '</div>'
    )


def main():
    os.makedirs(UTDATA_DIR, exist_ok=True)

    print(f"Leser {EXCEL_FIL} ...")
    wb = load_workbook(EXCEL_FIL, data_only=True)

    alle = [n for n in wb.sheetnames if n not in HOPP_OVER_ARK]
    sist_sett = set(SIST)
    ark_navn = [n for n in alle if n not in sist_sett] + [n for n in SIST if n in sist_sett and n in alle]
    print(f"  Fant {len(ark_navn)} ark: {', '.join(ark_navn)}")

    # Bygg innhold per ark
    ark_html = {}
    for navn in ark_navn:
        print(f"  Konverterer ark: {navn}")
        ws = wb[navn]
        if navn == "Lagstatistikk":
            tabell_html, seksjoner = bygg_tabell(ws, legg_til_anker=True)
            nav_html = generer_lagstat_nav(seksjoner)
            ark_html[navn] = nav_html + "\n" + tabell_html
        else:
            ark_html[navn] = bygg_tabell(ws)

    # Tab-knapper
    tab_knapper = []
    for i, navn in enumerate(ark_navn):
        aktiv = ' class="aktiv"' if i == 0 else ""
        safe_id = re.sub(r"[^a-zA-Z0-9_-]", "_", navn)
        tab_knapper.append(f'<button{aktiv} onclick="visArk(\'{safe_id}\')">{escape(navn)}</button>')

    # Ark-seksjoner
    ark_seksjoner = []
    for i, navn in enumerate(ark_navn):
        safe_id = re.sub(r"[^a-zA-Z0-9_-]", "_", navn)
        display = "block" if i == 0 else "none"
        ark_seksjoner.append(
            f'<div id="ark_{safe_id}" class="ark-innhold" style="display:{display}">\n'
            f'<h2>{escape(navn)}</h2>\n'
            f'{ark_html[navn]}\n'
            f'</div>'
        )

    tab_knapper.append('<button onclick="visArk(\'FIFA_ranking\')">FIFA-ranking</button>')
    ark_seksjoner.append(generer_fifa_ranking_seksjon())

    tab_html = '\n'.join(tab_knapper)
    ark_html_blokk = '\n'.join(ark_seksjoner)
    tidspunkt = __import__('datetime').datetime.now().strftime("%Y-%m-%d")

    template_fil = os.path.join(UTDATA_DIR, "template.html")
    if not os.path.exists(template_fil):
        print(f"FEIL: {template_fil} mangler — kan ikke generere HTML.", file=__import__('sys').stderr)
        __import__('sys').exit(1)
    with open(template_fil, encoding="utf-8") as f:
        html = f.read()
    print(f"  Bruker template: {template_fil}")

    html = html.replace("<!-- TABS -->", tab_html)
    html = html.replace("<!-- ARK -->", ark_html_blokk)
    html = html.replace("<!-- TIDSPUNKT -->", tidspunkt)

    with open(UTDATA_FIL, "w", encoding="utf-8") as f:
        f.write(html)

    size_kb = os.path.getsize(UTDATA_FIL) // 1024
    print(f"\nFerdig! Skrevet til: {UTDATA_FIL} ({size_kb} KB)")
    print(f"  Ark konvertert: {len(ark_navn)}")


if __name__ == "__main__":
    main()
