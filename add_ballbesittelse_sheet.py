#!/usr/bin/env python3
"""
add_ballbesittelse_sheet.py
Lager arket "Ballbesittelse" med kampstatistikk: besittelse, skudd (totalt,
på mål, utenfor, blokkert) per kamp. Kilde: wc2026_stats3.json.
"""
import io, json, re, shutil, sys, unicodedata
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
BASE_DIR = Path(__file__).parent

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    sys.exit(f"Mangler pakke: {e}")

EXCEL_PATH      = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
STATS_PATH      = BASE_DIR / "wc2026_stats3.json"
CACHE_PATH      = BASE_DIR / "kamper_resultater.json"
SLUTTSPILL_PATH = BASE_DIR / "sluttspill_cache.json"
SHEET_NAME      = "Ballbesittelse"

NORSK = {
    "Mexico": "Mexico", "South Africa": "Sør-Afrika", "Korea Republic": "Sør-Korea",
    "Czechia": "Tsjekkia", "Canada": "Canada", "Bosnia and Herzegovina": "Bosnia-Hercegovina",
    "Qatar": "Qatar", "Switzerland": "Sveits", "Brazil": "Brasil", "Morocco": "Marokko",
    "Haiti": "Haiti", "Scotland": "Skottland", "United States": "USA", "USA": "USA",
    "Paraguay": "Paraguay", "Australia": "Australia", "Turkiye": "Tyrkia", "Turkey": "Tyrkia",
    "Germany": "Tyskland", "Curacao": "Curaçao", "Ivory Coast": "Elfenbenskysten",
    "Ecuador": "Ecuador", "Netherlands": "Nederland", "Japan": "Japan",
    "Sweden": "Sverige", "Tunisia": "Tunisia", "Belgium": "Belgia", "Egypt": "Egypt",
    "IR Iran": "Iran", "Iran": "Iran", "New Zealand": "New Zealand", "Spain": "Spania",
    "Cape Verde": "Kapp Verde", "Saudi Arabia": "Saudi-Arabia", "Uruguay": "Uruguay",
    "France": "Frankrike", "Senegal": "Senegal", "Iraq": "Irak", "Norway": "Norge",
    "Argentina": "Argentina", "Algeria": "Algerie", "Austria": "Østerrike", "Jordan": "Jordan",
    "Portugal": "Portugal", "Congo DR": "DR Kongo", "DR Congo": "DR Kongo",
    "Uzbekistan": "Usbekistan", "Colombia": "Colombia", "England": "England",
    "Croatia": "Kroatia", "Ghana": "Ghana", "Panama": "Panama",
}

def _norm(s):
    s = unicodedata.normalize("NFD", (s or "").lower())
    return re.sub(r"\s+", "", "".join(c for c in s if unicodedata.category(c) != "Mn"))

def til_norsk(name):
    return NORSK.get(name, name)

def _poss_fill_font(pct: int):
    if pct >= 65:
        return PatternFill("solid", fgColor="0F2044"), Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    elif pct >= 55:
        return PatternFill("solid", fgColor="1A3C6B"), Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    elif pct >= 45:
        return PatternFill("solid", fgColor="BDD7EE"), Font(name="Calibri", bold=True, size=10, color="1A1A2E")
    else:
        return PatternFill("solid", fgColor="F0F5FB"), Font(name="Calibri", bold=True, size=10, color="6B7A99")

def _shot_font(val, other, bold=False):
    """Grønn bold hvis høyere enn motstanderen, ellers standard."""
    if val is not None and other is not None and val > other:
        return Font(name="Calibri", bold=True, size=10, color="0F5132")
    return Font(name="Calibri", size=10, color="1A1A2E")


def bygg_data() -> list[dict]:
    with open(STATS_PATH, encoding="utf-8") as f:
        stats = json.load(f)
    with open(CACHE_PATH, encoding="utf-8") as f:
        cache = json.load(f)

    kamp_idx = {}
    for gruppe, kamper in cache.items():
        for k in kamper:
            key = _norm(k["hjemme"]) + "|" + _norm(k["borte"])
            kamp_idx[key] = {"gruppe": gruppe, "hjemme": k["hjemme"], "borte": k["borte"]}

    if SLUTTSPILL_PATH.exists():
        with open(SLUTTSPILL_PATH, encoding="utf-8") as f:
            sluttspill = json.load(f)
        for runde, kamper in sluttspill.items():
            for k in kamper:
                h, b = k.get("hjemme", ""), k.get("borte", "")
                if h and b and h != "TBA" and b != "TBA":
                    key = _norm(h) + "|" + _norm(b)
                    kamp_idx[key] = {"gruppe": runde, "hjemme": h, "borte": b}

    rader = []
    ikke_matchet = []
    for entry in stats.values():
        h_no = til_norsk(entry["home"])
        a_no = til_norsk(entry["away"])
        key  = _norm(h_no) + "|" + _norm(a_no)
        info = kamp_idx.get(key)
        if not info:
            rev_key = _norm(a_no) + "|" + _norm(h_no)
            info = kamp_idx.get(rev_key)
        if not info:
            ikke_matchet.append(f"{h_no} vs {a_no}")
            continue

        # Hent score fra result-feltet i stats-filen
        res = entry.get("result", "")
        try:
            deler = res.split("-")
            score_str = f"{deler[0]}–{deler[1]}"
        except Exception:
            score_str = res

        if entry.get("home_poss") is None:
            continue

        rader.append({
            "gruppe":        info["gruppe"],
            "hjemme":        h_no,
            "borte":         a_no,
            "score":         score_str,
            "h_poss":        entry.get("home_poss"),
            "a_poss":        entry.get("away_poss"),
            "h_shots":       entry.get("home_shots_total"),
            "a_shots":       entry.get("away_shots_total"),
            "h_on":          entry.get("home_shots_on_target"),
            "a_on":          entry.get("away_shots_on_target"),
            "h_off":         entry.get("home_shots_off_target"),
            "a_off":         entry.get("away_shots_off_target"),
            "h_blk":         entry.get("home_shots_blocked"),
            "a_blk":         entry.get("away_shots_blocked"),
        })

    if ikke_matchet:
        print(f"  ADVARSEL: {len(ikke_matchet)} ikke matchet: {', '.join(ikke_matchet)}")

    gruppe_order = {f"Gruppe {x}": i for i, x in enumerate("ABCDEFGHIJKL")}
    sluttspill_order = {"16-delsfinaler": 20, "8-delsfinaler": 21, "Kvartfinaler": 22, "Semifinaler": 23, "Bronsefinale": 24, "Finale": 25}
    gruppe_order.update(sluttspill_order)
    rader.sort(key=lambda r: (gruppe_order.get(r["gruppe"], 99), r["hjemme"]))
    return rader


def skriv_ark(rader: list[dict]) -> None:
    backup = Path(str(EXCEL_PATH) + ".bak")
    shutil.copy2(EXCEL_PATH, backup)
    wb = load_workbook(EXCEL_PATH)

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]

    for after in ("Heatmap", "Scoringstidspunkt", "Lagstatistikk"):
        if after in wb.sheetnames:
            idx = wb.sheetnames.index(after) + 1
            break
    else:
        idx = len(wb.sheetnames)
    ws = wb.create_sheet(SHEET_NAME, idx)

    NAVY  = PatternFill("solid", fgColor="0F2044")
    BLUE  = PatternFill("solid", fgColor="1A3C6B")
    TEAL  = PatternFill("solid", fgColor="0E4D5C")   # mellomfarge for subheader
    EVEN  = PatternFill("solid", fgColor="F0F5FB")
    WHITE = PatternFill("solid", fgColor="FFFFFF")
    thin  = Side(style="thin",   color="E2E8F0")
    thck  = Side(style="medium", color="0F2044")
    bot   = Border(bottom=thin)
    ctr   = Alignment(horizontal="center", vertical="center")
    lft   = Alignment(horizontal="left",   vertical="center")

    f_title  = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    f_hdr    = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    f_sub    = Font(name="Calibri", bold=True, size=9,  color="FFFFFF")
    f_data   = Font(name="Calibri", size=10,            color="1A1A2E")
    f_muted  = Font(name="Calibri", size=10,            color="6B7A99")
    f_score  = Font(name="Calibri", bold=True, size=10, color="0F5132")
    f_grp    = Font(name="Calibri", bold=True, size=9,  color="FFFFFF")

    # Kolonner: A=Gruppe, B=Hjemme, C=Bes, D=Skudd, E=Treff, F=Bom, G=Blk,
    #           H=Score, I=Blk, J=Bom, K=Treff, L=Skudd, M=Bes, N=Borte
    COL_WIDTHS = {"A":10,"B":20,"C":7,"D":7,"E":7,"F":7,"G":6,
                  "H":8,"I":6,"J":7,"K":7,"L":7,"M":7,"N":20}
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # ── Rad 1: Tittel ─────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:N1")
    c = ws.cell(row=1, column=1, value="Kampstatistikk — VM 2026")
    c.font = f_title; c.fill = NAVY; c.alignment = lft

    # ── Rad 2: Subheader (Hjemme / Score / Borte) ─────────────────────────────
    ws.row_dimensions[2].height = 16
    # Hjemme-blokk B–G
    ws.merge_cells("B2:G2")
    c = ws.cell(row=2, column=2, value="← Hjemmelag")
    c.font = f_sub; c.fill = TEAL; c.alignment = ctr
    # Score H
    c = ws.cell(row=2, column=8, value="")
    c.fill = NAVY
    # Borte-blokk I–N
    ws.merge_cells("I2:N2")
    c = ws.cell(row=2, column=9, value="Bortelag →")
    c.font = f_sub; c.fill = TEAL; c.alignment = ctr
    # A2
    c = ws.cell(row=2, column=1); c.fill = NAVY

    # ── Rad 3: Kolonneoverskrifter ────────────────────────────────────────────
    ws.row_dimensions[3].height = 20
    hdrs = [
        (1,  "Gruppe",  ctr, NAVY),
        (2,  "Lag",     lft, BLUE),
        (3,  "Bes.",    ctr, BLUE),
        (4,  "Skudd",   ctr, BLUE),
        (5,  "Treff",   ctr, BLUE),
        (6,  "Bom",     ctr, BLUE),
        (7,  "Blk.",    ctr, BLUE),
        (8,  "Score",   ctr, NAVY),
        (9,  "Blk.",    ctr, BLUE),
        (10, "Bom",     ctr, BLUE),
        (11, "Treff",   ctr, BLUE),
        (12, "Skudd",   ctr, BLUE),
        (13, "Bes.",    ctr, BLUE),
        (14, "Lag",     lft, BLUE),
    ]
    for col, label, al, fill in hdrs:
        c = ws.cell(row=3, column=col, value=label)
        c.font = f_hdr; c.fill = fill; c.alignment = al

    # Tykk kant mellom hjemme-blokk og score, og score og borte
    def _set_border(row, col, left=None, right=None):
        cell = ws.cell(row=row, column=col)
        old = cell.border
        cell.border = Border(
            left=left or old.left, right=right or old.right,
            top=old.top, bottom=old.bottom)

    # ── Datarader ─────────────────────────────────────────────────────────────
    prev_gruppe = None
    for i, r in enumerate(rader):
        row = 4 + i
        ws.row_dimensions[row].height = 18
        bg = EVEN if i % 2 == 0 else WHITE

        grp_val = r["gruppe"] if r["gruppe"] != prev_gruppe else ""
        prev_gruppe = r["gruppe"]

        hp, ap = r["h_poss"] or 0, r["a_poss"] or 0
        h_poss_fill, h_poss_font = _poss_fill_font(hp)
        a_poss_fill, a_poss_font = _poss_fill_font(ap)

        def cell(col, val, font=None, fill=None, al=ctr, num_fmt=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font   = font or f_data
            c.fill   = fill or bg
            c.alignment = al
            c.border = Border(bottom=thin,
                              left=thck if col in (8, 9) else thin,
                              right=thck if col in (7, 8) else thin)
            if num_fmt:
                c.number_format = num_fmt
            return c

        cell(1,  grp_val,          f_grp if grp_val else f_muted, NAVY if grp_val else bg)
        cell(2,  r["hjemme"],      f_data, bg, lft)
        cell(3,  hp / 100 if hp else None, h_poss_font, h_poss_fill, ctr, "0%")
        cell(4,  r["h_shots"],     _shot_font(r["h_shots"], r["a_shots"]))
        cell(5,  r["h_on"],        _shot_font(r["h_on"],    r["a_on"]))
        cell(6,  r["h_off"],       f_data)
        cell(7,  r["h_blk"],       f_muted)
        cell(8,  r["score"],       f_score, bg)
        cell(9,  r["a_blk"],       f_muted)
        cell(10, r["a_off"],       f_data)
        cell(11, r["a_on"],        _shot_font(r["a_on"],    r["h_on"]))
        cell(12, r["a_shots"],     _shot_font(r["a_shots"], r["h_shots"]))
        cell(13, ap / 100 if ap else None, a_poss_font, a_poss_fill, ctr, "0%")
        cell(14, r["borte"],       f_data, bg, lft)

    # ── Forklaring ────────────────────────────────────────────────────────────
    leg_row = 4 + len(rader) + 1
    ws.row_dimensions[leg_row].height = 14
    note = ws.cell(row=leg_row, column=1,
                   value="Treff = skudd på mål  |  Bom = utenfor  |  Blk. = blokkert  |  Grønn bold = flest skudd/treff")
    note.font = Font(name="Calibri", italic=True, size=8, color="6B7A99")
    ws.merge_cells(f"A{leg_row}:N{leg_row}")

    leg_row += 1
    ws.row_dimensions[leg_row].height = 18
    ws.cell(row=leg_row, column=1, value="Besittelse:").font = Font(
        name="Calibri", bold=True, size=9, color="1A1A2E")
    legend = [
        ("≥65% (dominerende)", "0F2044", "FFFFFF"),
        ("55–64%",             "1A3C6B", "FFFFFF"),
        ("45–54% (jevnt)",     "BDD7EE", "1A1A2E"),
        ("≤44%",               "F0F5FB", "6B7A99"),
    ]
    for j, (lbl, bg_hex, fg_hex) in enumerate(legend):
        c = ws.cell(row=leg_row, column=j + 2, value=lbl)
        c.fill = PatternFill("solid", fgColor=bg_hex)
        c.font = Font(name="Calibri", size=8, color=fg_hex)
        c.alignment = ctr
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    try:
        wb.save(EXCEL_PATH)
    except Exception as e:
        shutil.copy2(backup, EXCEL_PATH)
        sys.exit(f"FEIL — rullet tilbake: {e}")


def main():
    print("=" * 55)
    print("  VM 2026 — Kampstatistikk (ballbesittelse + skudd)")
    print("=" * 55)

    if not Path(STATS_PATH).exists():
        sys.exit(f"FEIL: {STATS_PATH} finnes ikke.")

    print("\n[1/2] Matcher stats mot kamper...")
    rader = bygg_data()
    print(f"  {len(rader)} kamper matchet")
    for r in rader:
        print(f"  {r['gruppe']:9s}  {r['hjemme']:22s} {r['h_poss']:2d}%  "
              f"Skudd {r['h_shots']:2d}/{r['a_shots']:2d}  "
              f"Treff {r['h_on']:2d}/{r['a_on']:2d}  "
              f"{r['a_poss']:2d}%  {r['borte']}")

    print(f"\n[2/2] Skriver ark '{SHEET_NAME}'...")
    skriv_ark(rader)
    print("\nFerdig.")


if __name__ == "__main__":
    main()
