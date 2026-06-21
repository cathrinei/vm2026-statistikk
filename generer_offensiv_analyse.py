"""
Genererer docs/offensiv_analyse.html fra Lagstatistikk-arket i Excel.
Kjøres av daglig_oppdatering.yml etter excel_til_html.py.
"""

import os
import json
from datetime import datetime, timezone
from openpyxl import load_workbook

EXCEL_FIL = "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
UTDATA    = os.path.join("docs", "offensiv_analyse.html")


def les_lagstatistikk():
    wb = load_workbook(EXCEL_FIL, data_only=True)
    ws = wb["Lagstatistikk"]
    rows = [row for row in ws.iter_rows(values_only=True)]

    maal  = {}   # lag → {kamper, maal_for, maal_per_kamp}
    skudd = {}   # lag → {skudd, skudd_per_kamp, paa_maal, treff_pct}

    # --- Mål per kamp-tabell (starter på rad 3, header på rad 2) ---
    i = 0
    while i < len(rows):
        r = rows[i]
        if r[0] == "#" and r[1] == "Lag" and r[4] == "Mål for":
            i += 1
            while i < len(rows) and isinstance(rows[i][0], int):
                r = rows[i]
                lag     = r[1]
                kamper  = r[3]
                maal_for = r[4]
                # MF/kamp er lagret som streng med komma, beregn fra tall
                mpc = round(maal_for / kamper, 2) if kamper else 0
                maal[lag] = {"kamper": kamper, "maal_for": maal_for, "maal_per_kamp": mpc}
                i += 1
            break
        i += 1

    # --- Skudd per lag-tabell ---
    i = 0
    while i < len(rows):
        r = rows[i]
        if r[0] == "#" and r[1] == "Lag" and r[4] == "Skudd":
            i += 1
            while i < len(rows) and isinstance(rows[i][0], int):
                r = rows[i]
                lag  = r[1]
                sk   = r[4]
                spk  = round(sk / r[3], 1) if r[3] else 0
                pm   = r[6]
                pct  = int(r[7].rstrip("%")) if isinstance(r[7], str) else int(r[7] * 100)
                skudd[lag] = {"skudd": sk, "skudd_per_kamp": spk, "paa_maal": pm, "treff_pct": pct}
                i += 1
            break
        i += 1

    return maal, skudd


def bygg_data(maal, skudd):
    # Finn laget med flest skudd blant 0-mål-lag → markeres rødt
    nullmaal = {lag: skudd[lag]["skudd"] for lag, m in maal.items()
                if m["maal_for"] == 0 and lag in skudd}
    rodt_lag = max(nullmaal, key=nullmaal.get) if nullmaal else None

    lag_data = []
    for lag, m in maal.items():
        if lag not in skudd:
            continue
        s = skudd[lag]
        kamper    = m["kamper"]
        maal_for  = m["maal_for"]
        mpc       = m["maal_per_kamp"]
        spk       = s["skudd_per_kamp"]
        treff_pct = s["treff_pct"]
        effektivitet = round(maal_for / s["skudd"] * 100) if s["skudd"] else 0

        if lag == rodt_lag:
            farge = "red"
        elif maal_for == 0:
            farge = "gray"
        elif kamper == 1:
            farge = "green"
        else:
            farge = "blue"

        lag_data.append({
            "label":     lag,
            "x":         spk,
            "y":         mpc,
            "r":         treff_pct,
            "type":      farge,
            "kamper":    kamper,
            "maal":      maal_for,
            "skudd":     s["skudd"],
            "effektivitet": effektivitet,
        })
    return lag_data


def finn_metrikker(lag_data, maal, skudd):
    med_maal = [d for d in lag_data if d["maal"] > 0]

    best_mpc   = max(lag_data, key=lambda d: d["y"])
    best_treff = max(lag_data, key=lambda d: d["r"])
    nullmaal_skudd = max(
        (d for d in lag_data if d["maal"] == 0),
        key=lambda d: d["skudd"], default=None
    )
    best_eff = max(med_maal, key=lambda d: d["effektivitet"]) if med_maal else None

    return {
        "best_mpc":          best_mpc,
        "best_treff":        best_treff,
        "nullmaal_skudd":    nullmaal_skudd,
        "best_effektivitet": best_eff,
    }


TOPP_SCORERE = 12   # antall beste scorende lag i diagrammet
TOPP_NULLMAAL = 3   # antall verste 0-mål-lag i diagrammet


def velg_chart_lag(lag_data):
    """Velg alltid et fast antall lag uansett turneringsstadium."""
    med_maal  = sorted([d for d in lag_data if d["maal"] > 0],
                       key=lambda d: (-d["y"], -d["skudd"]))[:TOPP_SCORERE]
    nullmaal  = sorted([d for d in lag_data if d["maal"] == 0],
                       key=lambda d: -d["skudd"])[:TOPP_NULLMAAL]
    return med_maal + nullmaal


def generer_html(lag_data, metrikker, tidspunkt):
    chart_data = velg_chart_lag(lag_data)
    to_games = [d for d in chart_data if d["kamper"] >= 2]
    one_game = [d for d in chart_data if d["kamper"] == 1]

    def js_arr(data):
        items = []
        for d in data:
            items.append(
                f"  {{label:{json.dumps(d['label'])},x:{d['x']},y:{d['y']},r:{d['r']},type:{json.dumps(d['type'])}}}"
            )
        return "[\n" + ",\n".join(items) + "\n]"

    m = metrikker
    mpc_val  = f"{m['best_mpc']['y']:.1f}".replace(".", ",")
    mpc_lag  = m["best_mpc"]["label"]
    mpc_maal = m["best_mpc"]["maal"]
    mpc_kamp = m["best_mpc"]["kamper"]

    treff_val = m["best_treff"]["r"]
    treff_lag = m["best_treff"]["label"]
    treff_pm  = m["best_treff"]["r"] * m["best_treff"]["skudd"] // 100
    treff_sk  = m["best_treff"]["skudd"]

    null_lag   = m["nullmaal_skudd"]["label"] if m["nullmaal_skudd"] else "–"
    null_sk    = m["nullmaal_skudd"]["skudd"] if m["nullmaal_skudd"] else 0
    null_spk_f = m["nullmaal_skudd"]["x"] if m["nullmaal_skudd"] else 0
    null_spk   = int(null_spk_f) if null_spk_f == int(null_spk_f) else null_spk_f

    eff_val = m["best_effektivitet"]["effektivitet"] if m["best_effektivitet"] else 0
    eff_lag = m["best_effektivitet"]["label"] if m["best_effektivitet"] else "–"

    html = f"""<!DOCTYPE html>
<html lang="no">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>VM 2026 — Offensiv analyse</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

  :root {{
    --bg: #ffffff;
    --bg-secondary: #f5f4f0;
    --text-primary: #2c2c2a;
    --text-secondary: #5f5e5a;
    --border: rgba(0,0,0,0.12);
    --radius: 8px;
    --radius-lg: 12px;
  }}

  @media (prefers-color-scheme: dark) {{
    :root {{
      --bg: #1e1e1c;
      --bg-secondary: #2a2a28;
      --text-primary: #e8e7e2;
      --text-secondary: #a0a09a;
      --border: rgba(255,255,255,0.10);
    }}
  }}

  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    background: var(--bg);
    color: var(--text-primary);
    padding: 32px 24px;
    max-width: 820px;
    margin: 0 auto;
    line-height: 1.5;
  }}

  header {{ margin-bottom: 28px; }}
  header h1 {{ font-size: 20px; font-weight: 500; margin-bottom: 4px; }}
  header p {{ font-size: 13px; color: var(--text-secondary); }}

  .metrics {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
    gap: 10px;
    margin-bottom: 24px;
  }}

  .metric-card {{
    background: var(--bg-secondary);
    border-radius: var(--radius);
    padding: 14px 16px;
    border: 0.5px solid var(--border);
  }}
  .metric-label {{ font-size: 11px; color: var(--text-secondary); text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 6px; }}
  .metric-value {{ font-size: 26px; font-weight: 500; line-height: 1; margin-bottom: 4px; }}
  .metric-sub   {{ font-size: 12px; color: var(--text-secondary); }}

  .metric-card.blue .metric-value  {{ color: #185FA5; }}
  .metric-card.green .metric-value {{ color: #0F6E56; }}
  .metric-card.red .metric-value   {{ color: #a32d2d; }}
  .metric-card.amber .metric-value {{ color: #854F0B; }}

  @media (prefers-color-scheme: dark) {{
    .metric-card.blue .metric-value  {{ color: #85B7EB; }}
    .metric-card.green .metric-value {{ color: #5DCAA5; }}
    .metric-card.red .metric-value   {{ color: #F09595; }}
    .metric-card.amber .metric-value {{ color: #FAC775; }}
  }}

  .chart-section {{
    background: var(--bg);
    border: 0.5px solid var(--border);
    border-radius: var(--radius-lg);
    padding: 20px;
  }}
  .chart-header {{ font-size: 13px; color: var(--text-secondary); margin-bottom: 12px; }}
  .chart-wrap {{ position: relative; width: 100%; height: 420px; }}

  @media (max-width: 480px) {{
    body {{ padding: 20px 16px; }}
    header h1 {{ font-size: 17px; }}
    .metric-card {{ padding: 10px 12px; }}
    .metric-value {{ font-size: 22px; }}
    .chart-section {{ padding: 14px; }}
    .chart-wrap {{ height: 320px; }}
  }}

  .legend {{ display: flex; gap: 20px; margin-top: 12px; flex-wrap: wrap; font-size: 12px; color: var(--text-secondary); }}
  .legend-item {{ display: flex; align-items: center; gap: 6px; }}
  .legend-dot {{ width: 13px; height: 13px; border-radius: 50%; border-width: 1.5px; border-style: solid; display: inline-block; flex-shrink: 0; }}
</style>
</head>
<body>

<header>
  <a href="index.html" style="display:inline-flex;align-items:center;gap:5px;font-size:12px;color:var(--text-secondary);text-decoration:none;margin-bottom:10px;padding:4px 0;">← Tilbake til oversikten</a>
  <h1>VM 2026 — Offensiv dominans og effektivitet</h1>
  <p>Basert på gruppespilldata per {tidspunkt} · Boblestørrelse = treff% (skudd på mål)</p>
</header>

<div class="metrics">
  <div class="metric-card blue">
    <div class="metric-label">Mest mål per kamp</div>
    <div class="metric-value">{mpc_val}</div>
    <div class="metric-sub">{mpc_lag} · {mpc_maal} mål på {mpc_kamp} kamper</div>
  </div>
  <div class="metric-card green">
    <div class="metric-label">Beste treff%</div>
    <div class="metric-value">{treff_val}%</div>
    <div class="metric-sub">{treff_lag} · {treff_pm} av {treff_sk} skudd på mål</div>
  </div>
  <div class="metric-card red">
    <div class="metric-label">Skudd uten mål</div>
    <div class="metric-value">{null_sk}</div>
    <div class="metric-sub">{null_lag} · {null_spk} skudd/kamp · 0 mål</div>
  </div>
  <div class="metric-card amber">
    <div class="metric-label">Mest effektiv</div>
    <div class="metric-value">{eff_val}%</div>
    <div class="metric-sub">{eff_lag} · mål per skudd</div>
  </div>
</div>

<div class="chart-section">
  <div class="chart-header">
    <span id="chart-hint"></span> Større boble = høyere treffprosent.
  </div>
  <div class="chart-wrap">
    <canvas id="bubbleChart" role="img"
      aria-label="Boblediagram: skudd per kamp (x-akse) mot mål per kamp (y-akse). Boblestørrelse viser treffprosent.">
    </canvas>
  </div>
  <div class="legend">
    <div class="legend-item">
      <span class="legend-dot" style="background:rgba(50,102,173,0.65);border-color:#185FA5"></span>
      2 kamper spilt
    </div>
    <div class="legend-item">
      <span class="legend-dot" style="background:rgba(29,158,117,0.45);border-color:#0F6E56"></span>
      1 kamp spilt
    </div>
    <div class="legend-item">
      <span class="legend-dot" style="background:rgba(194,57,43,0.7);border-color:#a32d2d"></span>
      0 mål — flest skudd
    </div>
    <div class="legend-item">
      <span class="legend-dot" style="background:rgba(136,135,128,0.55);border-color:#5F5E5A"></span>
      0 mål — øvrige
    </div>
  </div>
</div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
const isDark   = matchMedia('(prefers-color-scheme: dark)').matches;
const isMobile = window.innerWidth <= 480;
const textColor = isDark ? '#a0a09a' : '#5f5e5a';
const gridColor = isDark ? 'rgba(255,255,255,0.07)' : 'rgba(0,0,0,0.07)';
document.getElementById('chart-hint').textContent =
  isMobile ? 'Trykk på en boble for detaljer.' : 'Hover over en boble for detaljer.';

const twoGames = {js_arr(to_games)};
const oneGame  = {js_arr(one_game)};

function bgColor(d) {{
  if (d.type === 'red')   return 'rgba(194,57,43,0.70)';
  if (d.type === 'gray')  return 'rgba(136,135,128,0.55)';
  if (d.type === 'green') return 'rgba(29,158,117,0.45)';
  return 'rgba(50,102,173,0.65)';
}}
function bdColor(d) {{
  if (d.type === 'red')   return '#a32d2d';
  if (d.type === 'gray')  return '#5F5E5A';
  if (d.type === 'green') return '#0F6E56';
  return '#185FA5';
}}
function toPoint(d) {{
  return {{ x: d.x, y: d.y, r: Math.sqrt(d.r) * 2.7, _label: d.label, _treff: d.r }};
}}

new Chart(document.getElementById('bubbleChart'), {{
  type: 'bubble',
  data: {{
    datasets: [
      {{
        label: '2 kamper spilt',
        data: twoGames.map(toPoint),
        backgroundColor: twoGames.map(bgColor),
        borderColor:     twoGames.map(bdColor),
        borderWidth: 1.5,
      }},
      {{
        label: '1 kamp spilt',
        data: oneGame.map(toPoint),
        backgroundColor: oneGame.map(bgColor),
        borderColor:     oneGame.map(bdColor),
        borderWidth: 1.5,
      }}
    ]
  }},
  options: {{
    responsive: true,
    maintainAspectRatio: false,
    layout: {{ padding: {{ top: 20, right: 20 }} }},
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{
        callbacks: {{
          label: ctx => {{
            const d = ctx.raw;
            return ` ${{d._label}}: ${{d.y}} mål/kamp · ${{d.x}} skudd/kamp · ${{d._treff}}% treff`;
          }}
        }}
      }}
    }},
    scales: {{
      x: {{
        title: {{ display: true, text: 'Skudd per kamp', color: textColor, font: {{ size: 12 }} }},
        min: 0,
        grid: {{ color: gridColor }},
        ticks: {{ color: textColor }}
      }},
      y: {{
        title: {{ display: true, text: 'Mål per kamp', color: textColor, font: {{ size: 12 }} }},
        min: -0.4,
        grid: {{ color: gridColor }},
        ticks: {{ color: textColor, stepSize: 1 }}
      }}
    }}
  }},
  plugins: [{{
    afterDraw(chart) {{
      const ctx = chart.ctx;
      const keyLabels = new Set(['Tyskland','Tyrkia','Nederland','Frankrike','Norge','Canada','Ecuador']);
      chart.data.datasets.forEach((ds, di) => {{
        ds.data.forEach((d, i) => {{
          const meta = chart.getDatasetMeta(di);
          if (!meta.data[i]) return;
          if (isMobile && !keyLabels.has(d._label)) return;
          const {{ x, y }} = meta.data[i];
          ctx.save();
          ctx.fillStyle = isDark ? '#c8c7c2' : '#2c2c2a';
          ctx.font = `500 ${{isMobile ? 10 : 11}}px -apple-system, sans-serif`;
          ctx.textAlign = 'center';
          ctx.fillText(d._label, x, y - d.r - 4);
          ctx.restore();
        }});
      }});
    }}
  }}]
}});
</script>
</body>
</html>"""
    return html


def main():
    print("Leser Lagstatistikk fra Excel...")
    maal, skudd = les_lagstatistikk()
    print(f"  Lag med mål-data:  {len(maal)}")
    print(f"  Lag med skudd-data: {len(skudd)}")

    lag_data  = bygg_data(maal, skudd)
    metrikker = finn_metrikker(lag_data, maal, skudd)
    tidspunkt = datetime.now(timezone.utc).strftime("%d.%m.%Y")

    html = generer_html(lag_data, metrikker, tidspunkt)

    os.makedirs("docs", exist_ok=True)
    with open(UTDATA, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\nFerdig! Skrevet til: {UTDATA}")
    print(f"  Beste mål/kamp:    {metrikker['best_mpc']['label']} ({metrikker['best_mpc']['y']})")
    print(f"  Beste treff%:      {metrikker['best_treff']['label']} ({metrikker['best_treff']['r']}%)")
    if metrikker["nullmaal_skudd"]:
        print(f"  Flest skudd/0 mål: {metrikker['nullmaal_skudd']['label']} ({metrikker['nullmaal_skudd']['skudd']} skudd)")
    if metrikker["best_effektivitet"]:
        print(f"  Mest effektiv:     {metrikker['best_effektivitet']['label']} ({metrikker['best_effektivitet']['effektivitet']}%)")


if __name__ == "__main__":
    main()
