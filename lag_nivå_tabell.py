"""
Genererer Excel-ark: spillere som ikke spiller i toppserien (nivå 2 og lavere).
Kolonner: Nivå, Klubb, Liga, Spiller, Nasjonallag
"""
import re, sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from level_map import LEVEL_MAP, LAND_NO

NAT_NO = {
    "Algeria": "Algerie", "Australia": "Australia", "Austria": "Østerrike",
    "Bosnia and Herzegovina": "Bosnia-Hercegovina", "Brazil": "Brasil",
    "Cabo Verde": "Kapp Verde", "Canada": "Canada", "Croatia": "Kroatia",
    "Curaçao": "Curaçao", "Democratic Republic of the Congo": "DR Kongo",
    "Ecuador": "Ecuador", "Egypt": "Egypt", "Germany": "Tyskland",
    "Ghana": "Ghana", "Haiti": "Haiti", "Iraq": "Irak",
    "Islamic Republic of Iran": "Iran", "Japan": "Japan",
    "Morocco": "Marokko", "Netherlands": "Nederland", "New Zealand": "New Zealand",
    "Norway": "Norge", "Panama": "Panama", "Paraguay": "Paraguay",
    "Qatar": "Qatar", "Republic of Korea": "Sør-Korea", "Scotland": "Skottland",
    "Senegal": "Senegal", "South Africa": "Sør-Afrika", "Sweden": "Sverige",
    "Switzerland": "Sveits", "Tunisia": "Tunisia",
    "United States of America": "USA", "Uruguay": "Uruguay",
}

# ── Les not_top_div.txt ───────────────────────────────────────────────────────

rows = []
with open('not_top_div.txt', encoding='utf-8') as f:
    content = f.read()

club_block_re = re.compile(r'^\[([^\]]+)\] (.+)$', re.MULTILINE)
player_re     = re.compile(r'^\s{4}(.+?):\s+(.+)$', re.MULTILINE)

for m_club in club_block_re.finditer(content):
    club = m_club.group(2).strip()
    if club not in LEVEL_MAP:
        continue
    niva, liga, land_en = LEVEL_MAP[club]
    land = LAND_NO.get(land_en, land_en)

    start = m_club.end()
    next_club = club_block_re.search(content, start)
    end = next_club.start() if next_club else len(content)
    block = content[start:end]

    for m_p in player_re.finditer(block):
        nasjonallag = NAT_NO.get(m_p.group(1).strip(), m_p.group(1).strip())
        spiller     = m_p.group(2).strip()
        rows.append((niva, club, land, liga, spiller, nasjonallag))

rows.sort(key=lambda x: (x[0], x[2], x[3], x[1], x[4], x[5]))

# ── Skriv Excel ───────────────────────────────────────────────────────────────

EXCEL_PATH = Path(__file__).parent / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
wb = load_workbook(EXCEL_PATH)

sheet_name = 'Nivå 2 og lavere'
if sheet_name in wb.sheetnames:
    del wb[sheet_name]
ws = wb.create_sheet(sheet_name)
ws.auto_filter.ref = "A2:F2"

thin  = Side(style='thin', color='E2E8F0')
brd   = Border(bottom=thin)
brd_hdr = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
                 top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
ctr   = Alignment(horizontal='center', vertical='center')
lft   = Alignment(horizontal='left', vertical='center')

# Tittelrad
ws.row_dimensions[1].height = 28
ws.merge_cells('A1:F1')
c = ws.cell(row=1, column=1, value=f'VM 2026 — Spillere på nivå 2 og lavere   {len(rows)} spillere')
c.font      = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
c.fill      = PatternFill('solid', fgColor='0F2044')
c.alignment = lft

headers = ['Nivå', 'Klubb', 'Land', 'Liga', 'Spiller', 'Nasjonallag']
FILLS = {
    'header': PatternFill('solid', start_color='1A3C6B'),
    2:        PatternFill('solid', start_color='EBF5EB'),
    3:        PatternFill('solid', start_color='FFF9E6'),
    4:        PatternFill('solid', start_color='FDE9D9'),
    5:        PatternFill('solid', start_color='F2DCE8'),
    6:        PatternFill('solid', start_color='E8E0F0'),
}
ws.row_dimensions[2].height = 20
for col, h in enumerate(headers, 1):
    c = ws.cell(row=2, column=col, value=h)
    c.font      = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    c.fill      = FILLS['header']
    c.alignment = ctr

for i, (niva, club, land, liga, spiller, nasjonallag) in enumerate(rows, 1):
    row = i + 2
    ws.row_dimensions[row].height = 17
    fill = FILLS.get(niva, PatternFill('solid', start_color='FFFFFF'))
    for col, (val, al) in enumerate(zip([niva, club, land, liga, spiller, nasjonallag],
                                         [ctr, lft, lft, lft, lft, lft]), 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name='Calibri', size=10, color='1A1A2E')
        c.fill      = fill
        c.alignment = al
        c.border    = brd

for col, width in zip('ABCDEF', [8, 26, 16, 22, 30, 22]):
    ws.column_dimensions[col].width = width

legend_row = len(rows) + 4
ws.cell(row=legend_row, column=1, value='Fargeforklaring:').font = Font(name='Calibri', bold=True, size=10, color='1A1A2E')

for offset, (niva_l, label) in enumerate([(2,'Nivå 2'), (3,'Nivå 3'), (4,'Nivå 4'), (5,'Nivå 5'), (6,'Nivå 6')], 1):
    c = ws.cell(row=legend_row + offset, column=1, value=label)
    c.fill = FILLS[niva_l]
    c.font = Font(name='Calibri', size=10, color='1A1A2E')

wb.save(EXCEL_PATH)

from collections import Counter
niva_teller   = Counter(r[0] for r in rows)
land_teller   = Counter(r[2] for r in rows)
klubb_teller  = Counter(r[1] for r in rows)

print(f'\nFerdig — {len(rows)} spillere skrevet til "{sheet_name}"')
print()
print('Fordeling per nivå:')
for niva in sorted(niva_teller):
    print(f'  Nivå {niva}: {niva_teller[niva]:>3} spillere')
print()
print('Fordeling per klubbland (topp 10):')
for land, ant in land_teller.most_common(10):
    print(f'  {land:<22} {ant:>3}')
print()
print('Klubber med flest spillere (nivå 2 og lavere):')
for klubb, ant in klubb_teller.most_common(10):
    print(f'  {klubb:<30} {ant:>3}')
