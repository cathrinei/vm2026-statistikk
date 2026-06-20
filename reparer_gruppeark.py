#!/usr/bin/env python3
"""
reparer_gruppeark.py
Erstatter gruppearkene (Gruppe A-L) i hoved-Excel med versjonen fra FULL_200626.xlsx.
Alle andre ark beholdes uendret.
"""
import io, shutil, sys
from copy import copy
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Mangler openpyxl: pip install openpyxl")

BASE_DIR  = Path(__file__).parent
MAIN_PATH = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill.xlsx"
FULL_PATH = BASE_DIR / "VM2026_avansert_gruppetabeller_og_sluttspill_FULL_200626.xlsx"
GROUPS    = [f"Gruppe {x}" for x in "ABCDEFGHIJKL"]


def kopier_ark(ws_src, ws_dst):
    # Merget celler
    for merge in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merge))

    # Kolonnebredder
    for col_ltr, col_dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[col_ltr].width = col_dim.width
        ws_dst.column_dimensions[col_ltr].hidden = col_dim.hidden

    # Radhøyder
    for row_num, row_dim in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[row_num].height = row_dim.height

    # Celler
    for row in ws_src.iter_rows():
        for cell in row:
            dst = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                dst.font          = copy(cell.font)
                dst.fill          = copy(cell.fill)
                dst.alignment     = copy(cell.alignment)
                dst.border        = copy(cell.border)
                dst.number_format = cell.number_format


backup = Path(str(MAIN_PATH) + ".bak")
shutil.copy2(MAIN_PATH, backup)
print(f"Backup: {backup}\n")

print(f"Leser {FULL_PATH.name} ...")
wb_full = load_workbook(FULL_PATH)
print(f"Leser {MAIN_PATH.name} ...")
wb_main = load_workbook(MAIN_PATH)

for gruppe in GROUPS:
    if gruppe not in wb_full.sheetnames:
        print(f"  {gruppe}: MANGLER i FULL — hopper over")
        continue

    ws_src = wb_full[gruppe]

    # Bevar posisjon i arkrekkefølgen
    if gruppe in wb_main.sheetnames:
        pos = wb_main.sheetnames.index(gruppe)
        del wb_main[gruppe]
    else:
        pos = 0

    ws_dst = wb_main.create_sheet(gruppe, pos)
    kopier_ark(ws_src, ws_dst)
    print(f"  {gruppe}: OK")

wb_main.save(MAIN_PATH)
print(f"\nLagret: {MAIN_PATH}")
print("\nNeste steg:")
print("  python add_minutter_column.py   (gjenopprett Min-kolonne)")
print("  python check_avvik.py --fix     (gjenopprett Gule/Røde kort)")
